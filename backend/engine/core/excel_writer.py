"""
DocAgent — Excel Writer (Upgraded)
Features:
  - Combined mode: all docs in one sheet per type (default)
  - Per-file mode: each document gets its own sheet
  - Fallback column derivation when schema not available
  - Append mode for adding to existing workbooks
  - Professional formatting with confidence highlighting
"""

from pathlib import Path
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2B5797")
SUBHEADER_FILL = PatternFill("solid", fgColor="4472C4")
DATA_FONT = Font(name="Arial", size=10)
LOW_CONF_FILL = PatternFill("solid", fgColor="FFF2CC")
MEDIUM_CONF_FILL = PatternFill("solid", fgColor="FFF8E1")
ERROR_FILL = PatternFill("solid", fgColor="FCE4EC")
SUCCESS_FILL = PatternFill("solid", fgColor="E8F5E9")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


class ExcelWriter:
    """Generates Excel output from extraction results."""

    def __init__(self, existing_path: str | Path = None, per_file_sheets: bool = False):
        """
        Args:
            existing_path: Load existing workbook to append to
            per_file_sheets: If True, each document gets its own sheet
        """
        if existing_path and Path(existing_path).exists():
            self.wb = load_workbook(str(existing_path))
        else:
            self.wb = Workbook()
            self.wb.remove(self.wb.active)
        self._sheets: dict[str, dict] = {}
        self._per_file = per_file_sheets
        self._file_count = 0

    def add_extraction_result(
        self,
        doc_type: str,
        filename: str,
        extracted_data: dict,
        validation_result=None,
        schema_fields: list[dict] = None,
        line_items_schema: list[dict] = None,
    ):
        """Add one document's extraction results."""
        if not schema_fields:
            schema_fields = self._derive_fields_from_data(extracted_data)

        if self._per_file:
            self._add_per_file_sheet(filename, doc_type, extracted_data, validation_result, schema_fields, line_items_schema)
        else:
            self._add_combined_row(doc_type, filename, extracted_data, validation_result, schema_fields, line_items_schema)

    def _add_combined_row(self, doc_type, filename, extracted_data, validation_result, schema_fields, line_items_schema):
        """Add a row to the combined sheet for this document type."""
        if doc_type not in self._sheets:
            self._create_combined_sheet(doc_type, schema_fields, line_items_schema)

        info = self._sheets[doc_type]
        ws = info["worksheet"]
        row = info["next_row"]
        field_names = info["field_names"]

        self._write_cell(ws, row, 1, filename)

        confidence = extracted_data.get("overall_confidence", "unknown")
        cc = self._write_cell(ws, row, 2, confidence)
        if confidence == "high":
            cc.fill = SUCCESS_FILL
        elif confidence == "low":
            cc.fill = LOW_CONF_FILL

        needs_review = "Yes" if (validation_result and hasattr(validation_result, 'needs_review') and validation_result.needs_review) else "No"
        rc = self._write_cell(ws, row, 3, needs_review)
        if needs_review == "Yes":
            rc.fill = LOW_CONF_FILL

        ext_data = extracted_data.get("extracted_data", {})
        for i, fname in enumerate(field_names):
            col = 4 + i
            fd = ext_data.get(fname, {})
            if isinstance(fd, dict):
                value = fd.get("value")
                fc = fd.get("confidence", "high")
            else:
                value = fd
                fc = "high"
            cell = self._write_cell(ws, row, col, value)
            if fc == "low":
                cell.fill = LOW_CONF_FILL
            elif fc == "medium":
                cell.fill = MEDIUM_CONF_FILL

        notes_col = 4 + len(field_names)
        meta = extracted_data.get("metadata", {})
        notes = meta.get("extraction_notes", "")
        if validation_result and hasattr(validation_result, 'errors') and validation_result.errors:
            notes += " | ERRORS: " + "; ".join(validation_result.errors)
        self._write_cell(ws, row, notes_col, notes)

        info["next_row"] = row + 1

        line_items = extracted_data.get("line_items", [])
        if line_items:
            li_schema = line_items_schema or self._derive_line_item_fields(line_items)
            if li_schema:
                self._write_line_items_combined(doc_type, filename, line_items, li_schema)

    def _add_per_file_sheet(self, filename, doc_type, extracted_data, validation_result, schema_fields, line_items_schema):
        """Create a dedicated sheet for this single file."""
        self._file_count += 1
        clean_name = Path(filename).stem[:25]
        sheet_title = f"{self._file_count}_{clean_name}"[:31]

        # Ensure unique
        while sheet_title in self.wb.sheetnames:
            sheet_title = f"{self._file_count}_{clean_name[:20]}_{self._file_count}"[:31]

        ws = self.wb.create_sheet(title=sheet_title)

        # Title row
        title_font = Font(name="Arial", bold=True, size=12, color="2B5797")
        ws.cell(row=1, column=1, value=f"Extraction: {filename}").font = title_font
        ws.cell(row=2, column=1, value=f"Type: {doc_type}").font = Font(name="Arial", size=10, color="666666")
        ws.cell(row=2, column=3, value=f"Confidence: {extracted_data.get('overall_confidence', 'unknown')}").font = Font(name="Arial", size=10)

        # Header fields section
        row = 4
        ws.cell(row=row, column=1, value="Field").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value="Value").font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value="Confidence").font = HEADER_FONT
        ws.cell(row=row, column=3).fill = HEADER_FILL
        ws.cell(row=row, column=3).border = THIN_BORDER

        ext_data = extracted_data.get("extracted_data", {})
        row = 5
        for field in schema_fields:
            fname = field["name"]
            fd = ext_data.get(fname, {})
            if isinstance(fd, dict):
                val = fd.get("value")
                conf = fd.get("confidence", "high")
            else:
                val = fd
                conf = "high"

            label = fname.replace("_", " ").title()
            if field.get("required"):
                label += " *"

            self._write_cell(ws, row, 1, label)
            vc = self._write_cell(ws, row, 2, val)
            cc = self._write_cell(ws, row, 3, conf)

            if conf == "low":
                vc.fill = LOW_CONF_FILL
            elif conf == "medium":
                vc.fill = MEDIUM_CONF_FILL
            row += 1

        # Line items section
        line_items = extracted_data.get("line_items", [])
        if line_items:
            li_schema = line_items_schema or self._derive_line_item_fields(line_items)
            if li_schema:
                row += 1
                ws.cell(row=row, column=1, value="Line Items").font = Font(name="Arial", bold=True, size=11, color="2B5797")
                row += 1

                headers = [f["name"].replace("_", " ").title() for f in li_schema]
                for ci, h in enumerate(headers):
                    cell = ws.cell(row=row, column=ci + 1, value=h)
                    cell.font = HEADER_FONT
                    cell.fill = SUBHEADER_FILL
                    cell.border = THIN_BORDER
                row += 1

                field_names = [f["name"] for f in li_schema]
                for item in line_items:
                    for ci, fn in enumerate(field_names):
                        fd = item.get(fn, {})
                        val = fd.get("value") if isinstance(fd, dict) else fd
                        self._write_cell(ws, row, ci + 1, val)
                    row += 1

        # Metadata
        meta = extracted_data.get("metadata", {})
        if meta.get("extraction_notes"):
            row += 1
            ws.cell(row=row, column=1, value="Notes:").font = Font(name="Arial", italic=True, size=9, color="888888")
            ws.cell(row=row, column=2, value=meta["extraction_notes"]).font = Font(name="Arial", size=9, color="888888")

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15

    def _create_combined_sheet(self, doc_type, schema_fields, line_items_schema):
        """Create a combined sheet for a document type."""
        title = doc_type[:31]
        if title in self.wb.sheetnames:
            title = f"{doc_type[:27]}_{len(self._sheets)}"[:31]

        ws = self.wb.create_sheet(title=title)

        headers = ["Source File", "Confidence", "Needs Review"]
        field_names = []

        if schema_fields:
            for f in schema_fields:
                name = f["name"]
                field_names.append(name)
                label = name.replace("_", " ").title()
                if f.get("required"):
                    label += " *"
                headers.append(label)
        headers.append("Notes")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = max(18, len(headers[col - 1]) + 6)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        self._sheets[doc_type] = {
            "worksheet": ws, "next_row": 2,
            "field_names": field_names, "headers": headers,
        }

    def _write_line_items_combined(self, doc_type, filename, items, schema):
        """Write line items to a combined sub-sheet."""
        sheet_name = f"{doc_type}_lines"[:31]

        if sheet_name not in self._sheets:
            ws = self.wb.create_sheet(title=sheet_name)
            headers = ["Source File"] + [f["name"].replace("_", " ").title() for f in schema]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = HEADER_FONT
                cell.fill = SUBHEADER_FILL
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = THIN_BORDER
                ws.column_dimensions[get_column_letter(col)].width = max(18, len(header) + 6)

            ws.freeze_panes = "A2"
            self._sheets[sheet_name] = {"worksheet": ws, "next_row": 2, "field_names": [f["name"] for f in schema]}

        info = self._sheets[sheet_name]
        ws = info["worksheet"]

        for item in items:
            row = info["next_row"]
            self._write_cell(ws, row, 1, filename)
            for i, fname in enumerate(info["field_names"]):
                fd = item.get(fname, {})
                val = fd.get("value") if isinstance(fd, dict) else fd
                cell = self._write_cell(ws, row, 2 + i, val)
                if isinstance(fd, dict) and fd.get("confidence") == "low":
                    cell.fill = LOW_CONF_FILL
            info["next_row"] = row + 1

    def _derive_fields_from_data(self, extracted_data):
        """Derive schema fields from extracted data when no schema provided."""
        ext = extracted_data.get("extracted_data", {})
        fields = []
        for key in ext:
            fd = ext[key]
            val = fd.get("value", "") if isinstance(fd, dict) else fd
            ftype = "string"
            if isinstance(val, (int, float)):
                ftype = "number"
            elif isinstance(val, str) and len(val) == 10 and val.count("-") == 2:
                ftype = "date"
            fields.append({"name": key, "type": ftype, "required": False})
        return fields

    def _derive_line_item_fields(self, items):
        """Derive line item schema from first item."""
        if not items:
            return []
        return [{"name": k, "type": "string"} for k in items[0]]

    def _write_cell(self, ws, row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        return cell

    def add_summary_sheet(self, job_stats):
        """Add a summary sheet."""
        ws = self.wb.create_sheet(title="Summary", index=0)
        tf = Font(name="Arial", bold=True, size=14, color="2B5797")
        lf = Font(name="Arial", bold=True, size=10)
        vf = Font(name="Arial", size=10)

        ws.cell(row=1, column=1, value="DocAgent — Extraction Report").font = tf
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = vf

        stats = [
            ("Total Documents", job_stats.get("total_docs", 0)),
            ("Successfully Extracted", job_stats.get("successful", 0)),
            ("Failed", job_stats.get("failed", 0)),
            ("Needs Review", job_stats.get("needs_review", 0)),
            ("Average Confidence", job_stats.get("avg_confidence", "N/A")),
            ("Total Processing Time", f"{job_stats.get('total_time_sec', 0):.1f}s"),
            ("Client", job_stats.get("client_name", "N/A")),
            ("LLM Provider", job_stats.get("primary_llm", "N/A")),
        ]

        for i, (label, value) in enumerate(stats, start=4):
            ws.cell(row=i, column=1, value=label).font = lf
            ws.cell(row=i, column=2, value=value).font = vf

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

    def save(self, output_path):
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(output_path))
        return output_path
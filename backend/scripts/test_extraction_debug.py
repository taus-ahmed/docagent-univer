"""
DocAgent — Extraction Pipeline Debug Harness
============================================

Runs the FULL web extraction pipeline on a single test document WITHOUT the
HTTP layer or the background thread, then dumps everything needed to inspect
exactly what the LLM returned and how it became Excel cell values.

Outputs (all under backend/debug_output/):
  - last_extraction_raw.json   raw LLM response (written by the llm_router hook)
  - extraction_result.json     the processed DocumentExtractionResult.extracted_data
  - cell_mapping.csv           cell_reference | value_written | source_field_from_json
  - test_result.xlsx           the final Excel output (same writer as the API)

Usage (run from the backend/ directory so engine modules resolve on sys.path):
  cd backend
  .venv\\Scripts\\activate
  python scripts/test_extraction_debug.py
  python scripts/test_extraction_debug.py --pdf data/test_uploads/mydoc.pdf
  python scripts/test_extraction_debug.py --template-id 7 --provider gemini
  python scripts/test_extraction_debug.py --template-file mytemplate.json

Template source (pick one; optional):
  --template-id N      load ColumnTemplate.description from the DB (needs Postgres)
  --template-file P    load a saved grid-layout JSON file (the description blob)
  (none)               no template -> orchestrator schema path -> flat export
"""

import argparse
import csv
import json
import os
import sys
from pathlib import Path

# ── Resolve paths and make engine + app importable (mirrors extract.py) ──────
_SCRIPTS_DIR = Path(__file__).resolve().parent
_BACKEND_DIR = _SCRIPTS_DIR.parent
_PROJECT_DIR = _BACKEND_DIR.parent
_ENGINE_DIR = _BACKEND_DIR / "engine"
_DEBUG_DIR = _BACKEND_DIR / "debug_output"
_UPLOADS_DIR = _BACKEND_DIR / "data" / "test_uploads"

for p in (str(_ENGINE_DIR), str(_BACKEND_DIR), str(_PROJECT_DIR)):
    if p not in sys.path:
        sys.path.insert(0, p)

os.environ.setdefault("PYTHONUTF8", "1")


def _pick_default_pdf() -> Path | None:
    if not _UPLOADS_DIR.exists():
        return None
    candidates = sorted(
        [f for f in _UPLOADS_DIR.iterdir()
         if f.suffix.lower() in {".pdf", ".png", ".jpg", ".jpeg", ".webp",
                                  ".tiff", ".tif", ".bmp", ".heic"}],
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def _load_template(args):
    """Return (template_data, sheet_data) where sheet_data is the raw grid JSON."""
    from app.api.routes.extract import _parse_template, _analyse_template_regions
    from app.models.models import ColumnTemplate

    raw_grid = None
    tpl_obj = None

    if args.template_id:
        from app.models.models import SessionLocal
        db = SessionLocal()
        try:
            tpl_obj = db.query(ColumnTemplate).filter(
                ColumnTemplate.id == int(args.template_id)
            ).first()
            if not tpl_obj:
                print(f"[TEST] template id {args.template_id} not found in DB")
                sys.exit(2)
            # detach a plain copy so we can close the session
            template_data = _parse_template(tpl_obj)
            if tpl_obj.description:
                try:
                    g = json.loads(tpl_obj.description)
                    if isinstance(g, dict) and "cells" in g:
                        raw_grid = g
                except Exception:
                    pass
            return template_data, raw_grid
        finally:
            db.close()

    if args.template_file:
        grid = json.loads(Path(args.template_file).read_text(encoding="utf-8"))
        # Build a throwaway ColumnTemplate to reuse _parse_template's logic
        tpl = ColumnTemplate(
            name="debug_template",
            document_type=args.doc_type or "other",
            description=json.dumps(grid),
            columns_json="[]",
        )
        template_data = _parse_template(tpl)
        if isinstance(grid, dict) and "cells" in grid:
            raw_grid = grid
        return template_data, raw_grid

    return None, None


def _run_extraction(pdf_path: Path, template_data, schema_path: str):
    """Run the real pipeline functions and return list of DocumentExtractionResult."""
    from app.api.routes.extract import (
        _extract_with_template, _extract_image_with_template, _is_image_file,
    )
    from orchestrator import Orchestrator

    orchestrator = Orchestrator(client_schema_path=schema_path)
    is_img = _is_image_file(pdf_path)

    if template_data:
        if is_img:
            return _extract_image_with_template(orchestrator, pdf_path, template_data)
        return _extract_with_template(orchestrator, pdf_path, template_data)
    else:
        if is_img:
            return _extract_image_with_template(orchestrator, pdf_path, None)
        return [orchestrator._process_single_document(pdf_path)]


def _flatten_sources(extracted_data: dict) -> dict:
    """Map normalized-string-value -> source field label, for reverse matching."""
    out = {}

    def add(label, value):
        if value is None:
            return
        sval = str(value).strip()
        if sval == "":
            return
        out.setdefault(sval, label)

    # form fields: extracted_data["extracted_data"] = {label: {value, confidence, ref}}
    inner = extracted_data.get("extracted_data", {})
    if isinstance(inner, dict):
        for label, fd in inner.items():
            if isinstance(fd, dict):
                add(f"{label} ({fd.get('ref','')})".strip(), fd.get("value"))
            else:
                add(label, fd)

    # cell-ref keyed fields
    ef = extracted_data.get("extracted_fields", {})
    if isinstance(ef, dict):
        for ref, v in ef.items():
            add(f"extracted_fields[{ref}]",
                v.get("value") if isinstance(v, dict) else v)

    # table rows
    for key, val in extracted_data.items():
        if key.endswith("_rows") or key == "table_rows":
            if isinstance(val, list):
                for i, row in enumerate(val):
                    if isinstance(row, dict):
                        for col, cv in row.items():
                            if not str(col).startswith("_"):
                                add(f"{key}[{i}].{col}", cv)
    return out


def _build_cell_mapping(xlsx_path: Path, extracted_data: dict):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    sources = _flatten_sources(extracted_data)

    rows = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or str(cell.value).strip() == "":
                continue
            cval = str(cell.value).strip()
            src = sources.get(cval, "")
            if not src:
                # fuzzy: numeric compare ignoring trailing zeros
                for sval, label in sources.items():
                    if sval and (sval == cval or sval.rstrip("0").rstrip(".") ==
                                 cval.rstrip("0").rstrip(".")):
                        src = label
                        break
            if not src:
                src = "(template static label / computed)"
            rows.append((cell.coordinate, cval, src))
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", default=None, help="path to test document")
    ap.add_argument("--template-id", default=None, help="ColumnTemplate id in DB")
    ap.add_argument("--template-file", default=None, help="grid-layout JSON file")
    ap.add_argument("--doc-type", default=None, help="override doc type")
    ap.add_argument("--provider", default="gemini", choices=["gemini", "groq", "env"],
                    help="force PRIMARY_LLM (default gemini); 'env' keeps .env value")
    args = ap.parse_args()

    # Force provider BEFORE importing app.config (settings is lru_cached at import)
    if args.provider != "env":
        os.environ["PRIMARY_LLM"] = args.provider

    pdf_path = Path(args.pdf) if args.pdf else _pick_default_pdf()
    if not pdf_path or not pdf_path.exists():
        print(f"[TEST] No test document found. Place a PDF in {_UPLOADS_DIR} "
              f"or pass --pdf <path>.")
        sys.exit(1)

    _DEBUG_DIR.mkdir(parents=True, exist_ok=True)

    from app.config import settings
    print(f"[TEST] document   : {pdf_path}")
    print(f"[TEST] PRIMARY_LLM : {settings.PRIMARY_LLM}")

    # Schema path for the orchestrator (demo schema is always seeded)
    from app.core.storage import get_storage
    storage = get_storage()
    schema_path = (storage.get_schema_path("demo_001")
                   or storage.get_schema_path("demo_001"))
    if schema_path is None:
        # fall back to the bundled demo schema file
        cand = _ENGINE_DIR / "demo_accounting.yaml"
        schema_path = cand if cand.exists() else None
    print(f"[TEST] schema_path : {schema_path}")

    template_data, raw_grid = _load_template(args)
    if args.doc_type and template_data:
        template_data["doc_type"] = args.doc_type
    print(f"[TEST] template    : "
          f"{'yes (mode=' + str((template_data or {}).get('regions', {}).get('primary_mode')) + ')' if template_data else 'NONE (orchestrator schema path)'}")

    # ── Run the pipeline ─────────────────────────────────────────────────────
    results = _run_extraction(pdf_path, template_data, str(schema_path) if schema_path else "")
    print(f"[TEST] produced {len(results)} result block(s)")

    # ── Save processed extraction result(s) ──────────────────────────────────
    all_extracted = []
    for r in results:
        ed = getattr(r, "extracted_data", None) or {}
        all_extracted.append({"filename": getattr(r, "filename", ""),
                              "success": getattr(r, "success", False),
                              "document_type": getattr(r, "document_type", ""),
                              "extracted_data": ed})
    (_DEBUG_DIR / "extraction_result.json").write_text(
        json.dumps(all_extracted, indent=2, default=str, ensure_ascii=False),
        encoding="utf-8")
    print(f"[TEST] processed result -> {_DEBUG_DIR / 'extraction_result.json'}")

    # ── Write the final Excel using the REAL export writer ────────────────────
    import openpyxl
    from app.api.routes.extract import _write_excel, _write_flat_table
    from app.models.models import DocumentResult

    db_docs = []
    for r in results:
        ed = getattr(r, "extracted_data", None) or {}
        d = DocumentResult(
            filename=getattr(r, "filename", pdf_path.name),
            document_type=getattr(r, "document_type", "") or "unknown",
            overall_confidence=ed.get("overall_confidence"),
            extraction_json=json.dumps(ed, default=str),
            needs_review=bool((ed.get("validation") or {}).get("flagged_count")),
        )
        db_docs.append(d)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    if raw_grid:
        from app.api.routes.extract import _analyse_template_regions
        regions = _analyse_template_regions(raw_grid)
        _write_excel(ws, db_docs, raw_grid, regions, openpyxl)
    else:
        _write_flat_table(ws, db_docs, openpyxl)

    xlsx_path = _DEBUG_DIR / "test_result.xlsx"
    wb.save(xlsx_path)
    print(f"[TEST] excel output -> {xlsx_path}")

    # ── Cell-by-cell mapping table ───────────────────────────────────────────
    primary_ed = all_extracted[0]["extracted_data"] if all_extracted else {}
    mapping = _build_cell_mapping(xlsx_path, primary_ed)
    csv_path = _DEBUG_DIR / "cell_mapping.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["cell_reference", "value_written", "source_field_from_json"])
        w.writerows(mapping)
    print(f"[TEST] cell mapping -> {csv_path}")

    # ── Console summary ──────────────────────────────────────────────────────
    print("\n=== CELL MAPPING (cell | value | source) ===")
    for ref, val, src in mapping[:120]:
        print(f"  {ref:<6} | {val[:40]:<40} | {src}")
    if len(mapping) > 120:
        print(f"  ... {len(mapping) - 120} more rows in {csv_path.name}")

    raw_file = _DEBUG_DIR / "last_extraction_raw.json"
    print(f"\n[TEST] raw LLM response -> {raw_file} "
          f"({'exists' if raw_file.exists() else 'MISSING — no LLM call was made'})")
    print("[TEST] done.")


if __name__ == "__main__":
    main()

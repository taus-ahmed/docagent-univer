"""
DocAgent v2 — Export Routes
POST /api/export/combined  — combined Excel
POST /api/export/perfile   — per-file Excel
"""

import io
import sys
import json
from pathlib import Path
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.auth import get_current_user
from app.models import get_db, User, ExtractionJob, DocumentResult, ColumnTemplate
from app.schemas.schemas import ExportRequest, ExportPerFileRequest

router = APIRouter(prefix="/api/export", tags=["export"])

# Pre-compute paths at module load time
_backend_dir = Path(__file__).resolve().parent.parent.parent.parent
_project_dir = _backend_dir.parent
_engine_dir  = _backend_dir / "engine"

def _ensure_paths():
    for p in [str(_engine_dir), str(_backend_dir), str(_project_dir)]:
        if p not in sys.path:
            sys.path.insert(0, p)


@router.post("/combined")
def export_combined(
    payload: ExportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.api.routes.extract import _get_job_or_404
    job = _get_job_or_404(payload.job_id, current_user, db)
    if job.status != "completed":
        raise HTTPException(status_code=400, detail="Job must be completed before exporting")

    docs = db.query(DocumentResult).filter(DocumentResult.job_id == payload.job_id).order_by(DocumentResult.id).all()
    if not docs:
        raise HTTPException(status_code=404, detail="No documents found")

    # Load template if provided
    template_columns = None
    if payload.template_id:
        tpl = db.query(ColumnTemplate).filter(ColumnTemplate.id == payload.template_id).first()
        if tpl and tpl.columns_json:
            try:
                raw = json.loads(tpl.columns_json)
                template_columns = sorted(
                    [{"name": (c if isinstance(c, str) else c.get("name","")),
                      "type": ("Text" if isinstance(c, str) else c.get("type","Text")),
                      "order": i}
                     for i, c in enumerate(raw)],
                    key=lambda x: x["order"]
                )
            except Exception:
                pass

    excel_bytes = _build_excel(job, docs, template_columns, per_file=False)
    filename = f"docagent_job{job.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/perfile")
def export_perfile(
    payload: ExportPerFileRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.api.routes.extract import _get_job_or_404
    job = _get_job_or_404(payload.job_id, current_user, db)
    if job.status != "completed":
        raise HTTPException(status_code=400, detail="Job must be completed before exporting")

    q = db.query(DocumentResult).filter(DocumentResult.job_id == payload.job_id)
    if payload.doc_ids:
        q = q.filter(DocumentResult.id.in_(payload.doc_ids))
    docs = q.order_by(DocumentResult.id).all()
    if not docs:
        raise HTTPException(status_code=404, detail="No documents found")

    template_columns = None
    if payload.template_id:
        tpl = db.query(ColumnTemplate).filter(ColumnTemplate.id == payload.template_id).first()
        if tpl and tpl.columns_json:
            try:
                raw = json.loads(tpl.columns_json)
                template_columns = sorted(
                    [{"name": (c if isinstance(c, str) else c.get("name","")),
                      "type": ("Text" if isinstance(c, str) else c.get("type","Text")),
                      "order": i}
                     for i, c in enumerate(raw)],
                    key=lambda x: x["order"]
                )
            except Exception:
                pass

    excel_bytes = _build_excel(job, docs, template_columns, per_file=True)
    filename = f"docagent_perfile_job{job.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_excel(job, docs: list, template_columns: list | None, per_file: bool) -> bytes:
    """Build Excel using openpyxl directly — no dependency on engine ExcelWriter."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FONT  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    HEADER_FILL  = PatternFill("solid", fgColor="4F46E5")
    DATA_FONT    = Font(name="Calibri", size=10)
    ALT_FILL     = PatternFill("solid", fgColor="F5F6F8")
    REVIEW_FILL  = PatternFill("solid", fgColor="FFF3CD")
    BORDER = Border(
        left=Side(style="thin", color="E3E6EC"),
        right=Side(style="thin", color="E3E6EC"),
        top=Side(style="thin", color="E3E6EC"),
        bottom=Side(style="thin", color="E3E6EC"),
    )

    wb = Workbook()
    wb.remove(wb.active)

    if per_file:
        # One sheet per document
        for doc in docs:
            ext = doc.get_extracted_data() or {}
            ext_data = ext.get("extracted_data", {})
            sheet_name = Path(doc.filename).stem[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Determine columns
            if template_columns:
                cols = [(c["name"], c["type"]) for c in template_columns]
            else:
                cols = [(k, "Text") for k in ext_data.keys()]

            # Header row
            ws.append(["Field", "Value", "Confidence"])
            for cell in ws[1]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 22

            # Data rows
            for i, (col_name, col_type) in enumerate(cols, 2):
                fd = ext_data.get(col_name)
                if fd is None:
                    value, confidence = "", ""
                elif isinstance(fd, dict):
                    value = fd.get("value", "") or ""
                    confidence = fd.get("confidence", "")
                else:
                    value = str(fd) if fd else ""
                    confidence = ""

                ws.cell(row=i, column=1, value=col_name).font = Font(name="Calibri", bold=True, size=10)
                ws.cell(row=i, column=2, value=str(value) if value else "").font = DATA_FONT
                ws.cell(row=i, column=3, value=confidence).font = DATA_FONT

                if i % 2 == 0:
                    for c in range(1, 4):
                        ws.cell(row=i, column=c).fill = ALT_FILL
                for c in range(1, 4):
                    ws.cell(row=i, column=c).border = BORDER

            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 40
            ws.column_dimensions["C"].width = 12
    else:
        # One combined sheet — all docs as rows
        if template_columns:
            col_names = [c["name"] for c in template_columns]
        else:
            # Collect all unique keys across all docs
            keys = []
            seen = set()
            for doc in docs:
                ext = (doc.get_extracted_data() or {}).get("extracted_data", {})
                for k in ext.keys():
                    if k not in seen:
                        keys.append(k)
                        seen.add(k)
            col_names = keys

        ws = wb.create_sheet(title="Extracted Data")
        headers = ["File", "Doc Type", "Confidence", "Status"] + col_names
        ws.append(headers)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BORDER
        ws.row_dimensions[1].height = 22

        for row_idx, doc in enumerate(docs, 2):
            ext = (doc.get_extracted_data() or {}).get("extracted_data", {})
            fill = REVIEW_FILL if doc.needs_review else (ALT_FILL if row_idx % 2 == 0 else None)

            row_data = [
                doc.filename,
                doc.document_type or "",
                doc.overall_confidence or "",
                "Review" if doc.needs_review else "OK",
            ]
            for col_name in col_names:
                fd = ext.get(col_name)
                if fd is None:
                    row_data.append("")
                elif isinstance(fd, dict):
                    v = fd.get("value", "")
                    row_data.append(str(v) if v is not None and v != "" else "")
                else:
                    row_data.append(str(fd) if fd else "")

            ws.append(row_data)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = DATA_FONT
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if fill:
                    cell.fill = fill

        # Auto-width
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for row_idx in range(2, len(docs) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 50))
            ws.column_dimensions[col_letter].width = max_len + 3

        ws.freeze_panes = "A2"

        # Summary sheet
        ws2 = wb.create_sheet(title="Summary")
        ws2.append(["DocAgent Export Summary"])
        ws2["A1"].font = Font(name="Calibri", bold=True, size=14)
        ws2.append([])
        ws2.append(["Job ID", job.id])
        ws2.append(["Total Documents", job.total_docs])
        ws2.append(["Successful", job.successful])
        ws2.append(["Failed", job.failed])
        ws2.append(["Needs Review", job.needs_review])
        ws2.append(["Processing Time", f"{job.total_time_sec:.1f}s"])
        ws2.append(["Exported At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

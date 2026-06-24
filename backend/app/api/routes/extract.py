"""
DocAgent v2 - Extract Routes (Vision-First Architecture)
=========================================================

ARCHITECTURE OVERVIEW
---------------------
Previous approach: binary TABLE vs FORM decision based on template layout.
Problem: fails on mixed layouts (Form 941, Balance Sheet header+table, etc.)

New approach: Vision-First with pdfplumber validation layer.

EXTRACTION PIPELINE (per document):
  1. TEMPLATE REGION ANALYSIS
     Read the template grid and identify ALL regions:
     - key_value pairs (label in col A, value in col B)
     - two_column_form (labels+values on both left and right sides)
     - table_header (column headers for repeating rows)
     - free_form (labels anywhere, values anywhere nearby)
     Region map is cached - computed once per template.

  2. PAGE IMAGE EXTRACTION (primary - Gemini Vision)
     Convert every PDF page to a base64 image.
     Send image + template region map + registry system prompt to Gemini.
     Gemini sees the document visually - reads any layout correctly.
     Returns JSON with every field value.

  3. pdfplumber TEXT EXTRACTION (parallel - validation layer)
     Extract all text from the PDF independently.
     For each AI-returned value, check whether it appears in the pdfplumber text.
     If yes -> HIGH confidence.
     If no  -> LOW confidence, flag for review.
     This catches AI hallucinations before they reach the Excel output.

  4. MULTI-DOCUMENT HANDLING
     One PDF may contain N separate documents (100 cheques, 50 invoices).
     Vision pre-pass detects document boundaries from the page images.
     Each detected document gets its own extraction pass.
     Each produces its own result block in the Excel output.

  5. TABLE DETECTION (for table-mode templates)
     pdfplumber reads the physical table structure from the PDF.
     Tables may start anywhere on the page (not just top-left).
     If table found -> extract directly, validate with AI.
     If not found   -> AI extracts from image, pdfplumber validates.

  6. EXCEL EXPORT
     Rebuilds the exact template grid per result.
     For table mode: header row + one data row per extracted line item.
     For form mode: one filled template block per document, stacked.
     For 100 cheques: 100 filled blocks in one sheet, separated by filename rows.

CONFIDENCE LEVELS:
  high   = AI value confirmed by pdfplumber text
  medium = AI value not confirmed but plausible (pdfplumber text was partial)
  low    = AI value not found anywhere in pdfplumber text (flag for review)
"""

import io
import re
import sys
import time
import json
import threading
import traceback
from pathlib import Path
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.config import settings
from app.core.auth import get_current_user
from app.core.storage import get_storage
from app.models import get_db, User, ExtractionJob, DocumentResult, ColumnTemplate
from app.schemas.schemas import (
    JobStatus, JobListItem, DocumentResultResponse,
    DocumentUpdateRequest, ExtractUploadResponse,
)

router = APIRouter(prefix="/api", tags=["extract"])

_backend_dir = Path(__file__).resolve().parent.parent.parent.parent
_project_dir = _backend_dir.parent
_engine_dir  = _backend_dir / "engine"

_IMAGE_EXTENSIONS = frozenset({
    ".jpg", ".jpeg", ".png", ".webp", ".tiff", ".tif", ".bmp",
    ".heic", ".heif", ".gif", ".avif",
})


def _is_image_file(file_path: Path) -> bool:
    return file_path.suffix.lower() in _IMAGE_EXTENSIONS


def _is_text_pdf(file_path: Path, min_chars: int = 80) -> bool:
    """Return True if the PDF has machine-readable text (not scanned/image-only).
    Checks first 3 pages — if any yields >= min_chars the PDF is text-based."""
    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages[:3]:
                if len((page.extract_text() or "").strip()) >= min_chars:
                    return True
        return False
    except Exception:
        return False


def _load_image_as_png_bytes(file_path: Path) -> bytes:
    """Load any supported image (including HEIC) as PNG bytes for vision LLM.
    Falls back to raw bytes if Pillow unavailable."""
    suffix = file_path.suffix.lower().lstrip(".")
    if suffix in {"heic", "heif"}:
        try:
            import pillow_heif
            pillow_heif.register_heif_opener()
        except ImportError:
            print("[IMAGE] pillow-heif not installed — run: pip install pillow-heif", flush=True)
    try:
        from PIL import Image
        import io as _io
        with Image.open(file_path) as img:
            if img.mode not in ("RGB", "RGBA"):
                img = img.convert("RGB")
            buf = _io.BytesIO()
            img.save(buf, format="PNG")
            return buf.getvalue()
    except Exception as exc:
        print(f"[IMAGE] Pillow failed ({exc}), reading raw bytes", flush=True)
        return file_path.read_bytes()


# ==============================================================================
# PAGE COUNT ENDPOINT
# ==============================================================================

@router.post("/extract/page-count")
async def get_pdf_page_count(
    files: list[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
):
    """Return page count for each uploaded file (PDFs only). Used by frontend
    to decide whether to show the page-selection modal before extraction."""
    import io
    results = []
    for f in files:
        content = await f.read()
        if f.filename and f.filename.lower().endswith(".pdf"):
            try:
                import pdfplumber
                with pdfplumber.open(io.BytesIO(content)) as pdf:
                    results.append({"filename": f.filename, "page_count": len(pdf.pages)})
            except Exception:
                results.append({"filename": f.filename, "page_count": 1})
        else:
            results.append({"filename": f.filename, "page_count": 1})
    return results


# ==============================================================================
# UPLOAD ENDPOINT
# ==============================================================================

@router.post("/extract/upload", response_model=ExtractUploadResponse, status_code=202)
async def upload_and_extract(
    files: list[UploadFile] = File(...),
    client_id: str = Form(...),
    template_id: Optional[int] = Form(None),
    options: Optional[str] = Form(None),  # JSON array: ["categorize","summary","anomaly","graphs"]
    selected_pages: Optional[str] = Form(None),  # JSON array of 1-based page numbers e.g. "[1,2,5]"
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    storage=Depends(get_storage),
):
    if len(files) > settings.MAX_FILES_PER_BATCH:
        raise HTTPException(status_code=400, detail=f"Max {settings.MAX_FILES_PER_BATCH} files.")

    for f in files:
        content = await f.read()
        await f.seek(0)
        error = storage.validate_upload(f.filename, len(content))
        if error:
            raise HTTPException(status_code=400, detail=f"{f.filename}: {error}")

    schema_path = storage.get_schema_path(client_id) or storage.get_schema_path("demo_001")
    if schema_path is None:
        raise HTTPException(status_code=404, detail="No schema found.")

    template_data = None
    if template_id:
        tpl = db.query(ColumnTemplate).filter(ColumnTemplate.id == template_id).first()
        if tpl:
            template_data = _parse_template(tpl)

    # Parse options
    selected_options: list = []
    if options:
        try:
            selected_options = json.loads(options)
        except Exception:
            selected_options = [o.strip() for o in options.split(",") if o.strip()]

    # Parse selected_pages (1-based list of page numbers to process)
    selected_pages_list: Optional[list] = None
    if selected_pages:
        try:
            parsed = json.loads(selected_pages)
            if isinstance(parsed, list) and parsed:
                selected_pages_list = [int(p) for p in parsed]
        except Exception:
            pass

    job = ExtractionJob(
        user_id=current_user.id,
        client_id=client_id,
        status="pending",
        total_docs=len(files),
        input_source="upload",
        started_at=datetime.utcnow(),
        schema_id=str(template_id) if template_id else None,
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    job_id = job.id

    saved_paths = []
    for f in files:
        content = await f.read()
        local_path, _ = storage.save_upload(content, f.filename, job_id, current_user.id)
        saved_paths.append(str(local_path))

    thread = threading.Thread(
        target=_run_extraction_sync,
        args=(job_id, saved_paths, str(schema_path), settings.DATABASE_URL,
              template_data, str(_project_dir), str(_backend_dir), str(_engine_dir)),
        kwargs={"options": selected_options, "selected_pages": selected_pages_list},
        daemon=True,
    )
    thread.start()

    return ExtractUploadResponse(
        job_id=job_id,
        message=f"Extraction started for {len(files)} file(s)",
        total_files=len(files),
        status="processing",
    )


# ==============================================================================
# TEMPLATE PARSING & REGION ANALYSIS
# ==============================================================================

def _parse_template(tpl: ColumnTemplate) -> Optional[dict]:
    """Parse a ColumnTemplate into extraction-ready format with region analysis."""
    if tpl.description:
        try:
            raw = json.loads(tpl.description)
            if isinstance(raw, dict) and "cells" in raw:
                template_data = {
                    "mode": "layout",
                    "layout": raw,
                    "doc_type": tpl.document_type or "other",
                    "name": tpl.name,
                }
                # Analyse regions once and cache in template_data
                try:
                    template_data["regions"] = _analyse_template_regions(raw)
                except Exception as region_err:
                    import traceback
                    print(f"[TEMPLATE] region analysis error: {region_err}", flush=True)
                    print(f"[TEMPLATE] traceback: {traceback.format_exc()}", flush=True)
                    # Fall back to empty regions so extraction can still proceed
                    template_data["regions"] = {
                        "primary_mode": "form_kv",
                        "explicit_targets": [],
                        "kv_pairs": [],
                        "two_col_pairs": [],
                        "table_regions": [],
                        "transposed_tables": [],
                        "section_label_rows": set(),
                        "has_explicit_targets": False,
                        "has_table": False,
                        "max_row": 0,
                        "max_col": 0,
                    }
                # Bug 5a: detect section-context risk for financial templates
                template_data["regions"]["needs_section_context"] = (
                    _detect_section_context_risk(
                        template_data["regions"], tpl.document_type or "other"
                    )
                )
                return template_data
            else:
                # Bug 4 case 1: JSON without a "cells" key — not a valid grid
                print(
                    f"[TEMPLATE] WARNING (Bug4): description is JSON but lacks 'cells' "
                    f"key — falling through to columns_json",
                    flush=True,
                )
        except json.JSONDecodeError:
            # Bug 4 case 2: plain text description — use it as unguided extraction context
            plain_text = tpl.description.strip() if tpl.description else ""
            if plain_text:
                print(
                    f"[TEMPLATE] WARNING (Bug4): description is plain text — "
                    f"routing to unguided extraction with context: {plain_text[:80]}",
                    flush=True,
                )
                return {
                    "mode": "layout",
                    "layout": {"cells": {}, "extractTargets": []},
                    "doc_type": tpl.document_type or "other",
                    "name": tpl.name,
                    "plain_text_description": plain_text,
                    "regions": {
                        "primary_mode": "unguided",
                        "explicit_targets": [],
                        "kv_pairs": [],
                        "two_col_pairs": [],
                        "table_regions": [],
                        "transposed_tables": [],
                        "section_label_rows": set(),
                        "has_explicit_targets": False,
                        "has_table": False,
                        "max_row": 0,
                        "max_col": 0,
                        "plain_text_description": plain_text,
                        "needs_section_context": False,
                    },
                }
        except Exception as e:
            import traceback
            print(f"[TEMPLATE] description parse error: {e}", flush=True)
            print(f"[TEMPLATE] traceback: {traceback.format_exc()}", flush=True)

    if not tpl.columns_json:
        return None

    try:
        raw = json.loads(tpl.columns_json)
        if isinstance(raw, dict) and "extractTargets" in raw:
            template_data = {
                "mode": "layout",
                "layout": raw,
                "doc_type": tpl.document_type or "other",
                "name": tpl.name,
            }
            template_data["regions"] = _analyse_template_regions(raw)
            return template_data

        header_cols = []
        for i, item in enumerate(raw if isinstance(raw, list) else []):
            if isinstance(item, str):
                col = {"name": item, "type": "Text", "order": i}
            else:
                col = {"name": item.get("name",""), "type": item.get("type","Text"),
                       "order": item.get("order", i)}
            header_cols.append(col)

        return {
            "mode": "columns",
            "header_cols": header_cols,
            "doc_type": tpl.document_type or "other",
            "name": tpl.name,
        }
    except Exception as e:
        print(f"[TEMPLATE] columns_json parse error: {e}", flush=True)
        return None


def _analyse_template_regions(layout: dict) -> dict:
    """
    Analyse the template grid and identify ALL extraction regions.

    This replaces the binary TABLE/FORM decision with a rich structural map
    that tells the AI exactly what the user designed and where.

    Detects:
      - key_value_pairs: label in col N, empty extract cell in col N+1
      - two_column_form: labels+values on BOTH left and right halves
      - table_header: a row of column headers with no values below
      - free_positions: explicit Extract here cells at any grid position
      - blank_near_label: empty cells adjacent to labels (auto-fill mode)

    Returns a regions dict that gets passed to the AI prompt.
    """
    cells = layout.get("cells", {})
    extract_targets = layout.get("extractTargets", [])

    # Parse all cells into a row/col grid
    grid = {}  # (row, col) -> {value, extractTarget, style}
    max_row, max_col = 0, 0
    for key, cell in cells.items():
        if not isinstance(cell, dict):
            continue
        parts = key.split(",")
        if len(parts) != 2:
            continue
        try:
            r, c = int(parts[0]), int(parts[1])
        except (ValueError, TypeError):
            continue
        max_row = max(max_row, r)
        max_col = max(max_col, c)
        # Safely get values — all fields may be None in malformed templates
        raw_val  = cell.get("value")
        val_str  = str(raw_val).strip() if raw_val is not None else ""
        merge_span   = cell.get("mergeSpan")
        merge_parent = cell.get("mergeParent")
        grid[(r, c)] = {
            "value":         val_str,
            "extractTarget": bool(cell.get("extractTarget", False)),
            "mergeParent":   merge_parent if isinstance(merge_parent, (list, dict)) else None,
            "mergeSpan":     merge_span   if isinstance(merge_span,   dict)         else None,
            "ref":           _cell_ref(r, c),
            "row":           r,
            "col":           c,
        }

    # Find all explicit Extract here targets
    # EXCLUDE: merge child cells whose parent is a section header (merged label row)
    # e.g. "Vendor Info" merged across 4 cols — its child cells should not be targets
    merged_section_header_rows = set()
    for (r, c), cell in grid.items():
        if cell["mergeSpan"] and cell["value"] and not cell["extractTarget"]:
            # This is a merged label cell (section header) — mark its row
            span_cols = (cell["mergeSpan"] or {}).get("cols", 1)
            if span_cols >= 2:  # only wide merges are section headers
                merged_section_header_rows.add(r)

    explicit_targets = []
    for (r, c), cell in grid.items():
        if cell["extractTarget"] and not cell["value"]:
            # Skip merge child cells in section header rows
            if cell["mergeParent"] and r in merged_section_header_rows:
                continue
            # Skip cells whose own row is a merged section header
            if r in merged_section_header_rows:
                continue
            # Find the nearest label (left, above, or two left)
            label = ""
            for dc in range(1, 4):  # look up to 3 cells to the left
                neighbor = grid.get((r, c - dc))
                if neighbor and neighbor["value"]:
                    label = neighbor["value"]
                    break
            if not label:
                neighbor = grid.get((r - 1, c))
                if neighbor and neighbor["value"]:
                    label = neighbor["value"]
            explicit_targets.append({
                "ref": _cell_ref(r, c),
                "row": r, "col": c,
                "label": label or f"field at {_cell_ref(r, c)}",
            })

    # Also check extractTargets list from layout
    for t in extract_targets:
        ref = _cell_ref(t.get("r", 0), t.get("c", 0))
        if not any(e["ref"] == ref for e in explicit_targets):
            explicit_targets.append({
                "ref": ref,
                "row": t.get("r", 0),
                "col": t.get("c", 0),
                "label": t.get("label", f"field at {ref}"),
            })

    # Detect key-value pairs (label col N, empty col N+1, same row)
    # Section label rows (rows above table headers) are excluded below
    # after rows_with_content is built
    kv_pairs = []
    for (r, c), cell in grid.items():
        if cell["value"] and not cell["extractTarget"]:
            right = grid.get((r, c + 1))
            if right and (right["extractTarget"] or not right["value"]):
                kv_pairs.append({
                    "label": cell["value"],
                    "label_ref": cell["ref"],
                    "value_ref": _cell_ref(r, c + 1),
                    "row": r,
                })

    # Detect two-column form layout
    # Pattern: labels in col A, values in col B, AND labels in col C, values in col D
    # More generally: labels in even cols, values in odd cols (or vice versa)
    two_col_pairs = []
    for (r, c), cell in grid.items():
        if cell["value"] and not cell["extractTarget"]:
            # Look for a value cell 2 columns to the right (two-column layout)
            right2 = grid.get((r, c + 2))
            if right2 and (right2["extractTarget"] or not right2["value"]):
                # Check there is also a label at c+1 or c+2 area
                mid = grid.get((r, c + 1))
                if mid and mid["value"] and not mid["extractTarget"]:
                    # Pattern: LabelA | LabelB | ValueA | ValueB
                    far_right = grid.get((r, c + 3))
                    if far_right and (far_right["extractTarget"] or not far_right["value"]):
                        two_col_pairs.append({
                            "left_label": cell["value"],
                            "left_label_ref": cell["ref"],
                            "left_value_ref": _cell_ref(r, c + 2),
                            "right_label": mid["value"],
                            "right_label_ref": mid["ref"],
                            "right_value_ref": _cell_ref(r, c + 3),
                            "row": r,
                        })

    # Detect table headers — the actual column header row for repeating data rows.
    # Rules for a valid table header row:
    #   1. Has 2+ non-empty cells
    #   2. NONE of the cells in that row are extract targets (those are form fields)
    #   3. NONE of the cells in adjacent columns of same row are extract targets
    #   4. The row BELOW is empty (data will go there) OR it is the last content row
    # This prevents two-column form rows from being misidentified as table headers.
    table_regions = []
    rows_with_content = {}
    for (r, c), cell in grid.items():
        if cell["value"] or cell["extractTarget"]:
            rows_with_content.setdefault(r, []).append(c)

    # Build a set of rows that contain extract targets — these are form rows.
    # IMPORTANT: Only mark a row as a form row if IT ITSELF contains extract targets.
    form_rows_set = set()
    for (r, c), cell in grid.items():
        if cell["extractTarget"]:
            form_rows_set.add(r)

    # Now compute pre_table_header_rows — rows directly above table header candidates.
    # These are section labels (e.g. "Earning Table", "Deduction Table") and should
    # be excluded from kv_pairs so they aren't sent to AI as form fields to fill.
    pre_table_header_rows = set()
    for r_candidate, cols_candidate in rows_with_content.items():
        if r_candidate in form_rows_set:
            continue
        value_cols_candidate = [c for c in cols_candidate
                                if grid.get((r_candidate, c))
                                and grid[(r_candidate, c)]["value"]
                                and not grid[(r_candidate, c)]["extractTarget"]]
        if len(value_cols_candidate) >= 2:
            # Bug 2: look 2 rows below — a single row of example values is still
            # a valid table header. Only skip if BOTH r+1 AND r+2 have content.
            below1 = [c for c in rows_with_content.get(r_candidate + 1, [])
                      if grid.get((r_candidate + 1, c))
                      and grid[(r_candidate + 1, c)]["value"]
                      and not grid[(r_candidate + 1, c)]["extractTarget"]]
            below2 = [c for c in rows_with_content.get(r_candidate + 2, [])
                      if grid.get((r_candidate + 2, c))
                      and grid[(r_candidate + 2, c)]["value"]
                      and not grid[(r_candidate + 2, c)]["extractTarget"]]
            if below1 and below2:
                continue  # two consecutive content rows below → not a table header
            # This row is a real table header — mark rows above it as section labels
            for r_above in range(max(0, r_candidate - 4), r_candidate):
                if r_above not in form_rows_set:
                    pre_table_header_rows.add(r_above)

    # Save unfiltered kv_pairs for parallel-column group detection.
    # The filter below removes rows that are section labels above table headers
    # (e.g. "Total" rows at row 9 that sit above the liabilities header at row 10).
    # Those rows ARE valid kv_pairs for structural detection — if we filter them out
    # before parallel detection, the band item count drops below the ≥3 threshold
    # and the balance sheet parallel groups are never found.
    kv_pairs_for_parallel = kv_pairs[:]   # snapshot BEFORE section-label filter

    # Filter kv_pairs to exclude section label rows (for AI prompts / mode logic)
    kv_pairs = [kv for kv in kv_pairs if kv["row"] not in pre_table_header_rows]

    for r, cols in sorted(rows_with_content.items()):
        # Skip rows that are form rows (have extract targets nearby)
        if r in form_rows_set:
            continue

        value_cols = [c for c in cols
                      if grid.get((r, c)) and grid[(r, c)]["value"]
                      and not grid[(r, c)]["extractTarget"]]

        if len(value_cols) < 2:
            continue

        # A valid table header MUST NOT have two consecutive content rows below it.
        # Form section headers (like "Summary", "Opening Bal | Closing Bal")
        # have content below them — they are NOT table headers.
        # A table header with one example row below (Bug 2: client puts sample data)
        # is still a valid table header.
        # Bug 2: check r+1 AND r+2 — only skip when BOTH rows have content.
        below1_cols = [c for c in rows_with_content.get(r + 1, [])
                       if grid.get((r + 1, c)) and grid[(r + 1, c)]["value"]
                       and not grid[(r + 1, c)]["extractTarget"]]
        below2_cols = [c for c in rows_with_content.get(r + 2, [])
                       if grid.get((r + 2, c)) and grid[(r + 2, c)]["value"]
                       and not grid[(r + 2, c)]["extractTarget"]]

        # Two consecutive content rows below → this is a form section, not a table header
        if below1_cols and below2_cols:
            continue

        # Also skip if this row is a merged section header
        if r in merged_section_header_rows:
            continue

        row_labels = [grid[(r, c)]["value"] for c in sorted(value_cols)]

        min_col = min(value_cols)
        max_col = max(value_cols)
        table_regions.append({
            "header_row": r,
            "start_col": min_col,
            "end_col": max_col,
            "start_ref": _cell_ref(r, min_col),
            "end_ref": _cell_ref(r, max_col),
            "column_names": row_labels,
            "is_header_only": True,
        })

    # Keep ALL valid table regions — sorted by row then by column position
    # Each table gets a label from the nearest section header above it
    # OR from its column position if it shares a row with another table
    if table_regions:
        # Sort by row first, then by start column
        table_regions.sort(key=lambda t: (t["header_row"], t["start_col"]))

        # Group tables that share the same header_row
        from itertools import groupby
        same_row_groups = {}
        for tbl in table_regions:
            same_row_groups.setdefault(tbl["header_row"], []).append(tbl)

        for row, tbls_on_row in same_row_groups.items():
            for idx, tbl in enumerate(tbls_on_row):
                # Find section label: row directly above table header (1-2 rows max)
                # MUST be blank row between form fields and table OR be a merged header
                # Do NOT pick up form field labels that are 3+ rows above
                section_label = ""
                for row_above in range(row - 1, max(0, row - 3), -1):
                    for c in range(tbl["start_col"], tbl["end_col"] + 1):
                        cell = grid.get((row_above, c))
                        if not cell or not cell["value"] or cell["extractTarget"]:
                            continue
                        # Accept as section label only if:
                        # 1. It's a wide merged cell (span >= 2 cols), OR
                        # 2. It's directly above (1 row) and the row between is blank
                        merge_span = cell.get("mergeSpan") or {}
                        is_merged = merge_span.get("cols", 1) >= 2
                        is_directly_above = (row_above == row - 1)
                        row_between_empty = not any(
                            (grid.get((r, c2)) or {}).get("value") or
                            (grid.get((r, c2)) or {}).get("extractTarget")
                            for r in range(row_above + 1, row)
                            for c2 in range(tbl["start_col"], tbl["end_col"] + 1)
                        )
                        if is_merged or (is_directly_above and row_between_empty):
                            section_label = cell["value"]
                            break
                    if section_label:
                        break

                if not section_label:
                    if len(tbls_on_row) > 1:
                        # Multiple tables on same row — use column range as label
                        start_letter = chr(ord('A') + tbl["start_col"])
                        end_letter   = chr(ord('A') + tbl["end_col"])
                        section_label = (
                            f"Table {_cell_ref(row, tbl['start_col'])[:1]}"
                            f"-{_cell_ref(row, tbl['end_col'])[:1]}"
                            f" ({', '.join(tbl['column_names'][:2])})"
                        )
                    else:
                        section_label = f"Table {table_regions.index(tbl) + 1}"

                tbl["section_label"] = section_label

                # Store col_range so AI knows which columns belong to this table
                tbl["col_range"] = (tbl["start_col"], tbl["end_col"])
                tbl["col_letters"] = (
                    chr(ord('A') + tbl["start_col"]),
                    chr(ord('A') + min(tbl["end_col"], 25))
                )

    print(f"[REGION] {len(table_regions)} table(s) detected: "
          f"{[t.get('section_label','?') for t in table_regions]}", flush=True)

    # Detect transposed tables (headers in col A, data in cols B+)
    transposed_tables = _detect_transposed_table(grid, max_row, max_col)
    if transposed_tables:
        print(f"[REGION] {len(transposed_tables)} transposed table(s) detected", flush=True)

    # Build section_label_rows — rows that:
    #   1. Have text content
    #   2. No extract targets on the row
    #   3. Are NOT table header rows themselves
    #   4. Are directly above a table header row (within 4 rows)
    # These rows are template structure labels, not data — never write as data rows.
    table_header_row_set = {t["header_row"] for t in table_regions}
    section_label_rows = set()
    for tbl in table_regions:
        hr = tbl["header_row"]
        # Only look 2 rows above the table header for section labels.
        # Looking further up risks capturing form section headers (like "Summary")
        # that belong to the form fields above, not to this table.
        for r_above in range(max(0, hr - 2), hr):
            if r_above in table_header_row_set:
                continue
            if r_above in form_rows_set:
                continue
            row_has_text = any(
                grid.get((r_above, c), {}).get("value", "")
                for c in range(max_col + 1)
            )
            if row_has_text:
                section_label_rows.add(r_above)

    # Build summary for AI
    has_explicit_targets = len(explicit_targets) > 0
    has_kv_pairs = len(kv_pairs) > 0
    has_two_col = len(two_col_pairs) > 0
    has_table = len(table_regions) > 0

    # Mixed mode: has BOTH form fields (targets or kv pairs) AND a table
    # This is the most important detection — Balance Sheet, Expense Report,
    # Purchase Order, Tax Form all have header fields + line items table
    if has_table and (has_explicit_targets or has_kv_pairs):
        primary_mode = "mixed"
    elif not has_table and has_explicit_targets:
        primary_mode = "form_with_targets"
    elif not has_table and has_kv_pairs:
        primary_mode = "form_kv"
    elif has_table and not has_explicit_targets and not has_kv_pairs:
        primary_mode = "table"
    elif has_two_col:
        # Two-column form detected but no explicit targets or kv_pairs
        primary_mode = "form_kv"
    else:
        # Bug 1: all region lists empty — fall back to unguided extraction
        # using only the doc_type system prompt
        primary_mode = "unguided"
        print(
            "[TEMPLATE] WARNING (Bug1): no regions detected in template grid "
            "— switching to unguided extraction (doc_type system prompt only)",
            flush=True,
        )

    print(f"[REGION] mode={primary_mode} targets={len(explicit_targets)} "
          f"kv={len(kv_pairs)} two_col={len(two_col_pairs)} "
          f"tables={len(table_regions)} grid={max_row+1}x{max_col+1}", flush=True)

    # ── PARALLEL COLUMN GROUPS DETECTION ──────────────────────────────────────────
    # Detect when template has two or more side-by-side label/value column bands
    # (e.g. balance sheet: Current Assets in A-B, Non-Current Assets in C-D).
    # When parallel groups ARE detected, ALWAYS override mode to parallel_groups.
    # The "table regions" on balance-sheet templates are the parallel group section
    # headers (e.g. row 0: "Current assets | Amount | Non-current assets | Amount"),
    # not real repeating-row tables. Using parallel_groups routes to _write_form_excel
    # which iterates every template row including dynamic fill rows.
    # Use the unfiltered snapshot so Total/header rows (which may have been
    # excluded from kv_pairs by the section-label filter above) still contribute
    # to band item counts — preventing false-negative parallel group detection.
    parallel_column_groups = _detect_parallel_column_groups(
        kv_pairs_for_parallel, grid, max_row, max_col
    )
    if parallel_column_groups:
        primary_mode = "parallel_groups"
        # FIX 1: when parallel groups are active, suppress all other region
        # interpretations so ONLY parallel-group instructions reach the AI.
        # Leaving these populated makes the prompt contradictory (telling Gemini
        # to fill cells AND extract tables AND transpose the same grid).
        table_regions = []
        transposed_tables = []
        two_col_pairs = []
        has_table = False

    return {
        "primary_mode":            primary_mode,
        "explicit_targets":        explicit_targets,
        "kv_pairs":                kv_pairs,
        "two_col_pairs":           two_col_pairs,
        "table_regions":           table_regions,
        "transposed_tables":       transposed_tables,
        "section_label_rows":      section_label_rows,
        "parallel_column_groups":  parallel_column_groups,
        "grid_size":               {"rows": max_row + 1, "cols": max_col + 1},
        "has_explicit_targets":    has_explicit_targets,
        "has_table":               has_table,
        "max_row":                 max_row,
        "max_col":                 max_col,
    }


def _detect_parallel_column_groups(kv_pairs: list, grid: dict,
                                    max_row: int, max_col: int) -> list:
    """
    Detect parallel column groups: two or more independent label/value column bands
    that occupy the same row range in the template.

    Example balance sheet layout (rows 0-N):
      col A = Current assets labels  | col B = empty (extract) |
      col C = Non-current labels     | col D = empty (extract)

    Returns [] when not detected, or a list of group dicts:
      [{group_id, label_col, value_col, label_col_letter, value_col_letter,
        section_label, items:[{label, label_ref, value_ref, row}]}]
    """
    if len(kv_pairs) < 4:
        return []

    # Group kv_pairs by their label column index
    col_groups: dict = {}
    for kv in kv_pairs:
        label_ref = kv.get("label_ref", "")
        if not label_ref:
            continue
        col_letters = "".join(ch for ch in label_ref if ch.isalpha()).upper()
        if not col_letters:
            continue
        col_idx = sum(
            (ord(ch) - 64) * (26 ** i)
            for i, ch in enumerate(reversed(col_letters))
        ) - 1
        col_groups.setdefault(col_idx, []).append(kv)

    # Need 2+ bands with at least 3 items each
    bands = sorted(
        [(col, kvs) for col, kvs in col_groups.items() if len(kvs) >= 3],
        key=lambda x: x[0],
    )
    if len(bands) < 2:
        return []

    # Verify that bands share significant row overlap (>= 40% of their union)
    band_row_sets = [set(kv["row"] for kv in kvs) for _, kvs in bands]
    has_overlap = False
    for i in range(len(band_row_sets)):
        for j in range(i + 1, len(band_row_sets)):
            union = band_row_sets[i] | band_row_sets[j]
            intersect = band_row_sets[i] & band_row_sets[j]
            if union and len(intersect) / len(union) >= 0.40:
                has_overlap = True
                break
        if has_overlap:
            break

    if not has_overlap:
        return []

    parallel_groups = []
    for i, (label_col, items) in enumerate(bands):
        value_col = label_col + 1
        l_letter = chr(ord('A') + min(label_col, 25))
        v_letter = chr(ord('A') + min(value_col, 25))

        # Section label: the last non-item row at-or-above the first item row
        # at this column (i.e. the heading row that names the group).
        item_rows = {kv["row"] for kv in items}
        first_item_row = min(item_rows)
        section_label = ""
        for r in range(0, first_item_row + 1):
            if r in item_rows:
                continue
            cell = grid.get((r, label_col))
            if cell and cell["value"] and not cell["extractTarget"]:
                section_label = cell["value"]  # keep updating — take last found

        if not section_label:
            section_label = f"Group {i + 1} (columns {l_letter}-{v_letter})"

        parallel_groups.append({
            "group_id":         i + 1,
            "label_col":        label_col,
            "value_col":        value_col,
            "label_col_letter": l_letter,
            "value_col_letter": v_letter,
            "section_label":    section_label,
            "items":            sorted(items, key=lambda x: x["row"]),
        })

    print(
        "[REGION] parallel_column_groups detected: "
        + ", ".join(
            f'"{g["section_label"]}" ({g["label_col_letter"]}-{g["value_col_letter"]})'
            for g in parallel_groups
        ),
        flush=True,
    )
    return parallel_groups


# ==============================================================================
# REGISTRY LOADER
# ==============================================================================

def _load_registry():
    if not hasattr(_load_registry, "_cache"):
        try:
            import importlib.util
            reg_file = Path(__file__).resolve().parent / "prompt_registry.py"
            if reg_file.exists():
                spec = importlib.util.spec_from_file_location("prompt_registry", reg_file)
                mod = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(mod)
                _load_registry._cache = mod
                print(f"[REGISTRY] Loaded from {reg_file}", flush=True)
            else:
                _load_registry._cache = None
        except Exception as e:
            print(f"[REGISTRY] Load error: {e}", flush=True)
            _load_registry._cache = None
    return _load_registry._cache


def _get_system_prompt(doc_type): r=_load_registry(); return r.get_system_prompt(doc_type) if r else f"You are an expert {doc_type} extraction specialist."
def _get_unguided_prompt(): r=_load_registry(); return r.get_unguided_prompt() if r else "You are an expert document data extraction AI. Extract all visible fields as JSON."
def _get_table_rules(doc_type): r=_load_registry(); return (r.get_table_rules(doc_type) or "") if r else ""
def _get_numeric_fields(doc_type): r=_load_registry(); return r.get_numeric_fields(doc_type) if r else []
def _get_date_fields(doc_type): r=_load_registry(); return r.get_date_fields(doc_type) if r else []
def _classify_by_hints(text): r=_load_registry(); return r.classify_by_hints(text) if r else None


# ==============================================================================
# FINANCIAL SECTION CONTEXT RISK DETECTION  (Bug 5)
# ==============================================================================

_FINANCIAL_SECTIONED_TYPES = frozenset({
    "balance_sheet", "income_statement", "audit_report", "payslip",
})

_GENERIC_FINANCIAL_COL_WORDS = frozenset({
    "total", "amount", "value", "balance", "net", "gross", "subtotal",
    "debit", "credit", "description",
})


def _detect_section_context_risk(regions: dict, doc_type: str) -> bool:
    """
    Return True when the template has generic repeated column headers (e.g. Total,
    Amount, Value) in a financial document type that contains section-level totals.
    These templates create ambiguity: the AI may pick the wrong section's total
    without explicit positional guidance.
    """
    if doc_type not in _FINANCIAL_SECTIONED_TYPES:
        return False
    table_regions = regions.get("table_regions", [])
    kv_pairs = regions.get("kv_pairs", [])
    explicit_targets = regions.get("explicit_targets", [])

    # Collect all column names from tables
    all_col_names_lower = [
        cn.lower().strip()
        for tr in table_regions
        for cn in tr.get("column_names", [])
    ]
    # Count how many generic financial words appear in column names
    generic_col_count = sum(
        1 for cn in all_col_names_lower
        if any(g in cn for g in _GENERIC_FINANCIAL_COL_WORDS)
    )
    # Also flag if kv_pairs or targets have multiple generic labels
    generic_field_count = sum(
        1 for kv in kv_pairs
        if any(g in kv.get("label", "").lower() for g in _GENERIC_FINANCIAL_COL_WORDS)
    ) + sum(
        1 for t in explicit_targets
        if any(g in t.get("label", "").lower() for g in _GENERIC_FINANCIAL_COL_WORDS)
    )

    risk = generic_col_count >= 2 or generic_field_count >= 3
    if risk:
        print(
            f"[BUG5] Section-context risk detected: doc_type={doc_type} "
            f"generic_cols={generic_col_count} generic_fields={generic_field_count}",
            flush=True,
        )
    return risk


# ==============================================================================
# CELL HELPERS
# ==============================================================================

def _cell_ref(r: int, c: int) -> str:
    col_letter = ""
    n = c
    while True:
        col_letter = chr(65 + (n % 26)) + col_letter
        n = n // 26 - 1
        if n < 0:
            break
    return f"{col_letter}{r + 1}"

def _col_letter(c: int) -> str:
    letter = ""
    n = c
    while True:
        letter = chr(65 + (n % 26)) + letter
        n = n // 26 - 1
        if n < 0:
            break
    return letter


# ==============================================================================
# VISION-FIRST PROMPT BUILDER
# ==============================================================================

def _smart_truncate(doc_text: str, primary_mode: str, regions: dict) -> str:
    """
    Smart text truncation for long documents.
    Also replaces the PAGE BREAK marker with an explicit label the AI understands.
    """
    if not doc_text:
        return ""

    MAX_FORM  = 3000
    MAX_TABLE = 8000

    # Replace page break marker with explicit AI-readable label
    # This is critical — the AI must know content continues on next page
    doc_text = doc_text.replace(
        "--- PAGE BREAK ---",
        "\n\n=== DOCUMENT CONTINUES ON NEXT PAGE — ALL CONTENT BELOW IS PART OF THE SAME DOCUMENT ===\n\n"
    )

    # Short doc — no truncation needed
    if len(doc_text) <= MAX_TABLE:
        return doc_text

    if primary_mode in ("form_with_targets", "form_kv"):
        return doc_text[:MAX_FORM]

    # For table/mixed — smart extraction
    table_regions = regions.get("table_regions", [])
    lines = doc_text.split('\n')

    # Always include first 1500 chars for header fields
    header_text = doc_text[:1500]
    collected   = [header_text]
    used_chars  = len(header_text)

    if table_regions:
        for tr in table_regions:
            col_names   = tr.get("column_names", [])
            section_lbl = tr.get("section_label", "")
            if not col_names:
                continue

            section_start = -1
            search_terms  = col_names[:2] + ([section_lbl] if section_lbl else [])

            for i, line in enumerate(lines):
                if any(term.lower() in line.lower() for term in search_terms if term):
                    section_start = i
                    break

            if section_start == -1:
                continue

            section_lines = lines[section_start:section_start + 100]
            section_text  = '\n'.join(section_lines)

            budget = min(2000, MAX_TABLE - used_chars - 100)
            if budget <= 0:
                break

            snippet = section_text[:budget]
            collected.append(f"\n--- SECTION: {section_lbl or 'Table'} ---\n{snippet}")
            used_chars += len(snippet)

    result = '\n'.join(collected)

    if used_chars < MAX_TABLE:
        remaining = doc_text[1500:MAX_TABLE - used_chars + 1500]
        result += '\n' + remaining

    return result[:MAX_TABLE]


def _detect_transposed_table(grid: dict, max_row: int, max_col: int) -> list:
    """
    Detect transposed (horizontal) tables where:
    - Headers are in COLUMN A (rows going down)
    - Data is in columns B, C, D... (each column = one record)

    Example:
      A1=Name    B1=Alice  C1=Bob
      A2=Dept    B2=HR     C2=Eng
      A3=Salary  B3=50000  C3=60000

    Returns list of transposed table regions, each with:
      header_col, start_row, end_row, row_names, data_cols
    """
    transposed = []

    # Check if column 0 has multiple consecutive labels
    col0_labels = []
    for r in range(max_row + 1):
        cell = grid.get((r, 0))
        if cell and cell["value"] and not cell["extractTarget"]:
            col0_labels.append((r, cell["value"]))

    if len(col0_labels) < 3:
        return []  # Not enough labels in col A to be a transposed table

    # Check if columns 1+ have data in the same rows
    label_rows = [r for r, _ in col0_labels]
    data_cols  = []
    for c in range(1, min(max_col + 1, 20)):
        filled = sum(
            1 for r in label_rows
            if grid.get((r, c)) and grid[(r, c)]["value"]
        )
        if filled >= len(label_rows) * 0.5:  # at least 50% of label rows have data
            data_cols.append(c)

    if data_cols:
        transposed.append({
            "header_col":  0,
            "start_row":   min(label_rows),
            "end_row":     max(label_rows),
            "row_names":   [lbl for _, lbl in col0_labels],
            "data_cols":   data_cols,
            "is_transposed": True,
        })

    return transposed


def _preserve_currency(value: str) -> tuple:
    """
    Extract currency symbol before stripping it.
    Returns (normalized_number, currency_code).
    Used when currency context is important.
    """
    symbols = {
        '$': 'USD', '£': 'GBP', '€': 'EUR', '₹': 'INR',
        '¥': 'JPY', '₩': 'KRW', 'A$': 'AUD', 'C$': 'CAD',
        'HK$': 'HKD', 'S$': 'SGD', 'NZ$': 'NZD',
    }
    s = str(value).strip()
    for symbol, code in sorted(symbols.items(), key=lambda x: -len(x[0])):
        if s.startswith(symbol):
            number = s[len(symbol):].replace(',', '').strip()
            return number, code
    return s, ""


def _build_page_anchor_map(doc_text: str, table_regions: list) -> dict:
    """
    Scan document text to find which page each table's content appears on.
    Searches for: section label name, column header names, and common synonyms.
    Zero hardcoding — purely derived from template structure vs document text.

    Returns: {section_label: page_number}
    """
    if not doc_text or not table_regions:
        return {}

    pages = doc_text.split("--- PAGE BREAK ---")
    if len(pages) <= 1:
        return {}

    page_map = {}
    for tbl in table_regions:
        section   = tbl.get("section_label", "")
        col_names = tbl.get("column_names", [])
        if not col_names and not section:
            continue

        # Build search terms:
        # 1. Section label words (e.g. "Earning Table" -> ["earning", "table"])
        # 2. Column header names
        # 3. First word of each column name (catches "Earning Type" -> "earning")
        search_terms = set()
        if section:
            for word in section.lower().split():
                if len(word) > 3:  # skip short words like "and", "the"
                    search_terms.add(word)
        for col in col_names:
            search_terms.add(col.lower())
            first_word = col.lower().split()[0] if col else ""
            if len(first_word) > 3:
                search_terms.add(first_word)

        # Score each page by how many search terms appear
        best_page  = 0
        best_score = 0
        for page_num, page_text in enumerate(pages, 1):
            page_lower = page_text.lower()
            score = sum(1 for term in search_terms if term in page_lower)
            if score > best_score:
                best_score = score
                best_page  = page_num

        # Only anchor if we found a meaningful match
        if best_page > 0 and best_score > 0:
            page_map[section] = best_page

    # If all tables map to the same page, anchoring adds no value — skip it
    unique_pages = set(page_map.values())
    if len(unique_pages) <= 1:
        # Try to differentiate by finding where the MOST table-specific content is
        # Use the LAST significant keyword from each table's columns
        page_map_refined = {}
        for tbl in table_regions:
            section   = tbl.get("section_label", "")
            col_names = tbl.get("column_names", [])
            if not section or not col_names:
                continue
            # Use the last column name as a differentiator
            # (first col is often shared, e.g. "Amount" appears in both tables)
            differentiator = col_names[-1].lower() if len(col_names) > 1 else col_names[0].lower()
            # Also try section label first word
            sec_word = section.lower().split()[0] if section else ""

            best_page  = 0
            best_score = 0
            for page_num, page_text in enumerate(pages, 1):
                page_lower = page_text.lower()
                score = (
                    (2 if sec_word and sec_word in page_lower else 0) +
                    (1 if differentiator in page_lower else 0)
                )
                if score > best_score:
                    best_score = score
                    best_page  = page_num
            if best_page > 0:
                page_map_refined[section] = best_page

        # Use refined map if it produces different pages
        if len(set(page_map_refined.values())) > 1:
            return page_map_refined

    return page_map


# ==============================================================================
# NEIGHBOR-MATRIX BINDING MAP  (layout-based extraction for unlabeled templates)
# ==============================================================================

def _slug(s: str) -> str:
    """Lowercase, underscore-joined slug for section keys."""
    return re.sub(r'[^a-z0-9]+', '_', (s or "").lower()).strip('_')


def compute_binding_map(template_data: dict, grid: dict):
    """
    COMPONENT 1/2/3 — analyse every cell in the template bounding box, examine its
    8 neighbours (resolved through the merge map), detect section boundaries, and
    assign each cell a role. Returns a dict keyed by "r,c" plus a "_meta" entry, or
    None on any failure (caller falls back to _analyse_template_regions gracefully).

    Roles: merged_child | section_header | column_header | label | static_text |
           value_target | table_data | section_spacer | unknown
    """
    try:
        cells  = grid.get("cells", {})  if isinstance(grid, dict) else {}
        merges = grid.get("merges", {}) if isinstance(grid, dict) else {}

        parsed = {}
        max_r = max_c = 0
        for key, cell in cells.items():
            if not isinstance(cell, dict):
                continue
            parts = key.split(",")
            if len(parts) != 2:
                continue
            try:
                r, c = int(parts[0]), int(parts[1])
            except (ValueError, TypeError):
                continue
            max_r = max(max_r, r); max_c = max(max_c, c)
            parsed[(r, c)] = {
                "value":   str(cell.get("value") or "").strip(),
                "extract": bool(cell.get("extractTarget")),
                "span":    cell.get("mergeSpan") if isinstance(cell.get("mergeSpan"), dict) else None,
            }
        if not parsed:
            return None

        # COMPONENT 2 — merge_map: child cell -> parent cell.
        merge_map = {}
        for pkey, span in (merges or {}).items():
            pp = str(pkey).split(",")
            if len(pp) != 2:
                continue
            try:
                pr, pc = int(pp[0]), int(pp[1])
            except (ValueError, TypeError):
                continue
            sr = int((span or {}).get("rows", 1) or 1)
            sc = int((span or {}).get("cols", 1) or 1)
            for dr in range(sr):
                for dc in range(sc):
                    if (dr, dc) != (0, 0):
                        merge_map[(pr + dr, pc + dc)] = (pr, pc)
        for (r, c), info in parsed.items():           # honour per-cell mergeSpan too
            sp = info.get("span")
            if sp:
                sr = int(sp.get("rows", 1) or 1); sc = int(sp.get("cols", 1) or 1)
                for dr in range(sr):
                    for dc in range(sc):
                        if (dr, dc) != (0, 0):
                            merge_map.setdefault((r + dr, c + dc), (r, c))

        def val(r, c):     return parsed.get((r, c), {}).get("value", "")
        def is_extract(r, c):
            cell = parsed.get((r, c)); return bool(cell and cell["extract"])
        def empty(r, c):
            cell = parsed.get((r, c)); return (cell is None) or (not cell["value"])

        # Column headers — per column, the topmost text cell whose cell directly
        # below is empty/extract (it heads a column of blanks = a real header).
        col_header = {}
        for c in range(max_c + 1):
            for r in range(max_r + 1):
                if val(r, c):
                    if empty(r + 1, c) or is_extract(r + 1, c):
                        col_header[c] = {"row": r, "text": val(r, c)}
                    break   # only the topmost text cell of the column can be a header

        # COMPONENT 3 — section boundaries (walk rows top-to-bottom).
        row_section = {}
        current_section = ""
        prev_blank = True
        for r in range(max_r + 1):
            row_has_content = any(val(r, c) or is_extract(r, c) for c in range(max_c + 1))
            c0 = val(r, 0)
            if c0 and c0 != current_section and (empty(r + 1, 0) or is_extract(r + 1, 0) or prev_blank):
                current_section = c0
            row_section[r] = _slug(current_section) or "section"
            prev_blank = not row_has_content

        binding = {}
        for r in range(max_r + 1):
            for c in range(max_c + 1):
                key = f"{r},{c}"
                if (r, c) in merge_map:                         # 1. merged child
                    pr, pc = merge_map[(r, c)]
                    binding[key] = {"role": "merged_child", "parent": f"{pr},{pc}",
                                    "extract": False, "section": row_section.get(r, "")}
                    continue
                v = val(r, c)
                span = parsed.get((r, c), {}).get("span")
                wide = isinstance(span, dict) and int(span.get("cols", 1) or 1) >= 2
                if v:                                           # 2. text cell
                    ch = col_header.get(c)
                    if wide:
                        binding[key] = {"role": "section_header", "text": v, "extract": False}
                    elif ch and ch["row"] == r:
                        binding[key] = {"role": "column_header", "text": v,
                                        "col_index": c, "extract": False}
                    elif empty(r, c + 1) or is_extract(r, c + 1):
                        binding[key] = {"role": "label", "text": v,
                                        "owns": f"{r},{c+1}", "extract": False}
                    elif empty(r + 1, c) or is_extract(r + 1, c):
                        binding[key] = {"role": "label", "text": v,
                                        "owns": f"{r+1},{c}", "extract": False}
                    else:
                        binding[key] = {"role": "static_text", "text": v, "extract": False}
                    binding[key]["section"] = row_section.get(r, "")
                else:                                           # 3. empty cell
                    left_lbl  = val(r, c - 1) if c > 0 else ""
                    above_lbl = val(r - 1, c) if r > 0 else ""
                    left_is_header = bool(c > 0 and col_header.get(c - 1, {}).get("row") == r)
                    this_is_header_row = bool(col_header.get(c) and col_header[c]["row"] == r)
                    if left_lbl and not left_is_header and not this_is_header_row:
                        binding[key] = {"role": "value_target", "label": left_lbl,
                                        "label_cell": f"{r},{c-1}", "extract": True,
                                        "section": row_section.get(r, "")}
                    elif above_lbl and col_header.get(c) is None:
                        binding[key] = {"role": "value_target", "label": above_lbl,
                                        "label_cell": f"{r-1},{c}", "extract": True,
                                        "section": row_section.get(r, "")}
                    else:
                        ch = col_header.get(c)
                        if ch and ch["row"] < r:               # b. inherit from column header above
                            binding[key] = {
                                "role": "table_data", "col_header": ch["text"],
                                "col_header_cell": f"{ch['row']},{c}", "col_index": c,
                                "row_index": r - ch["row"], "row_group": r,
                                "section": row_section.get(r, ""),
                                "extract": True, "inherit_from": f"{ch['row']},{c}",
                            }
                        else:
                            neigh = [(r + dr, c + dc) for dr in (-1, 0, 1) for dc in (-1, 0, 1)
                                     if (dr, dc) != (0, 0)]
                            if all(empty(nr, nc) for nr, nc in neigh):
                                binding[key] = {"role": "section_spacer", "extract": False,
                                                "section": row_section.get(r, "")}
                            else:
                                binding[key] = {"role": "unknown", "extract": True,
                                                "needs_review": True,
                                                "section": row_section.get(r, "")}

        for r in range(max_r + 1):                              # row_siblings for table_data
            sibs = [f"{r},{c}" for c in range(max_c + 1)
                    if binding.get(f"{r},{c}", {}).get("role") == "table_data"]
            if len(sibs) > 1:
                for k in sibs:
                    binding[k]["row_siblings"] = sibs

        # ── Section-aware COLUMN GROUPS (generic multi-section support) ───────
        # Pair each value column with the label column to its left, then split
        # each pair's fill (table_data) rows into VERTICAL SECTIONS delimited by
        # header rows in the label column. Works for side-by-side AND stacked
        # repeated sections (balance sheet, payslip, P&L, ...) — no hardcoding.
        _VALUE_KW = {"amount", "amt", "value", "total", "balance", "price", "cost",
                     "qty", "quantity", "debit", "credit", "$", "net", "gross", "sum"}

        def _chdr(c):
            ch = col_header.get(c)
            return ch["text"] if ch else ""

        # Distinguish a SECTION HEADER from a TOTALS row (both have label text in the
        # label column and an empty value cell → value_target structurally). Generic,
        # no document-type terms. A section header: label text is not a total-like
        # word, is not purely numeric, and is followed by empty fill rows below it.
        _TOTAL_INDICATORS = ("total", "sum", "subtotal", "grand", "net", "balance",
                             "amount due", "amount payable")

        def _is_section_header(row_idx, label_col):
            label_text = val(row_idx, label_col)
            if not label_text:
                return False
            low = label_text.lower()
            if any(ind in low for ind in _TOTAL_INDICATORS):
                return False
            if (label_text.replace(",", "").replace(".", "").replace("-", "")
                          .replace("$", "").replace("%", "").isdigit()):
                return False
            empty_below = 0
            for r in range(row_idx + 1, row_idx + 5):
                if val(r, label_col):
                    break
                empty_below += 1
            return empty_below >= 1

        _td_cols = sorted({b["col_index"] for b in binding.values()
                           if isinstance(b, dict) and b.get("role") == "table_data"
                           and isinstance(b.get("col_index"), int)})
        _usedc, _pairs = set(), []
        for c in _td_cols:                       # value columns paired with the label col to their left
            if c in _usedc or not any(kw in _chdr(c).lower() for kw in _VALUE_KW):
                continue
            lc = next((x for x in range(c - 1, -1, -1) if x in _td_cols and x not in _usedc), None)
            _pairs.append((lc, c)); _usedc.add(c)
            if lc is not None:
                _usedc.add(lc)
        for c in _td_cols:                       # leftovers: pair consecutively
            if c in _usedc:
                continue
            rc = next((x for x in _td_cols if x > c and x not in _usedc), None)
            if rc is not None:
                _pairs.append((c, rc)); _usedc.add(c); _usedc.add(rc)
            else:
                _pairs.append((c, None)); _usedc.add(c)

        column_groups, _gid = [], 0
        for (lc, vc) in _pairs:
            # A column group needs BOTH a value column and a label column to its left.
            # A value column with no label column (lc is None) is a stray column — e.g.
            # a grand-total label column like "Final Assets Total" whose header happens
            # to contain "total". Skip it so it does NOT fall back to column 0's header
            # rows and create duplicate groups.
            if vc is None or lc is None:
                continue
            fill_rows = sorted(r for r in range(max_r + 1)
                               if binding.get(f"{r},{vc}", {}).get("role") == "table_data")
            if not fill_rows:
                continue
            anchor = lc if lc is not None else 0
            # Header rows delimit the vertical sections within this pair. Use the
            # section-header test (label text that is not a total/numeric word and is
            # followed by empty fill rows) so a section header with an EMPTY value cell
            # (e.g. "Current liabilities") is detected, while totals rows are excluded.
            header_rows = sorted(r for r in range(max_r + 1)
                                 if _is_section_header(r, anchor))
            buckets = {}
            for fr in fill_rows:
                hr = max((h for h in header_rows if h < fr), default=-1)
                buckets.setdefault(hr, []).append(fr)
            for hr in sorted(buckets):
                rows = sorted(buckets[hr])
                slabel = (val(hr, anchor) if hr >= 0
                          else (_chdr(lc) if lc is not None else _chdr(vc)))
                _gid += 1
                column_groups.append({
                    "group_id": _gid,
                    "section_label": slabel or f"section_{_gid}",
                    "label_col": lc, "value_col": vc,
                    "label_col_letter": _col_letter(lc) if lc is not None else "",
                    "value_col_letter": _col_letter(vc),
                    "start_row": rows[0], "end_row": rows[-1],
                    "rows": rows,
                })
        column_groups.sort(key=lambda g: (g["start_row"], g["value_col"]))

        # Dedupe: drop any group that duplicates another's columns + row range
        # (safety net against the same header rows producing repeated groups).
        _seen_g, _dedup_g = set(), []
        for g in column_groups:
            _gk = (g["label_col"], g["value_col"], g["start_row"], g["end_row"])
            if _gk in _seen_g:
                continue
            _seen_g.add(_gk)
            _dedup_g.append(g)
        column_groups = _dedup_g

        binding["_meta"] = {
            "max_row": max_r, "max_col": max_c,
            "has_table_data": any(b.get("role") == "table_data" for b in binding.values()),
            "column_headers": {c: ch["text"] for c, ch in col_header.items()},
            "column_groups": column_groups,
        }
        return binding
    except Exception as e:
        import traceback as _tb
        print(f"[BINDING] compute_binding_map failed: {e}", flush=True)
        print(_tb.format_exc(), flush=True)
        return None


def _build_layout_prompt_parts(binding_map: dict, layout: dict, doc_text: str,
                               system_instruction: str) -> tuple:
    """
    COMPONENT 4 — build the "extract & place" prompt for layout templates. Uses the
    section-aware column_groups from the binding map so EVERY vertical section
    (side-by-side AND stacked) is described with its own explicit row range.
    """
    import re as _re
    meta           = binding_map.get("_meta", {})
    column_groups  = meta.get("column_groups", [])
    col_headers    = meta.get("column_headers", {})

    # Fallback: derive minimal groups if column_groups is unexpectedly empty.
    if not column_groups:
        td_cols = sorted({b["col_index"] for b in binding_map.values()
                          if isinstance(b, dict) and b.get("role") == "table_data"
                          and isinstance(b.get("col_index"), int)})
        for i in range(0, len(td_cols), 2):
            lc = td_cols[i]
            vc = td_cols[i + 1] if i + 1 < len(td_cols) else None
            if vc is None:
                continue
            rows = sorted(int(k.split(",")[0]) for k, b in binding_map.items()
                          if isinstance(b, dict) and b.get("role") == "table_data"
                          and b.get("col_index") == vc)
            if rows:
                column_groups.append({
                    "group_id": i // 2 + 1,
                    "section_label": str(col_headers.get(lc, col_headers.get(str(lc), "")) or f"section_{i}"),
                    "label_col_letter": _col_letter(lc), "value_col_letter": _col_letter(vc),
                    "start_row": rows[0], "end_row": rows[-1],
                })

    fixed = sorted(
        (_cell_ref(*map(int, k.split(","))), b["label"])
        for k, b in binding_map.items()
        if isinstance(b, dict) and b.get("role") == "value_target" and b.get("label")
    )

    desc = [
        "=== EXTRACTION TASK (LAYOUT MODE) ===",
        "This template provides a LAYOUT, not a fixed field list. The line-item rows "
        "are BLANK. Read the document, extract ALL line items (both the item NAME and "
        "its AMOUNT) for EACH section below, and place them into the grid. Each section "
        "has its OWN row range and columns — keep sections separate.",
        "",
    ]
    out_sections, _seen_slugs = [], {}
    _col_first_use = {}        # (label_col, value_col) -> first section that used it
    _all_cols = []             # every unique column letter used, in order
    for g in column_groups:
        sec_label = g.get("section_label") or f"section_{g.get('group_id', 0)}"
        lcl = (g.get("label_col_letter") or "").upper()
        vcl = (g.get("value_col_letter") or "").upper()
        start_ss = int(g.get("start_row", 0)) + 1     # 1-based spreadsheet rows
        end_ss   = int(g.get("end_row", 0)) + 1
        slug = _re.sub(r'[^a-z0-9]+', '_', sec_label.lower()).strip('_') or f"section_{g.get('group_id', 0)}"
        if slug in _seen_slugs:
            _seen_slugs[slug] += 1; slug = f"{slug}_{_seen_slugs[slug]}"
        else:
            _seen_slugs[slug] = 1
        for col in (lcl, vcl):
            if col and col not in _all_cols:
                _all_cols.append(col)
        desc.append(f'SECTION: "{sec_label}" (rows {start_ss} to {end_ss})')
        desc.append(f'  Column {lcl} = labels, Column {vcl} = values')
        desc.append(f'  MANDATORY: Extract EVERY individual line item found in the '
                    f'"{sec_label}" portion of the document. Each line item gets its '
                    f'own row (name into column {lcl}, amount into column {vcl}, rows '
                    f'{start_ss}-{end_ss}). The section TOTAL goes in the fixed cell '
                    f'listed below — NOT in this section\'s rows.')
        # FIX 1: when this section reuses columns from an earlier section, say so.
        prev = _col_first_use.get((lcl, vcl))
        if prev:
            desc.append(f'  NOTE: This section uses the SAME columns ({lcl}/{vcl}) as '
                        f'"{prev}". Reuse those columns — do NOT create new columns.')
        else:
            _col_first_use[(lcl, vcl)] = sec_label
        desc.append(f'  This section MUST be present in your response with at least one '
                    f'row. Do not skip any section.')
        desc.append("")
        out_sections.append((slug, lcl, vcl, start_ss))

    # FIX 1 + completeness: global directives inserted at the TOP, before sections.
    if out_sections:
        _names = ", ".join(f'"{s[0]}"' for s in out_sections)
        desc.insert(3, f'COLUMN CONSTRAINT: The ONLY valid columns in your response are: '
                       f'{", ".join(_all_cols)}. Any column not in this list is FORBIDDEN '
                       f'— do not use any other column. Multiple sections reuse the same '
                       f'columns; place each section\'s items in ITS specified columns.')
        desc.insert(4, f'There are {len(out_sections)} sections. Your layout_sections '
                       f'response MUST contain ALL of them: {_names}. Each must have '
                       f'at least one row.')
        desc.insert(5, "")

    if fixed:
        desc.append("FIXED CELLS (labels already in the template — return the value only):")
        desc += [f"  {ref} = {lbl}" for ref, lbl in fixed]
        desc.append("")

    seclines = [
        f'      "{slug}": {{ "rows": [\n'
        f'        {{"label_col": "{lcl}", "value_col": "{vcl}", "row": {first_row}, '
        f'"label": "<item name from document>", "value": "<amount>"}}\n'
        f'      ] }}'
        for slug, lcl, vcl, first_row in out_sections
    ]
    fixed_example = ", ".join(f'"{ref}": {{"value": "<amount>", "confidence": "high"}}'
                              for ref, _ in fixed[:4])

    output_format = (
        "Return ONLY valid JSON:\n{\n"
        '  "document_type": "detected type",\n'
        '  "overall_confidence": "high",\n'
        '  "documents": [{\n'
        '    "doc_index": 0,\n'
        '    "layout_sections": {\n' + ",\n".join(seclines) + "\n    },\n"
        '    "extracted_fields": {' + fixed_example + "}\n"
        "  }]\n}\n"
        "RULES:\n"
        "- label_col/value_col are column LETTERS; row is the spreadsheet row NUMBER.\n"
        "- Target a cell by combining them: column letter + row number (e.g. A + 2 = A2).\n"
        "- Extract EVERY line item present in the document — do not stop at the number\n"
        "  of available rows; add as many row objects as the document actually has.\n"
        "- Numbers: strip $ and commas. Dates: YYYY-MM-DD. Missing: \"\".\n"
        "- Put each section's items ONLY into that section's columns."
    )

    si = system_instruction + (
        "\n\nLAYOUT EXTRACTION MODE:\n"
        "The template's line-item rows are intentionally blank. Read the document and "
        "reconstruct each section's line items (name + amount), then place them into "
        "the designated label/value columns. Extract the document's ACTUAL content — "
        "do not invent rows and do not omit any line item."
    )
    doc_text_use = _smart_truncate(doc_text, "table", {})
    user_prompt = (
        "\n".join(desc)
        + "\n=== DOCUMENT TEXT ===\n"
        + (doc_text_use if doc_text_use else "(See document image)")
        + "\n\n=== OUTPUT FORMAT ===\n" + output_format + "\n"
    )
    print(f"[LAYOUT] prompt built: {len(column_groups)} section group(s), "
          f"{len(fixed)} fixed cell(s)", flush=True)
    return si, user_prompt


def _convert_extracted_fields_to_layout(extracted_fields: dict, binding_map: dict) -> dict:
    """
    ISSUE 1 fallback — when a weaker model returns extracted_fields instead of the
    required layout_sections format, reconstruct layout_sections from the flat cell
    map using the binding map's column groups (pairing each value-column cell with
    the label-column cell on the same row).
    """
    import re as _re
    if (not isinstance(extracted_fields, dict) or not extracted_fields
            or not isinstance(binding_map, dict)):
        return {}
    col_headers = binding_map.get("_meta", {}).get("column_headers", {})

    def hdr(c):
        return str(col_headers.get(c, col_headers.get(str(c), "")) or "")

    VALUE_KW = {"amount", "amt", "value", "total", "balance", "price", "cost",
                "qty", "quantity", "debit", "credit", "$", "net", "gross", "sum"}
    td_cols = sorted({b["col_index"] for b in binding_map.values()
                      if isinstance(b, dict) and b.get("role") == "table_data"
                      and isinstance(b.get("col_index"), int)})

    def is_val(c):
        return any(kw in hdr(c).lower() for kw in VALUE_KW)

    used, groups = set(), []
    for c in td_cols:
        if c in used or not is_val(c):
            continue
        lc = next((x for x in range(c - 1, -1, -1) if x in td_cols and x not in used), None)
        groups.append((lc, c)); used.add(c)
        if lc is not None:
            used.add(lc)
    for c in td_cols:
        if c in used:
            continue
        rc = next((x for x in td_cols if x > c and x not in used), None)
        if rc is not None:
            groups.append((c, rc)); used.add(c); used.add(rc)

    def _v(x):
        return str((x.get("value", "") if isinstance(x, dict) else x) or "").strip()

    parsed = {}
    for ref, raw_v in extracted_fields.items():
        m = _re.match(r'^([A-Za-z]+)(\d+)$', str(ref).strip())
        if not m:
            continue
        col = 0
        for ch in m.group(1).upper():
            col = col * 26 + (ord(ch) - 64)
        parsed[(int(m.group(2)) - 1, col - 1)] = _v(raw_v)

    out = {}
    for lc, vc in groups:
        if vc is None:
            continue
        slug = _re.sub(r'[^a-z0-9]+', '_', (hdr(lc) or hdr(vc) or "section").lower()).strip('_') or "section"
        rows = []
        for (row0, col0), v in sorted(parsed.items()):
            if col0 != vc or not v:
                continue
            label = parsed.get((row0, lc), "") if lc is not None else ""
            rows.append({"label_col": _col_letter(lc) if lc is not None else "",
                         "value_col": _col_letter(vc), "row": row0 + 1,
                         "label": label, "value": v})
        if rows:
            out[slug] = {"rows": rows}
    return out


def _build_vision_prompt(template_data: dict, doc_text: str = "") -> tuple:
    """
    Build extraction prompt split into (system_instruction, user_prompt).

    system_instruction = registry expert persona (stable, doc-type specific).
    user_prompt = template structure + page anchors + self-verification + doc text.
    """
    doc_type      = template_data.get("doc_type", "other")
    regions       = template_data.get("regions", {})
    layout        = template_data.get("layout", {})
    primary_mode  = regions.get("primary_mode", "form_kv")
    table_regions = regions.get("table_regions", [])

    # System instruction — expert persona from registry (or unguided when type is unknown)
    if primary_mode == "unguided" and doc_type in ("other", "", None):
        system_instruction = _get_unguided_prompt()
    else:
        system_instruction = _get_system_prompt(doc_type)
    table_rules = _get_table_rules(doc_type)
    if table_rules and primary_mode in ("table", "mixed"):
        system_instruction += f"\n\nTABLE RULES:\n{table_rules}"

    # Bug 5b: inject critical financial accuracy rule when section-context risk is present
    if regions.get("needs_section_context"):
        system_instruction += (
            "\n\nCRITICAL FINANCIAL ACCURACY RULE:\n"
            "This document has MULTIPLE SECTIONS, each with its own subtotal row. "
            "Generic column headers like 'Total' or 'Amount' appear in EVERY section.\n"
            "You MUST match each value to its CORRECT section — do NOT use the wrong "
            "section's total.\n"
            "Rules:\n"
            "1. Read the SECTION HEADER (e.g. 'Current Assets', 'Non-Current Assets', "
            "'Operating Expenses') to identify which section each row belongs to.\n"
            "2. The 'Total Current Assets' row belongs ONLY to the Current Assets section.\n"
            "   Do NOT use that value for 'Total Non-Current Assets' or 'Total Assets'.\n"
            "3. Extract every row in POSITIONAL ORDER — the section header is the "
            "   primary key for resolving ambiguity.\n"
            "4. If you are unsure which section a value belongs to, use the row's "
            "   position in the document as a tiebreaker.\n"
            "FINANCIAL ERROR WARNING: Mixing up section totals (e.g. Total Current "
            "Assets vs Total Assets) is a critical error with legal and financial "
            "consequences. Verify EACH total against its section header before writing it."
        )
        print(f"[BUG5] CRITICAL FINANCIAL ACCURACY RULE injected for {doc_type}", flush=True)

    # Issue 1: inject PARALLEL COLUMN GROUPS RULE when the template has multiple
    # independent label/value column bands on the same row range.
    parallel_groups = regions.get("parallel_column_groups", [])
    if parallel_groups:
        group_lines = [
            f"  GROUP {pg['group_id']} \"{pg['section_label']}\": "
            f"labels in column {pg['label_col_letter']}, "
            f"amounts into column {pg['value_col_letter']}"
            for pg in parallel_groups
        ]
        system_instruction += (
            "\n\nPARALLEL COLUMN GROUPS RULE:\n"
            "This template has SIDE-BY-SIDE independent column groups. "
            "Each group occupies the SAME rows but DIFFERENT columns.\n"
            + "\n".join(group_lines) + "\n"
            "CRITICAL RULES:\n"
            "1. Fill ALL groups. Never leave any group empty.\n"
            "2. Each group's amounts go ONLY into that group's designated value column.\n"
            "3. Read the document's LEFT section for GROUP 1 "
            "and RIGHT/NEXT section for GROUP 2.\n"
            "4. Total rows: match EACH total to its OWN group's value column.\n"
            "   'Total Current Assets' → GROUP 1 value column.\n"
            "   'Total Non-Current Assets' → GROUP 2 value column.\n"
            "5. Do NOT copy the same value into multiple groups."
        )
        print(
            f"[PROMPT] PARALLEL COLUMN GROUPS RULE injected "
            f"({len(parallel_groups)} groups)",
            flush=True,
        )

    # ── COMPONENT 6: layout-based extraction trigger ─────────────────────────
    # When the binding map shows table_data cells (column headers but blank,
    # unlabeled line-item rows), switch to the "extract & place" layout prompt.
    # Field-based templates (explicit label→value pairs) keep the logic below.
    binding_map = template_data.get("binding_map")
    if isinstance(binding_map, dict) and binding_map.get("_meta", {}).get("has_table_data"):
        try:
            print("[PROMPT] layout-based extraction selected (table_data present)", flush=True)
            return _build_layout_prompt_parts(binding_map, layout, doc_text, system_instruction)
        except Exception as _lp_e:
            print(f"[LAYOUT] prompt build failed ({_lp_e}) — using field-based prompt", flush=True)

    # Smart text truncation
    doc_text_use = _smart_truncate(doc_text, primary_mode, regions)

    fields_description      = _build_fields_description(regions, layout)
    extraction_instructions = _build_extraction_instructions(regions, primary_mode, "")
    output_format           = _build_output_format(regions, primary_mode)
    verify_block            = _build_verification_block(primary_mode, table_regions)

    # Build page anchor map — tells AI which page each table lives on
    # Derived purely from scanning the document text for column header names
    # Zero hardcoding — works for any table with any column names
    page_anchor_block = ""
    if table_regions and doc_text and "--- PAGE BREAK ---" in doc_text:
        page_map = _build_page_anchor_map(doc_text, table_regions)
        if page_map:
            anchor_lines = ["=== PAGE LOCATION OF EACH TABLE ==="]
            anchor_lines.append(
                "The document has multiple pages. Each table's data is on a specific page."
            )
            anchor_lines.append(
                "This is determined by where the column headers appear in the text."
            )
            for section, page_num in page_map.items():
                tbl = next((t for t in table_regions
                           if t.get("section_label") == section), None)
                cols = tbl.get("column_names", []) if tbl else []
                col_str = f"({', '.join(cols[:3])})" if cols else ""
                anchor_lines.append(
                    f"  Table \"{section}\" {col_str}: "
                    f"data is on PAGE {page_num} of the document"
                )
            anchor_lines.append(
                "Extract rows for EACH table from its indicated page."
            )
            anchor_lines.append(
                "Do not stop at page 1 if a table is indicated on page 2 or later."
            )
            page_anchor_block = "\n".join(anchor_lines)

    user_prompt = f"""=== EXTRACTION TASK ===

{fields_description}

=== INSTRUCTIONS ===
{extraction_instructions}

{page_anchor_block + chr(10) if page_anchor_block else ""}=== SELF-VERIFICATION (run this before returning) ===
{verify_block}

=== RULES ===
1. MATCH BY MEANING: Labels in the template and the document will differ.
   Match by concept, not exact text.
   Examples: "Rcpt No"="Receipt Number", "Amt"="Amount", "Inv Ref"="Invoice Reference"
   "No of Emp"="Number of Employees", "Fed"="Federal", "YTD"="Year to Date"
   "EIN"="Employer ID Number", "SSN"="Social Security Number"
2. BLANK ROWS IN TEMPLATE are placeholders only — they do NOT limit row count.
   If the document has 15 rows, return 15 rows. If it has 2, return 2.
3. NUMBERS: Remove $ £ € ₹ and commas. "(2.85)" means -2.85.
4. DATES: Always YYYY-MM-DD format.
5. MISSING VALUES: Use "" — never "N/A", "null", or "not found".
6. ALL PAGES: Extract from every page. Never stop at page 1.
7. HEADERS ARE NOT DATA: Section label rows and column header rows are
   template structure only. Never include them as data rows in table_rows.

=== DOCUMENT TEXT ===
{doc_text_use if doc_text_use else "(See document image)"}

=== OUTPUT FORMAT ===
{output_format}
"""
    # Log the output format section so we can verify AI is getting correct instructions
    print(f"[PROMPT] mode={primary_mode} n_tables={len(table_regions)} "
          f"targets={len(regions.get('explicit_targets',[]))} "
          f"kv={len(regions.get('kv_pairs',[]))}", flush=True)
    if table_regions:
        for i, tr in enumerate(table_regions):
            print(f"[PROMPT] table_{i+1}: '{tr.get('section_label','')}' "
                  f"cols={tr.get('column_names',[])} "
                  f"key={'table_'+str(i+1)+'_rows'}", flush=True)
    # Log key section of output format to verify separate arrays are being sent
    fmt_preview = output_format[:300].replace('\n',' ')
    print(f"[PROMPT] output_format preview: {fmt_preview}", flush=True)

    return system_instruction, user_prompt


def _build_verification_block(primary_mode: str, table_regions: list) -> str:
    """
    Build a generic self-verification checklist the AI runs before returning.
    No hardcoded numbers, names, or document-specific logic.
    Works for any template structure.
    """
    checks = []

    # Universal checks for all modes
    checks.append(
        "STEP 1 — COUNT CHECK:\n"
        "  Go back to the document text.\n"
        "  Count every data row that exists in EACH table section.\n"
        "  Compare that count to the rows you are about to return.\n"
        "  If your row count is LESS than what the document contains — add the missing rows.\n"
        "  Common mistake: stopping at page 1 when data continues on page 2."
    )

    checks.append(
        "STEP 2 — COMPLETENESS CHECK:\n"
        "  For every table header you identified in the template:\n"
        "    - Did you extract rows for it? If any table has zero rows, that is wrong.\n"
        "    - Scan the document text specifically for that section and add any missing rows.\n"
        "  For every form field marked for extraction:\n"
        "    - Is the value filled? If blank, scan the document again for that label."
    )

    checks.append(
        "STEP 3 — FORMAT CHECK:\n"
        "  For every numeric value you extracted:\n"
        "    - Remove currency symbols ($ £ € ₹) and thousand-separator commas\n"
        "    - Convert accounting negatives: (1,245.00) → -1245.00\n"
        "    - Expand suffixes: 8.41K → 8410, 1.2M → 1200000\n"
        "  For every date value:\n"
        "    - Convert to YYYY-MM-DD\n"
        "    - 'April 30, 2024' → '2024-04-30', '30/04/2024' → '2024-04-30'\n"
        "  For every text value:\n"
        "    - Remove line breaks within a value\n"
        "    - Trim leading/trailing whitespace"
    )

    checks.append(
        "STEP 4 — CROSS-CHECK:\n"
        "  Pick 3 random values you extracted.\n"
        "  Find each one in the document text and confirm it matches exactly.\n"
        "  If any value does not match — correct it before returning."
    )

    if table_regions:
        checks.append(
            "STEP 5 — TABLE INTEGRITY:\n"
            "  For each table, verify:\n"
            "    - Every row has values in ALL columns (no empty columns mid-row)\n"
            "    - No row is a header row, section title, or total/subtotal row\n"
            "    - Rows from different tables are not mixed together\n"
            "    - If a table spans a page break, rows from BOTH pages are included"
        )

    return "\n\n".join(checks)


def _build_fields_description(regions: dict, layout: dict) -> str:
    """
    Build a complete, unambiguous description of the template structure for the AI.
    Handles: multiple tables, two-column layouts, mixed mode, any position.
    """
    lines = []
    primary_mode = regions.get("primary_mode", "form_kv")
    table_regions = regions.get("table_regions", [])
    explicit_targets = regions.get("explicit_targets", [])
    kv_pairs = regions.get("kv_pairs", [])
    two_col_pairs = regions.get("two_col_pairs", [])

    # ── PARALLEL COLUMN GROUPS (takes priority over flat kv_pairs) ───────────────
    parallel_groups = regions.get("parallel_column_groups", [])
    if parallel_groups:
        # Build the target list DIRECTLY from the template grid (not from
        # pg["items"], which can be incomplete when empty value cells aren't
        # stored as grid entries). For each group's column band, list EVERY row
        # that has a label in the label column whose value cell is empty/extract
        # — that is an extraction target. This guarantees Gemini is told exactly
        # which cells to fill instead of guessing.
        grid_cells = layout.get("cells", {})

        def _grid_val(row, col):
            c = grid_cells.get(f"{row},{col}")
            return str(c.get("value") or "").strip() if isinstance(c, dict) else ""

        def _is_target_value_cell(row, col):
            """True when the value cell is empty/missing or explicitly an
            extract target (i.e. NOT a static header like 'Amount')."""
            c = grid_cells.get(f"{row},{col}")
            if not isinstance(c, dict):
                return True                      # not stored → empty → target
            if c.get("extractTarget"):
                return True
            return not str(c.get("value") or "").strip()

        # ── Pass 1: collect every extraction target across all groups, tracking
        #    the SECTION each cell falls under (the nearest header row above it,
        #    i.e. the most recent label whose value cell is a static header). ──
        groups_targets = []   # [(gid, group_label, value_col_letter, [(ref, label, section)])]
        label_counts = {}     # label_text -> occurrences across the whole template
        for pg in parallel_groups:
            gid       = pg["group_id"]
            slbl      = pg["section_label"]
            label_col = pg["label_col"]
            value_col = pg["value_col"]
            v_c       = pg["value_col_letter"]

            label_rows = sorted({
                int(k.split(",")[0])
                for k in grid_cells
                if "," in k
                and k.split(",")[1].isdigit()
                and int(k.split(",")[1]) == label_col
            })

            current_section = slbl
            tgts = []
            for r in label_rows:
                label = _grid_val(r, label_col)
                if not label:
                    continue
                if not _is_target_value_cell(r, value_col):
                    # static header cell (e.g. "Amount") → this row names a section
                    current_section = label
                    continue
                ref = _cell_ref(r, value_col)
                tgts.append((ref, label, current_section))
                label_counts[label] = label_counts.get(label, 0) + 1
            groups_targets.append((gid, slbl, v_c, tgts))

        # ── Pass 2: emit the listing. Ambiguous labels (the same label text used
        #    more than once anywhere in the template) get an inline section hint;
        #    unique labels need none. ──
        lines.append("=== EXTRACTION TARGETS — fill EXACTLY these cells ===")
        lines.append(
            "This template has INDEPENDENT side-by-side column groups. Below is "
            "the COMPLETE list of every cell you must fill, each with the label "
            "that sits next to it. Fill EVERY listed cell."
        )
        lines.append("")

        all_refs = []
        for gid, slbl, v_c, tgts in groups_targets:
            lines.append(f"GROUP {gid} — \"{slbl}\" (Column {v_c}):")
            if tgts:
                for ref, label, section in tgts:
                    hint = (f" (from {section} section)"
                            if label_counts.get(label, 0) > 1 and section else "")
                    lines.append(f"  {ref} = {label}{hint}")
                    all_refs.append(ref)
            else:
                lines.append("  (no extraction targets detected in this group)")
            lines.append("")

        # FIX 1: per-cell matching rule replaces the contradictory per-group
        # "Put ONLY values from the <section> section" footer.
        lines.append(
            "Each cell reference above is mapped to its exact source label. "
            "Extract the value matching that EXACT label from the document. "
            "Do not use section membership to decide — use the label name only. "
            "If a label appears multiple times in the document, use the value "
            "from the row that matches the label text exactly."
        )
        lines.append("")

        # FIX 3: strong, explicit closing directive with a dynamically-generated
        # key list (never hardcoded).
        lines.append(
            "MANDATORY: Your response must contain extracted_fields with EXACTLY "
            f"these keys: [{', '.join(all_refs)}]. "
            "Every key must be present. Use \"\" for any value not found. "
            "Match each key to its label above — ignore section boundaries, use "
            "label text only."
        )
        lines.append("")

    # ── FORM FIELDS ────────────────────────────────────────────────────────────
    elif explicit_targets:
        lines.append("=== FORM FIELDS (marked Extract here) ===")
        lines.append("Fill each cell reference with the matching value from the document.")
        lines.append("Use the cell reference as the key in extracted_fields.")
        for t in explicit_targets:
            lines.append(f"  [{t['ref']}] {t['label']}")
        lines.append("")

    elif kv_pairs and not table_regions:
        lines.append("=== FORM FIELDS (label-value pairs) ===")
        lines.append("Labels are in the left column. Fill the right column.")
        for kv in kv_pairs[:30]:  # limit to avoid token waste
            lines.append(f"  Label: \"{kv['label']}\" -> Fill cell: {kv['value_ref']}")
        lines.append("")

    if two_col_pairs:
        lines.append("=== TWO-COLUMN FORM LAYOUT ===")
        lines.append("This template has label-value pairs on BOTH the left AND right sides of each row.")
        lines.append("Extract BOTH sides. Do NOT treat this as a table.")
        for tc in two_col_pairs[:20]:
            lines.append(f"  LEFT:  [{tc['left_value_ref']}] = \"{tc['left_label']}\"")
            lines.append(f"  RIGHT: [{tc['right_value_ref']}] = \"{tc['right_label']}\"")
        lines.append("")

    # ── TABLES ─────────────────────────────────────────────────────────────────
    # FIX 1: never render table instructions for parallel-group templates,
    # even if table_regions somehow survived region analysis.
    if table_regions and primary_mode != "parallel_groups":
        n = len(table_regions)

        # Group by row to detect same-row tables
        same_row = {}
        for tr in table_regions:
            same_row.setdefault(tr["header_row"], []).append(tr)
        has_same_row = any(len(v) > 1 for v in same_row.values())

        if n == 1:
            lines.append("=== TABLE ===")
        else:
            lines.append(f"=== {n} TABLES ===")
            if has_same_row:
                lines.append(
                    "IMPORTANT: Some tables appear SIDE BY SIDE on the same row.\n"
                    "They are separated by empty columns. Use the column range to\n"
                    "identify which data belongs to which table."
                )
            lines.append("")

        # Bug 5c: detect generic column headers that repeat across sections
        needs_section_ctx = regions.get("needs_section_context", False)

        for i, tr in enumerate(table_regions):
            section   = tr.get("section_label", f"Table {i+1}")
            col_start = chr(ord('A') + tr.get("start_col", 0))
            col_end   = chr(ord('A') + min(tr.get("end_col", 0), 25))
            col_range = f"columns {col_start} to {col_end}"

            lines.append(f"TABLE {i+1}: \"{section}\"")
            lines.append(f"  Position: row {tr['header_row']+1}, {col_range}")
            lines.append(f"  Columns:  {', '.join(tr['column_names'])}")
            lines.append(
                f"  Extract:  ALL data rows for this table from the document.\n"
                f"            There is NO row limit — blank rows in the template\n"
                f"            are just placeholders. Extract every row that exists."
            )
            if tr.get("col_range") and n > 1:
                lines.append(
                    f"  NOTE: Data for this table is in {col_range} of the document.\n"
                    f"        Do not mix it with data from other tables."
                )
            if needs_section_ctx:
                lines.append(
                    f"  SECTION ACCURACY: This table has generic column names. "
                    f"Each row belongs to a specific SECTION of the document "
                    f"(identified by a section header above it). "
                    f"Match every value to its CORRECT section — positional order matters."
                )
            lines.append("")

    if not lines:
        plain_text_desc = regions.get("plain_text_description", "")
        if plain_text_desc:
            # Bug 4 + Bug 1: template has plain-text description only
            lines.append("=== UNGUIDED EXTRACTION (template description provided) ===")
            lines.append(f"Template description: {plain_text_desc}")
            lines.append("Extract all fields relevant to this document based on the description above.")
        elif primary_mode == "unguided":
            lines.append("=== UNGUIDED EXTRACTION ===")
            lines.append("No template structure was detected. Extract all fields from the document.")
        else:
            lines.append("=== AUTO-EXTRACT MODE ===")
            lines.append("The template has labelled fields. Match each label to its value in the document.")
            lines.append("Use cell references as keys.")

    # ── TRANSPOSED TABLES ──────────────────────────────────────────────────────
    transposed = regions.get("transposed_tables", [])
    # FIX 1: never render transposed-table instructions for parallel-group templates.
    if transposed and primary_mode != "parallel_groups":
        lines.append("=== TRANSPOSED TABLE (horizontal layout) ===")
        lines.append("This template has a HORIZONTAL table where:")
        lines.append("  - Row LABELS are in the leftmost column (column A)")
        lines.append("  - Each subsequent column contains one complete record")
        lines.append("  - Extract each column as a separate record/row")
        lines.append("")
        for tt in transposed:
            lines.append(f"  Row labels: {', '.join(tt['row_names'][:6])}")
            lines.append(f"  Data columns: {len(tt['data_cols'])} records")
            lines.append("  Return each column as one object in table_rows")
        lines.append("")

    return "\n".join(lines)


def _build_extraction_instructions(regions: dict, primary_mode: str,
                                    table_rules: str) -> str:
    """Build extraction instructions covering all known failure scenarios."""
    instructions = []
    table_regions = regions.get("table_regions", [])
    n_tables = len(table_regions)

    # ── UNIVERSAL RULES (always included) ─────────────────────────────────────
    instructions.append("""=== UNIVERSAL EXTRACTION RULES ===
1. MATCH BY MEANING: Template labels and document labels will differ.
   Match by concept, not exact text.
   "Rcpt No" = "Receipt Number", "Amt Rcvd" = "Amount Received",
   "Inv Ref" = "Invoice Reference", "Pmt Mthd" = "Payment Method".
   An abbreviation in the template refers to the full label in the document.

2. SECTION HEADERS ARE NOT DATA: Cells like "Bank Information", "Payment Details",
   "Earning Table", "Deduction Table" are visual section organizers.
   NEVER extract a section header text as a field value.

3. MULTI-PAGE DOCUMENTS: The document may span 2 or more pages.
   Extract data from ALL pages. Tables may continue across page breaks.
   A row on page 2 is just as valid as a row on page 1.
   Do NOT stop extracting at the end of page 1.

4. NUMBERS: Strip all currency symbols ($, £, €, ₹) and commas.
   Negative values: "(2.85)" means -2.85. "8.41K" means 8410.
   Page-break splits: if "7,513.0" appears on page 1 and "3" on page 2,
   the correct value is 7513.03 — combine them.

5. DATES: Always YYYY-MM-DD. "Feb 10, 2024" -> "2024-02-10".

6. EMPTY CELLS: If a field has no corresponding value, use "".
   Never invent values. Never copy a label as a value.

7. CELL REFERENCES: Use the exact cell reference from the template (B3, D10, etc.)
   as the key in extracted_fields. The label is for your understanding only.

8. FORMULA CELLS: If the template shows =SUM(...), calculate and return the number.""")

    # ── MODE-SPECIFIC INSTRUCTIONS ─────────────────────────────────────────────
    if primary_mode == "parallel_groups":
        parallel_groups = regions.get("parallel_column_groups", [])
        gmap = "\n".join(
            f'  GROUP {pg["group_id"]} "{pg["section_label"]}": '
            f'value column = {pg["value_col_letter"]}'
            for pg in parallel_groups
        )
        instructions.append(
            "=== PARALLEL COLUMN GROUP EXTRACTION ===\n"
            "This template has multiple INDEPENDENT column groups side by side.\n"
            "You MUST extract values for EVERY group listed below.\n\n"
            + gmap + "\n\n"
            "RULES:\n"
            "1. Scan the document for EACH group's section independently.\n"
            "2. Put each group's amounts into its designated value column ONLY.\n"
            "3. Zero filled cells in any group = extraction failure.\n"
            "4. Match totals to their group:\n"
            "   'Total Current Assets' → GROUP 1 value column.\n"
            "   'Total Non-Current Assets' → GROUP 2 value column.\n"
            "5. If the document has content in two side-by-side columns, "
            "read BOTH columns independently."
        )

    elif primary_mode == "unguided":
        # Bug 1: no regions found — use doc_type persona only
        instructions.append(
            "=== UNGUIDED EXTRACTION ===\n"
            "No template layout could be parsed for this document. "
            "Extract ALL fields and values present in the document based on "
            "your domain expertise for this document type.\n"
            "Return every label-value pair you can identify. "
            "Use the label text as the key in extracted_fields.\n"
            "For any tables, extract all rows into table_rows."
        )

    elif primary_mode in ("form_with_targets", "form_kv"):
        instructions.append("""=== FORM EXTRACTION ===
Fill each marked field with exactly ONE value from the document.
Blank rows between sections are visual spacers — skip them.
Merged header cells are section titles — do not extract their text as values.
If a field appears in both a section header and as a real value,
extract the real value (the data, not the heading).""")

    elif primary_mode == "table":
        instructions.append("""=== TABLE EXTRACTION ===
Extract EVERY data row from the document table.
The table may start anywhere on the page — find the column headers.
Skip: the header row itself, subtotal rows, total rows, blank rows.
Include ALL data rows including those on page 2.""")
        if table_rules:
            instructions.append(f"Document-specific rules:\n{table_rules}")

    elif primary_mode == "mixed":
        if n_tables == 1:
            instructions.append("""=== MIXED MODE: FORM FIELDS + TABLE ===
PART 1 - FORM FIELDS: Fill each labelled extraction cell with its value.
PART 2 - TABLE: Extract every data row from the table section.
Both parts are equally important. Do not skip either.""")
        else:
            # Bug 3: multiple tables — use SEPARATE array keys matching _build_output_format
            # The old code used a single table_rows array + "Table" field, which
            # contradicted the output format and caused AI to misroute rows.
            table_key_map = []
            for i, tr in enumerate(table_regions):
                name  = tr.get("section_label", f"Table {i+1}")
                key   = re.sub(r'[^a-z0-9]', '_', name.lower()).strip('_') + "_rows"
                col_s = chr(ord('A') + tr.get("start_col", 0))
                col_e = chr(ord('A') + min(tr.get("end_col", 0), 25))
                cols  = ", ".join(tr.get("column_names", []))
                table_key_map.append(
                    f'  JSON key "{key}" → table "{name}" '
                    f'(columns: {cols}, template cols {col_s}–{col_e})'
                )
            instructions.append(
                f"=== MIXED MODE: FORM FIELDS + {n_tables} SEPARATE TABLES ===\n"
                f"PART 1 - FORM FIELDS: Fill each labelled extraction cell.\n"
                f"PART 2 - {n_tables} TABLES: Each table gets its OWN JSON array key.\n\n"
                f"CRITICAL — use these EXACT JSON array key names:\n"
                + "\n".join(table_key_map) + "\n\n"
                f"RULES:\n"
                f"- DO NOT use a single 'table_rows' key — use the separate keys above\n"
                f"- Each row goes into its table's OWN array key only\n"
                f"- Extract EVERY row from EVERY table — zero rows from any table is WRONG\n"
                f"- Blank rows in the template are placeholders, not row limits\n"
                f"- See PAGE LOCATION section for which page each table is on\n"
                f"- Do NOT include column header rows as data rows"
            )
        if table_rules:
            instructions.append(f"Document-specific rules:\n{table_rules}")

    # ── TWO-COLUMN REMINDER ────────────────────────────────────────────────────
    if regions.get("two_col_pairs"):
        instructions.append("""=== TWO-COLUMN LAYOUT REMINDER ===
Extract BOTH left-side and right-side values on each row.
This is a form, not a table. Each row has 2 label-value pairs.
Left label and right label are independent fields.""")

    return "\n\n".join(instructions)


def _build_output_format(regions: dict, primary_mode: str) -> str:
    """
    Build the JSON output format specification.
    Handles: single table, multiple tables, form only, mixed mode.
    """
    table_regions = regions.get("table_regions", [])
    n_tables = len(table_regions)
    has_form = primary_mode != "table"

    # Build example extracted_fields
    field_example = '"B3": {"value": "extracted value", "confidence": "high"}'

    # Build table examples
    def table_example(tr):
        cols = tr.get("column_names", ["Col1", "Col2"])[:4]
        return "{" + ", ".join(f'"{c}": "value"' for c in cols) + "}"

    if primary_mode == "unguided":
        # Bug 1: no template regions — generic free-form output
        return f"""Return ONLY valid JSON:
{{
  "document_type": "detected type",
  "overall_confidence": "medium",
  "document_count": 1,
  "documents": [{{
    "doc_index": 0,
    "doc_hint": "brief description",
    "extracted_fields": {{
      "Field Name 1": {{"value": "extracted value", "confidence": "high"}},
      "Field Name 2": {{"value": "extracted value", "confidence": "medium"}}
    }},
    "table_rows": [],
    "notes": ""
  }}]
}}
RULES:
- extracted_fields keys = the label/field name from the document
- Include every field you can identify
- Numbers: no $ or commas. Dates: YYYY-MM-DD."""

    elif primary_mode == "table" and n_tables == 1:
        ex = table_example(table_regions[0])
        return f"""Return ONLY valid JSON:
{{
  "document_type": "detected type",
  "overall_confidence": "high",
  "document_count": 1,
  "documents": [{{
    "doc_index": 0,
    "doc_hint": "brief description",
    "table_rows": [{ex}, {ex}],
    "row_count": 2,
    "notes": ""
  }}]
}}
RULES: table_rows = one object per data row. Skip headers, totals, blank rows. Numbers: no $ or commas."""

    elif primary_mode == "mixed" and n_tables == 1:
        ex = table_example(table_regions[0])
        return f"""Return ONLY valid JSON:
{{
  "document_type": "detected type",
  "overall_confidence": "high",
  "document_count": 1,
  "documents": [{{
    "doc_index": 0,
    "doc_hint": "brief description",
    "extracted_fields": {{{field_example}}},
    "table_rows": [{ex}],
    "row_count": 1,
    "notes": ""
  }}]
}}
RULES:
- extracted_fields: cell refs as keys (B3, D10), one value per field
- table_rows: one object per data row, using exact column names
- Numbers: no $ or commas. Dates: YYYY-MM-DD."""

    elif primary_mode == "mixed" and n_tables > 1:
        # Multiple tables — use ONE table_rows array with a "Table" column
        # Show a SEPARATE example row for each table with its OWN columns
        # This prevents the AI from merging columns across tables
        section_names = [tr.get("section_label", f"Table {i+1}")
                        for i, tr in enumerate(table_regions)]

        # Build separate array key and example for each table
        # Using SEPARATE arrays per table is the ONLY reliable way to prevent
        # the AI from merging columns across tables into one row structure
        table_blocks = []
        table_keys   = []
        for tr in table_regions:
            name    = tr.get("section_label", "")
            cols    = tr.get("column_names", [])
            col_s   = chr(ord('A') + tr.get("start_col", 0))
            col_e   = chr(ord('A') + min(tr.get("end_col", 0), 25))
            # Make a safe JSON key from section label
            key     = re.sub(r'[^a-z0-9]', '_', name.lower()).strip('_') + "_rows"
            table_keys.append(key)
            ex_row  = {c: "value" for c in cols}
            table_blocks.append(
                f'    "{key}": [\n'
                f'      {json.dumps(ex_row)}\n'
                f'    ]  /* ALL rows from "{name}" table (cols {col_s}-{col_e}) */'
            )

        blocks_str = ",\n".join(table_blocks)

        table_rules_str = []
        for i, (tr, key) in enumerate(zip(table_regions, table_keys)):
            name = tr.get("section_label", f"Table {i+1}")
            cols = ", ".join(tr.get("column_names", []))
            col_s = chr(ord('A') + tr.get("start_col", 0))
            col_e = chr(ord('A') + min(tr.get("end_col", 0), 25))
            table_rules_str.append(
                f'  "{key}": rows from "{name}" section '
                f'(columns: {cols}, template cols {col_s}-{col_e})'
            )

        return f"""Return ONLY valid JSON:
{{
  "document_type": "detected type",
  "overall_confidence": "high",
  "document_count": 1,
  "documents": [{{
    "doc_index": 0,
    "doc_hint": "brief description",
    "extracted_fields": {{{field_example}}},
{blocks_str},
    "notes": ""
  }}]
}}
RULES FOR {len(table_regions)} SEPARATE TABLES:
Each table gets its OWN array key. Do NOT use a single "table_rows" array.
Do NOT merge columns from different tables into the same row.

{chr(10).join(table_rules_str)}

For each table array:
- Extract ALL data rows from that table section in the document
- Include rows from ALL pages (document may continue on page 2+)
- Each row uses ONLY that table's columns — do not add columns from other tables
- Blank rows in the template are placeholders, not row limits
- Numbers: no $ or commas. Dates: YYYY-MM-DD
- Do NOT include header rows or section title rows as data rows"""

    else:
        # Form only
        return f"""Return ONLY valid JSON:
{{
  "document_type": "detected type",
  "overall_confidence": "high",
  "document_count": 1,
  "documents": [{{
    "doc_index": 0,
    "doc_hint": "brief description",
    "extracted_fields": {{
      {field_example},
      "D3": {{"value": "another value", "confidence": "high"}}
    }},
    "notes": ""
  }}]
}}
RULES:
- extracted_fields keys MUST be cell references (B3, D10, etc.)
- Include every marked field, even if value is ""
- Numbers: no $ or commas. Dates: YYYY-MM-DD."""


# ==============================================================================
# PDFPLUMBER VALIDATION LAYER
# ==============================================================================

def _validate_with_pdfplumber(extracted_fields: dict, doc_text: str,
                               table_rows: list = None) -> dict:
    """
    Cross-check AI-extracted values against pdfplumber text.

    For each extracted value:
      - Search for it (or a close match) in the pdfplumber text
      - If found: keep confidence as-is or upgrade to high
      - If not found: downgrade confidence to low, flag for review

    This is the safety net against AI hallucinations.
    A value the AI invented will not appear anywhere in the pdfplumber text.
    A value the AI correctly extracted will appear in the text.
    """
    if not doc_text:
        return {"validated": extracted_fields, "flagged": [], "confidence_map": {}}

    doc_text_lower = doc_text.lower()
    validated = {}
    flagged = []
    confidence_map = {}

    for ref, field_data in extracted_fields.items():
        if isinstance(field_data, dict):
            value = str(field_data.get("value", "")).strip()
            ai_confidence = field_data.get("confidence", "medium")
        else:
            value = str(field_data).strip()
            ai_confidence = "medium"

        if not value or value == "":
            validated[ref] = {"value": "", "confidence": "high", "validated": True}
            continue

        # Check if value appears in pdfplumber text
        found = _check_value_in_text(value, doc_text_lower)

        if found:
            # Confirmed - upgrade to high confidence
            final_confidence = "high"
            validated[ref] = {"value": value, "confidence": final_confidence,
                               "validated": True, "pdfplumber_confirmed": True}
        else:
            # Not found - this might be hallucinated or normalised
            # Check if a variant of the value appears (e.g. normalised date)
            variant_found = _check_value_variants(value, doc_text)

            if variant_found:
                final_confidence = "medium"
                validated[ref] = {"value": value, "confidence": final_confidence,
                                   "validated": True, "pdfplumber_confirmed": "variant"}
            else:
                # Value not found at all - flag for review
                final_confidence = "low"
                validated[ref] = {"value": value, "confidence": final_confidence,
                                   "validated": False, "pdfplumber_confirmed": False,
                                   "needs_review": True}
                flagged.append({"ref": ref, "value": value, "reason": "not found in document text"})

        confidence_map[ref] = final_confidence

    # Validate table rows
    validated_rows = []
    if table_rows:
        for row in table_rows:
            validated_row = {}
            for col, val in row.items():
                val_str = str(val).strip()
                if not val_str:
                    validated_row[col] = {"value": "", "confidence": "high"}
                    continue
                found = _check_value_in_text(val_str, doc_text_lower)
                if found:
                    validated_row[col] = {"value": val_str, "confidence": "high"}
                else:
                    variant = _check_value_variants(val_str, doc_text)
                    validated_row[col] = {
                        "value": val_str,
                        "confidence": "medium" if variant else "low",
                    }
            validated_rows.append(validated_row)

    return {
        "validated_fields": validated,
        "validated_rows": validated_rows,
        "flagged": flagged,
        "confidence_map": confidence_map,
        "flag_count": len(flagged),
    }


def _check_value_in_text(value: str, doc_text_lower: str) -> bool:
    """Check if a value appears in the document text."""
    if not value or len(value) < 2:
        return True  # Short values - don't flag

    val_lower = value.lower().strip()

    # Direct match
    if val_lower in doc_text_lower:
        return True

    # Numeric match - strip formatting
    val_numeric = re.sub(r'[^0-9.]', '', value)
    if len(val_numeric) >= 4:
        doc_numeric = re.sub(r'[^0-9.]', '', doc_text_lower)
        if val_numeric in doc_numeric:
            return True

    return False


def _check_value_variants(value: str, doc_text: str) -> bool:
    """
    Check if any variant of the value appears in the document.
    Handles date normalisation, number formatting etc.
    """
    # Date variants: "2024-01-31" -> check for "January 31", "31/01", "01/31" etc.
    date_match = re.match(r'(\d{4})-(\d{2})-(\d{2})', value)
    if date_match:
        y, m, d = date_match.groups()
        month_names = ["january","february","march","april","may","june",
                       "july","august","september","october","november","december"]
        month_name = month_names[int(m) - 1]
        variants = [
            f"{d}/{m}/{y}", f"{m}/{d}/{y}", f"{d}-{m}-{y}",
            f"{month_name} {int(d)}", f"{int(d)} {month_name}",
            f"{month_name[:3]} {int(d)}", f"{int(d)} {month_name[:3]}",
        ]
        doc_lower = doc_text.lower()
        return any(v in doc_lower for v in variants)

    # Amount variants: "8410.00" -> check for "$8,410", "8,410.00" etc.
    amount_match = re.match(r'^(\d+)\.(\d{2})$', value)
    if amount_match:
        integer_part = amount_match.group(1)
        # Check for the integer part at least
        if len(integer_part) >= 4 and integer_part in doc_text.replace(",", ""):
            return True

    return False


# ==============================================================================
# PDFPLUMBER DIRECT TABLE EXTRACTION
# ==============================================================================

def _extract_pdf_table_direct(file_path: Path, template_data: dict) -> Optional[list]:
    """
    Extract table rows directly from PDF using pdfplumber.
    Handles tables at ANY position on the page (not just top-left).
    Returns list of row dicts or None if no matching table found.
    """
    try:
        import pdfplumber

        regions = template_data.get("regions", {})
        table_regions = regions.get("table_regions", [])

        if not table_regions:
            return None

        col_names = []
        for tr in table_regions:
            col_names.extend(tr.get("column_names", []))
        col_names = list(dict.fromkeys(col_names))

        if not col_names:
            return None

        skip_kw = {"subtotal", "grand total", "shipping", "discount",
                   "charges", "refund", "paid", "free", "balance due"}
        all_rows = []

        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                # Strategy 1: line-based (visible borders)
                tables = page.extract_tables({
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "snap_tolerance": 4,
                    "join_tolerance": 4,
                })
                # Strategy 2: text-based (no visible borders)
                if not tables:
                    tables = page.extract_tables({
                        "vertical_strategy": "text",
                        "horizontal_strategy": "text",
                        "snap_tolerance": 6,
                        "join_tolerance": 6,
                    })

                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    header_row_idx = _find_header_row(table, col_names)
                    if header_row_idx is None:
                        continue
                    pdf_headers = [_clean_cell(c) for c in table[header_row_idx]]
                    col_map = _match_columns(col_names, pdf_headers)
                    if not col_map:
                        continue

                    for row in table[header_row_idx + 1:]:
                        if not row or all(not _clean_cell(c) for c in row):
                            continue
                        first_val = _clean_cell(row[0]) if row else ""
                        if not first_val:
                            continue
                        if any(kw in first_val.lower() for kw in skip_kw):
                            continue
                        if re.match(r'^\d{1,4}$', first_val):
                            continue

                        row_dict = {}
                        for col_name in col_names:
                            pdf_idx = col_map.get(col_name)
                            if pdf_idx is not None and pdf_idx < len(row):
                                val = _clean_cell(row[pdf_idx])
                                val = re.sub(r'^\$', '', val).replace(',', '')
                            else:
                                val = ""
                            row_dict[col_name] = val

                        if any(v for v in row_dict.values()):
                            all_rows.append(row_dict)

        print(f"[DIRECT] {file_path.name}: {len(all_rows)} rows", flush=True)
        return all_rows if all_rows else None

    except Exception as e:
        print(f"[DIRECT] Failed {file_path.name}: {e}", flush=True)
        return None


def _clean_cell(val) -> str:
    if val is None: return ""
    return " ".join(str(val).strip().split())

def _find_header_row(table, col_names):
    col_lower = [c.lower() for c in col_names]
    best_idx, best_score = None, 0
    for i, row in enumerate(table[:8]):
        if not row: continue
        row_vals = [_clean_cell(c).lower() for c in row]
        score = sum(1 for col in col_lower if any(col in cell or cell in col for cell in row_vals if cell))
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx if best_score >= max(1, len(col_names) // 2) else None

def _match_columns(col_names, pdf_headers):
    import difflib
    mapping = {}
    pdf_lower = [h.lower() for h in pdf_headers]
    for col in col_names:
        cl = col.lower()
        if cl in pdf_lower:
            mapping[col] = pdf_lower.index(cl); continue
        for i, h in enumerate(pdf_lower):
            if cl in h or h in cl:
                mapping[col] = i; break
        else:
            matches = difflib.get_close_matches(cl, pdf_lower, n=1, cutoff=0.6)
            if matches:
                mapping[col] = pdf_lower.index(matches[0])
    return mapping

def _clean_text_for_table(text: str) -> str:
    text = re.sub(r'(\d{8,})\n(\d{1,3})\n', lambda m: m.group(1)+m.group(2)+'\n', text)
    text = re.sub(r'(\d{8,})\n(\d{1,3})$', lambda m: m.group(1)+m.group(2), text)
    text = re.sub(r'(-\s*)\n(#\d+)', r'\1 \2', text)
    return text


# ==============================================================================
# VALUE NORMALISATION
# ==============================================================================

def _normalise_values(row: dict, doc_type: str) -> dict:
    numeric_fields = _get_numeric_fields(doc_type)
    cleaned = {}
    for k, v in row.items():
        if v is None:
            cleaned[k] = ""; continue
        s = str(v).strip()
        kl = k.lower().replace(" ", "_")
        if any(kl == f or kl.endswith(f) for f in numeric_fields):
            s = re.sub(r'[£€$¥₹,\s]', '', s)
            s = re.sub(r'^\((.+)\)$', r'-\1', s)
        cleaned[k] = s
    return cleaned

def _filter_ghost_rows(rows: list, col_names: list) -> list:
    if not rows or not col_names: return rows
    first_col = col_names[0]
    # Only skip clear summary rows — never skip "tax" as it's a valid row type
    skip_exact = {"subtotal", "grand total", "shipping", "discount",
                  "charges", "refund", "balance due", "amount due"}
    clean = []
    for row in rows:
        first_val = str(row.get(first_col, "")).strip()
        if not first_val: continue
        if re.match(r'^[\d,.\-\s]{1,6}$', first_val) and not re.search(r'[a-zA-Z]', first_val): continue
        if first_val.lower() in skip_exact: continue
        clean.append(row)
    return clean


# ==============================================================================
# MULTI-DOCUMENT DETECTION
# ==============================================================================

def _detect_document_boundaries_vision(doc_images_b64: list, orchestrator,
                                        filename: str, doc_type: str) -> list:
    """
    Detect if a PDF contains multiple separate documents using vision.
    Returns list of {index, page_indices, hint} for each detected document.

    This is used for PDFs like "100 cheques in one file" where each page
    or each half-page is a separate document.
    """
    if not doc_images_b64:
        return [{"index": 0, "page_indices": list(range(1)), "hint": "full document"}]

    total_pages = len(doc_images_b64)

    if total_pages == 1:
        # Single page - could still contain multiple documents side by side
        detection_prompt = f"""Look at this document image carefully.

Does this page contain MULTIPLE SEPARATE {doc_type} documents on it?
(For example: two cheques on one page, two receipts, multiple invoices)

Return ONLY this JSON:
{{
  "document_count": <integer>,
  "layout": "single|side_by_side|stacked|grid",
  "documents": [
    {{"index": 0, "hint": "brief description, e.g. cheque to Pacific Steel"}}
  ]
}}

If there is only ONE document: document_count = 1."""

        try:
            detection = orchestrator.llm.extract(
                image_b64=doc_images_b64[0],
                prompt=detection_prompt
            )
            if detection.success and detection.parsed_json:
                raw = detection.parsed_json
                count = raw.get("document_count", 1)
                if count > 1:
                    print(f"[DETECT] {filename}: {count} docs on single page", flush=True)
                    return [
                        {"index": d["index"], "page_indices": [0],
                         "hint": d.get("hint", f"doc {d['index']+1}"),
                         "sub_index": d["index"],
                         "total_on_page": count}
                        for d in raw.get("documents", [{"index": 0}])
                    ]
        except Exception as e:
            print(f"[DETECT] single page detection error: {e}", flush=True)

        return [{"index": 0, "page_indices": [0], "hint": "full document"}]

    else:
        # Multi-page: ask vision model whether this is ONE multi-page document
        # or MULTIPLE separate documents merged into one PDF.
        doc_type_label = doc_type if doc_type and doc_type != "other" else "document"
        boundary_prompt = f"""You are analyzing page 1 of a {total_pages}-page PDF.

Your task: decide if this PDF is ONE document that spans multiple pages,
OR multiple SEPARATE {doc_type_label} documents merged into a single file.

KEY DISTINCTION:
- ONE multi-page document: contract, report, multi-page invoice, bank statement
  with months on separate pages, financial report — all clearly one document.
- MULTIPLE separate documents: 40 individual invoices in one PDF, 12 separate
  receipts scanned together, batch of cheques, individual monthly statements
  combined into one file — each is a complete, independent document.

If you see page numbers like "Page 1 of 5" or "Page 2/10", this is ONE document.
If you see each page has its own header/footer/date/number, it may be MULTIPLE docs.

Return ONLY this JSON (no markdown, no explanation):
{{
  "is_multi_document": true_or_false,
  "document_count": <integer, 1 if single>,
  "reasoning": "one sentence why",
  "documents": [
    {{"doc_number": 1, "start_page": 1, "end_page": 2, "hint": "brief description"}},
    {{"doc_number": 2, "start_page": 3, "end_page": 3, "hint": "brief description"}}
  ]
}}

Total pages in this PDF: {total_pages}
If is_multi_document is false, set document_count to 1 and documents to
[{{"doc_number": 1, "start_page": 1, "end_page": {total_pages}, "hint": "full document"}}]"""

        try:
            detection = orchestrator.llm.extract(
                image_b64=doc_images_b64[0],
                prompt=boundary_prompt,
            )
            if detection.success and detection.parsed_json:
                raw = detection.parsed_json
                is_multi = raw.get("is_multi_document", False)
                count    = raw.get("document_count", 1)
                docs     = raw.get("documents", [])
                reasoning = raw.get("reasoning", "")

                if is_multi and count > 1 and docs:
                    result_segs = []
                    for d in docs:
                        sp = max(0, int(d.get("start_page", 1)) - 1)       # 0-indexed
                        ep = min(total_pages - 1, int(d.get("end_page", sp + 1)) - 1)
                        result_segs.append({
                            "index":        int(d.get("doc_number", len(result_segs) + 1)) - 1,
                            "page_indices": list(range(sp, ep + 1)),
                            "hint":         d.get("hint", f"document {len(result_segs)+1}"),
                        })
                    print(
                        f"[DETECT] {filename}: {len(result_segs)} documents detected "
                        f"across {total_pages} pages — {reasoning}",
                        flush=True,
                    )
                    return result_segs
                else:
                    # Single multi-page document
                    print(
                        f"[DETECT] {filename}: {total_pages} pages -> single document — {reasoning}",
                        flush=True,
                    )
                    return [{"index": 0,
                             "page_indices": list(range(total_pages)),
                             "hint": "full document"}]

        except Exception as e:
            print(f"[DETECT] multi-page boundary detection error: {e}", flush=True)

        # Safe fallback: treat as single document (not one-per-page, which fragments long docs)
        print(
            f"[DETECT] {filename}: {total_pages} pages -> single-document fallback (detection unavailable)",
            flush=True,
        )
        return [{"index": 0, "page_indices": list(range(total_pages)), "hint": "full document"}]


# ==============================================================================
# VALUE NORMALIZATION
# ==============================================================================

def _normalize_value(v) -> str:
    """
    Normalize a single extracted value:
    - Numbers: strip $£€₹ and commas, convert (x) to -x, expand K/M
    - Dates: normalize to YYYY-MM-DD
    - Empty/null: return ""
    """
    if v is None:
        return ""
    s = str(v).strip()
    if not s or s.lower() in ("null", "n/a", "none", "not found", "–", "-"):
        return ""

    # Negative accounting format: (2.85) -> -2.85
    m = re.match(r'^\(([0-9,]+\.?[0-9]*)\)$', s)
    if m:
        return "-" + m.group(1).replace(",", "")

    # Month name date formats
    months = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
              "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
    m_name = re.match(
        r'^(\d{1,2})\s+([A-Za-z]{3,9})\s+(\d{4})$|^([A-Za-z]{3,9})[. ]+(\d{1,2}),?\s+(\d{4})$', s)
    if m_name:
        try:
            if m_name.group(1):
                d, mo, y = int(m_name.group(1)), months.get(m_name.group(2).lower()[:3]), int(m_name.group(3))
            else:
                mo = months.get(m_name.group(4).lower()[:3])
                d, y = int(m_name.group(5)), int(m_name.group(6))
            if mo:
                return f"{y}-{mo:02d}-{d:02d}"
        except Exception:
            pass

    # Numeric date formats
    date_patterns = [
        (r'^(\d{1,2})/(\d{1,2})/(\d{4})$',  lambda m: f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"),
        (r'^(\d{1,2})-(\d{1,2})-(\d{4})$',  lambda m: f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"),
    ]
    for pattern, formatter in date_patterns:
        m = re.match(pattern, s)
        if m:
            try:
                return formatter(m)
            except Exception:
                pass

    # Currency/number cleanup: strip symbols and commas
    num_candidate = re.sub(r'[$£€₹,\s]', '', s)
    km = re.match(r'^(-?[0-9.]+)[Kk]$', num_candidate)
    if km:
        try:
            return str(float(km.group(1)) * 1000)
        except Exception:
            pass
    mm = re.match(r'^(-?[0-9.]+)[Mm][Mm]?$', num_candidate)
    if mm:
        try:
            return str(float(mm.group(1)) * 1_000_000)
        except Exception:
            pass

    return s


def _normalize_field_values(fields: dict) -> dict:
    """Normalize all values in an extracted_fields dict."""
    if not fields:
        return fields
    result = {}
    for k, v in fields.items():
        if isinstance(v, dict):
            val = v.get("value", "")
            result[k] = {**v, "value": _normalize_value(val)}
        else:
            result[k] = _normalize_value(v)
    return result


def _normalize_row_values(row: dict) -> dict:
    """Normalize all values in a table row."""
    if not row:
        return row
    return {k: _normalize_value(v) for k, v in row.items()}


# ==============================================================================
# RESULT PROCESSORS
# ==============================================================================

def _fix_split_decimals(fields: dict) -> dict:
    """
    Fix numeric values split across page breaks.
    e.g. {"Net Pay": "7513.0", "unknown_1": "3"} -> {"Net Pay": "7513.03"}
    Pattern: a value ending in ".<digit>" or ".<digit>0" followed by a lone
    1-2 digit orphan value that completes the decimal.
    """
    if not fields:
        return fields
    result = {}
    items = list(fields.items())
    skip_next = False
    for i, (key, val) in enumerate(items):
        if skip_next:
            skip_next = False
            continue
        val_str = str(val).strip() if val is not None else ""
        # Check if this value looks like a truncated decimal: ends with "."
        # or has fewer decimal digits than expected (e.g. "7513.0" when full is "7513.03")
        if i + 1 < len(items):
            next_key, next_val = items[i + 1]
            next_str = str(next_val).strip() if next_val is not None else ""
            # If next value is a lone 1-2 digit number and current ends with decimal
            if (re.match(r'^\d{1,2}$', next_str) and
                re.match(r'^\d[\d,]*\.\d{0,2}$', val_str.replace(',', ''))):
                # Combine: "7513.0" + "3" = "7513.03"
                clean = val_str.replace(',', '')
                if clean.endswith('.'):
                    merged = clean + next_str
                elif len(clean.split('.')[-1]) < 2:
                    merged = clean + next_str
                else:
                    merged = val_str
                    result[key] = val
                    continue
                try:
                    result[key] = float(merged)
                    skip_next = True
                    continue
                except ValueError:
                    pass
        result[key] = val
    return result


def _fix_split_decimals_row(row: dict) -> dict:
    """Apply decimal split fix to a single table row."""
    return _fix_split_decimals(row) if row else row


def _validate_row_alignment(
    table_rows: list, table_regions: list
) -> tuple:
    """
    Bug 6: Check that all columns in each table have consistent row presence.
    When the AI returns rows where some columns are systematically absent for
    the first N rows, the data is misaligned (column 2 started at row 13 instead of 1).

    Returns (rows, is_misaligned, warning_msg).
    If misalignment is detected, attempt re-alignment by removing leading empty rows.
    """
    if not table_rows or not table_regions:
        return table_rows, False, ""

    # Collect expected columns per table_source
    expected_cols: dict = {}
    for tr in table_regions:
        name = tr.get("section_label", "")
        key = re.sub(r'[^a-z0-9]', '_', name.lower()).strip('_')
        expected_cols[key] = tr.get("column_names", [])
    # Also accept "table_rows" as the default source
    if table_regions:
        expected_cols["table"] = table_regions[0].get("column_names", [])

    # Group rows by _table_source (or use the single table case)
    groups: dict = {}
    for row in table_rows:
        src = row.get("_table_source", "table")
        groups.setdefault(src, []).append(row)

    warnings = []
    final_rows = []
    is_misaligned = False

    for src, rows in groups.items():
        cols = expected_cols.get(src, [])
        if not cols or len(rows) < 3:
            final_rows.extend(rows)
            continue

        # For each expected column, count how many rows have a non-empty value
        col_fill = {}
        for col in cols:
            col_fill[col] = sum(1 for r in rows if r.get(col, "") not in ("", None))

        if not col_fill:
            final_rows.extend(rows)
            continue

        max_fill = max(col_fill.values())
        min_fill = min(col_fill.values())

        # If the fill counts differ by more than 25% across columns → misalignment
        if max_fill > 0 and (max_fill - min_fill) / max_fill > 0.25:
            is_misaligned = True
            # Find columns with low fill
            sparse_cols = [c for c, f in col_fill.items() if f < max_fill * 0.75]
            warn = (
                f"[BUG6] Row-column misalignment in source='{src}': "
                f"max_fill={max_fill}, min_fill={min_fill}, "
                f"sparse_cols={sparse_cols} — flagging for review"
            )
            print(warn, flush=True)
            warnings.append(warn)
            # Attempt repair: strip leading rows that are empty in ALL sparse columns
            repaired = []
            leading_empty_stripped = 0
            for row in rows:
                if not repaired and all(
                    row.get(c, "") in ("", None) for c in sparse_cols
                ):
                    leading_empty_stripped += 1
                    continue
                repaired.append(row)
            if repaired and leading_empty_stripped > 0:
                print(
                    f"[BUG6] Stripped {leading_empty_stripped} leading-empty rows "
                    f"to repair alignment for '{src}'",
                    flush=True,
                )
                final_rows.extend(repaired)
            else:
                final_rows.extend(rows)
        else:
            final_rows.extend(rows)

    return final_rows, is_misaligned, "; ".join(warnings)


def _regions_to_jsonable(regions: dict) -> dict:
    """FIX 4: make the regions dict JSON-serializable so it can be persisted in
    document_results.extraction_json and reused verbatim at export time."""
    out = dict(regions or {})
    slr = out.get("section_label_rows")
    if isinstance(slr, (set, frozenset)):
        out["section_label_rows"] = sorted(slr)
    return out


def _coerce_extracted_fields(raw_doc: dict, regions: dict) -> tuple:
    """
    FIX 3: enforce the extracted_fields (cell-ref keyed) schema on the AI response
    for form / parallel-group templates.

    Returns (extracted_fields_dict, warning_str). warning_str is non-empty when the
    response had to be remapped/flattened or could not be reconciled — the caller
    uses it to flag the document for review instead of silently writing nothing.
    """
    import re as _re
    primary_mode = regions.get("primary_mode", "form_kv")
    ef = raw_doc.get("extracted_fields", {})
    ref_re = _re.compile(r'^[A-Z]+[0-9]+$')

    form_modes = ("parallel_groups", "form_kv", "form_with_targets", "mixed")
    if primary_mode not in form_modes:
        return (ef if isinstance(ef, dict) else {}), ""

    # Ordered (ref, label) bindings exactly as presented to the AI in the prompt.
    ordered = []
    for pg in regions.get("parallel_column_groups", []):
        for it in pg.get("items", []):
            if it.get("value_ref"):
                ordered.append((it["value_ref"], (it.get("label") or "")))
    for kv in regions.get("kv_pairs", []):
        if kv.get("value_ref"):
            ordered.append((kv["value_ref"], (kv.get("label") or "")))
    for t in regions.get("explicit_targets", []):
        if t.get("ref"):
            ordered.append((t["ref"], (t.get("label") or "")))
    label_to_ref = {(lbl or "").strip().lower(): ref
                    for ref, lbl in ordered if (lbl or "").strip()}

    # Case 1: already a cell-ref-keyed dict → accept as-is.
    if isinstance(ef, dict) and ef:
        ref_keys = [k for k in ef if ref_re.match(str(k))]
        if len(ref_keys) >= max(1, len(ef) // 2):
            return ef, ""
        # Case 2: label-keyed dict → remap labels back to cell refs.
        remapped = {}
        for k, v in ef.items():
            if ref_re.match(str(k)):
                remapped[str(k)] = v
            else:
                ref = label_to_ref.get(str(k).strip().lower())
                if ref:
                    remapped[ref] = v
        if remapped:
            return remapped, (f"AI returned label-keyed fields; remapped "
                              f"{len(remapped)} to cell refs (verify)")

    # Case 3: no usable extracted_fields but table_rows present → flatten by label match.
    table_rows = list(raw_doc.get("table_rows", []) or [])
    for key, val in raw_doc.items():
        if key.endswith("_rows") and key != "table_rows" and isinstance(val, list):
            table_rows.extend(val)
    if table_rows and ordered:
        _num_re = _re.compile(r'^[\$\-\(\)]?[\d,]+(?:\.\d+)?\)?$')
        flat = {}
        for row in table_rows:
            if not isinstance(row, dict):
                continue
            row_label, row_value = "", ""
            for k, v in row.items():
                if str(k).startswith("_"):
                    continue
                sv = str(v).strip()
                if not sv:
                    continue
                if any(ch.isalpha() for ch in sv) and not row_label:
                    row_label = sv
                elif _num_re.match(sv.replace(" ", "")) and not row_value:
                    row_value = sv
            if not row_label:
                continue
            ref = label_to_ref.get(row_label.strip().lower())
            if not ref:
                rl = row_label.strip().lower()
                for lbl, r in label_to_ref.items():
                    if lbl and (lbl in rl or rl in lbl):
                        ref = r
                        break
            if ref and row_value:
                flat[ref] = {"value": row_value, "confidence": "medium"}
        if flat:
            return flat, (f"AI returned table_rows instead of extracted_fields; "
                          f"flattened {len(flat)} values to cell refs (verify)")

    # Case 4: cannot reconcile → keep whatever we had, flag for review.
    if isinstance(ef, dict) and ef:
        return ef, ""
    return {}, ("AI response had no usable extracted_fields for a form/parallel "
                "template — flagged for review")


def _process_vision_result(raw_doc: dict, template_data: dict, filename: str,
                            doc_type: str, elapsed: float, extraction,
                            doc_text: str, seg_hint: str = "",
                            doc_index: int = 0) -> object:
    """
    Process the result from a single document extracted by vision.
    Applies pdfplumber validation to every extracted value.
    """
    from orchestrator import DocumentExtractionResult

    # Unwrap the documents[] wrapper if Gemini nested the result inside it.
    # Some call sites pass extraction.parsed_json directly (e.g. the image path),
    # so the document fields (layout_sections, extracted_fields, table_rows) end up
    # nested under documents[0] instead of at the top level — which silently lost
    # layout_sections and extracted_fields. Idempotent for callers that already
    # unwrapped (documents[0] has no "documents" key).
    if (isinstance(raw_doc, dict) and isinstance(raw_doc.get("documents"), list)
            and raw_doc["documents"] and isinstance(raw_doc["documents"][0], dict)):
        raw_doc = raw_doc["documents"][0]

    regions = template_data.get("regions", {})
    layout = template_data.get("layout", {})
    cells = layout.get("cells", {})
    primary_mode = regions.get("primary_mode", "form_kv")

    # FIX 3: coerce/validate the AI response into the cell-ref extracted_fields
    # schema the writer expects; _schema_warn is set when reconciliation was needed.
    extracted_fields_raw, _schema_warn = _coerce_extracted_fields(raw_doc, regions)
    confidence_raw = raw_doc.get("confidence", "medium")
    if isinstance(confidence_raw, dict):
        confidence_raw = "medium"

    # -- Collect table rows from ALL tables in response ------------------------
    # For single-table responses: raw_doc["table_rows"]
    # For multi-table responses: raw_doc["earning_table_rows"], raw_doc["deduction_table_rows"] etc.
    # We collect all into a single list, adding a "table_source" key to each row
    table_rows_raw = []
    all_doc_keys = list(raw_doc.keys())
    print(f"[COLLECT] AI response keys: {all_doc_keys}", flush=True)
    if raw_doc.get("table_rows"):
        print(f"[COLLECT] table_rows: {len(raw_doc['table_rows'])} rows", flush=True)
        table_rows_raw.extend(raw_doc["table_rows"])
    # Collect any additional table arrays (multi-table templates)
    for key, val in raw_doc.items():
        if key.endswith("_rows") and key != "table_rows" and isinstance(val, list):
            print(f"[COLLECT] {key}: {len(val)} rows", flush=True)
            for row in val:
                if isinstance(row, dict):
                    row["_table_source"] = key.replace("_rows", "")
                    table_rows_raw.append(row)
    print(f"[COLLECT] total rows collected: {len(table_rows_raw)}", flush=True)

    # -- Bug 6: validate row/column alignment -----------------------------------
    alignment_warning = ""
    alignment_misaligned = False
    if table_rows_raw:
        table_rows_raw, alignment_misaligned, alignment_warning = _validate_row_alignment(
            table_rows_raw, regions.get("table_regions", [])
        )
        if alignment_misaligned:
            print(
                f"[BUG6] Alignment issue detected in {filename} — "
                f"document will be flagged for review",
                flush=True,
            )

    # -- Fix page-break decimal splits -----------------------------------------
    extracted_fields_raw = _fix_split_decimals(extracted_fields_raw)
    table_rows_raw = [_fix_split_decimals_row(r) for r in table_rows_raw]

    # -- Normalize numeric and date values -------------------------------------
    extracted_fields_raw = _normalize_field_values(extracted_fields_raw)
    table_rows_raw = [_normalize_row_values(r) for r in table_rows_raw]

    # -- pdfplumber validation -------------------------------------------------
    # Skip cross-validation for scanned/image PDFs — pdfplumber returns empty text
    # for scanned docs, causing every correctly-extracted field to be flagged as
    # low confidence / needs_review even when the LLM was right.
    _skip_plumber_validation = not doc_text or len(doc_text.strip()) < 80
    if _skip_plumber_validation:
        validated_fields = {
            k: {"value": (v.get("value", "") if isinstance(v, dict) else str(v or "")),
                "confidence": "medium", "validated": True}
            for k, v in extracted_fields_raw.items()
        }
        validated_rows = [
            {col: {"value": str(val or ""), "confidence": "medium"}
             for col, val in row.items()}
            for row in table_rows_raw
        ]
        flagged = []
        validation = {"validated_fields": validated_fields, "validated_rows": validated_rows,
                      "flagged": [], "confidence_map": {}, "flag_count": 0}
    else:
        validation = _validate_with_pdfplumber(extracted_fields_raw, doc_text, table_rows_raw)
        validated_fields = validation["validated_fields"]
        validated_rows = validation["validated_rows"]
        flagged = validation["flagged"]

    # FIX 3: a schema-coercion warning escalates the document to needs_review
    # so a silently-unwritable AI response never passes as clean.
    if _schema_warn:
        flagged = list(flagged) + [{"ref": "_schema", "value": "", "issue": _schema_warn}]
        print(f"[SCHEMA] {filename}: {_schema_warn}", flush=True)

    # -- Build human-readable extracted_data -----------------------------------
    ref_to_label = {}
    for key, cell in cells.items():
        if not isinstance(cell, dict): continue
        val = cell.get("value", "").strip()
        is_extract = cell.get("extractTarget", False)
        if is_extract and val:
            parts = key.split(",")
            if len(parts) == 2:
                ref_to_label[_cell_ref(int(parts[0]), int(parts[1]))] = val

    for t in regions.get("explicit_targets", []):
        if t["ref"] not in ref_to_label:
            ref_to_label[t["ref"]] = t["label"]
    for kv in regions.get("kv_pairs", []):
        if kv["value_ref"] not in ref_to_label:
            ref_to_label[kv["value_ref"]] = kv["label"]
    for tc in regions.get("two_col_pairs", []):
        if tc["left_value_ref"] not in ref_to_label:
            ref_to_label[tc["left_value_ref"]] = tc["left_label"]
        if tc["right_value_ref"] not in ref_to_label:
            ref_to_label[tc["right_value_ref"]] = tc["right_label"]

    # Auto-resolve unlabelled refs
    for ref in validated_fields:
        if ref not in ref_to_label:
            try:
                col_str = "".join(ch for ch in ref if ch.isalpha()).upper()
                row_str = "".join(ch for ch in ref if ch.isdigit())
                c_idx = sum((ord(ch)-64)*(26**i) for i,ch in enumerate(reversed(col_str)))-1
                r_idx = int(row_str)-1
                left_val = (cells.get(f"{r_idx},{c_idx-1}") or {}).get("value","").strip() if c_idx>0 else ""
                above_val = (cells.get(f"{r_idx-1},{c_idx}") or {}).get("value","").strip() if r_idx>0 else ""
                ref_to_label[ref] = left_val or above_val or ref
            except Exception:
                ref_to_label[ref] = ref

    extracted_data = {}
    for ref, label in ref_to_label.items():
        vf = validated_fields.get(ref, {"value": "", "confidence": "high"})
        extracted_data[label] = {
            "value": vf.get("value", ""),
            "confidence": vf.get("confidence", "high"),
            "ref": ref,
        }

    # Static label cells
    for key, cell in cells.items():
        if not isinstance(cell, dict): continue
        val = cell.get("value","").strip()
        if val and not cell.get("extractTarget"):
            parts = key.split(",")
            if len(parts) == 2:
                ref = _cell_ref(int(parts[0]), int(parts[1]))
                extracted_data[f"_label_{ref}"] = {"value": val, "confidence": "high"}

    # Table rows - normalise
    # For multi-table docs, rows have _table_source set to identify which table they belong to.
    # Each table has its own column set — we must not merge columns across tables.
    normalised_rows = []
    if table_rows_raw:
        table_regions_list = regions.get("table_regions", [])

        # Build a map: table_source_key -> column_names for that table
        # table_source is set to the array key minus "_rows" (e.g. "earning_table")
        source_to_cols = {}
        for tr in table_regions_list:
            section = tr.get("section_label", "")
            import re as _re
            key = _re.sub(r'[^a-z0-9]', '_', section.lower()).strip('_')
            source_to_cols[key] = tr.get("column_names", [])

        # Also build merged col_names for single-table case
        all_col_names = []
        for tr in table_regions_list:
            all_col_names.extend(tr.get("column_names", []))
        all_col_names = list(dict.fromkeys(all_col_names))

        for row in table_rows_raw:
            if not isinstance(row, dict):
                continue

            # Determine which table this row belongs to
            table_source = row.get("_table_source", "")
            row_col_names = source_to_cols.get(table_source, all_col_names)

            if not row_col_names:
                row_col_names = [k for k in row.keys()
                                if not k.startswith("_")]

            # Filter ghost rows using THIS table's first column
            first_col = row_col_names[0] if row_col_names else None
            if first_col:
                first_val = str(row.get(first_col, "")).strip()
                if not first_val:
                    continue
                # Only skip rows that are clearly summary/total rows
                # Do NOT skip "tax" — it's a legitimate deduction type
                skip_kw = {"subtotal", "grand total", "shipping", "discount",
                           "charges", "refund", "balance due", "amount due"}
                if any(first_val.lower() == kw for kw in skip_kw):
                    continue
                # Skip rows where first value is ONLY a number (row number artifacts)
                if re.match(r'^\d{1,3}$', first_val):
                    continue

            # Build clean row with only this table's columns
            clean = {}
            for col in row_col_names:
                clean[col] = str(row.get(col, "") or "").strip()
            # Preserve table source for Excel writer
            if table_source:
                clean["_table_source"] = table_source

            normalised_rows.append(_normalise_values(clean, doc_type))

    has_table = bool(normalised_rows)
    overall_confidence = raw_doc.get("overall_confidence", "medium")

    r = DocumentExtractionResult(filename=filename)
    r.document_type = doc_type

    # Build per-table row maps using separate array keys
    per_table_rows = {}
    for tr in regions.get("table_regions", []):
        section = tr.get("section_label", "")
        key = re.sub(r'[^a-z0-9]', '_', section.lower()).strip('_') + "_rows"
        # Get rows for this table from the raw response
        raw_rows = raw_doc.get(key, [])
        if not raw_rows:
            # Fall back: filter table_rows_raw by Table field or _table_source
            raw_rows = [
                r2 for r2 in table_rows_raw
                if isinstance(r2, dict) and (
                    r2.get("Table", "").strip().lower() == section.strip().lower() or
                    r2.get("_table_source", "").replace("_", " ").strip().lower() in section.lower()
                )
            ]
        # Normalise these rows
        col_names = tr.get("column_names", [])
        normed = []
        for row in raw_rows:
            if isinstance(row, dict):
                clean = {col: str(row.get(col, "") or "").strip() for col in col_names}
                normed.append(_normalise_values(clean, doc_type))
        per_table_rows[key] = normed

    r.extracted_data = {
        "document_type": doc_type,
        "overall_confidence": overall_confidence,
        "extraction_method": "vision_primary",
        # FIX 4: persist the region analysis so export reuses it (no re-analysis)
        "template_regions": _regions_to_jsonable(regions),
        # COMPONENT 5: layout-mode "extract & place" output (empty for field-based)
        "layout_sections": (raw_doc.get("layout_sections")
                            if isinstance(raw_doc.get("layout_sections"), dict) else {}),
        # Section-aware column groups — used by the layout writer's column safety net
        "binding_column_groups": (template_data.get("binding_map", {})
                                  .get("_meta", {}).get("column_groups", [])),
        "table_mode": has_table and primary_mode == "table",
        "mixed_mode": has_table and primary_mode == "mixed",
        "table_rows": normalised_rows,
        "column_headers": [tr["column_names"] for tr in regions.get("table_regions",[])],
        "row_count": len(normalised_rows),
        "extracted_data": extracted_data,
        "extracted_fields": {k: v.get("value","") if isinstance(v,dict) else v
                              for k,v in validated_fields.items()},
        "segment_hint": seg_hint,
        "doc_index": doc_index,
        "validation": {
            "flagged_count": len(flagged),
            "flagged_fields": flagged,
            "confidence_map": validation.get("confidence_map", {}),
            "alignment_misaligned": alignment_misaligned,
            "alignment_warning": alignment_warning,
        },
        **per_table_rows,   # adds earning_table_rows, deduction_table_rows etc.
    }
    r.extraction_response = extraction
    r.processing_time_ms = elapsed
    r.success = True

    status = f"{len(flagged)} flags" if flagged else "clean"
    print(f"[EXTRACT] vision: {filename} doc#{doc_index} -> "
          f"{len(extracted_data)} fields, {len(normalised_rows)} rows, "
          f"{overall_confidence} confidence, {status}", flush=True)
    return r


def _make_table_result(rows, template_data, filename, doc_type, elapsed, method, extraction=None, confidence="high"):
    from orchestrator import DocumentExtractionResult
    regions = template_data.get("regions", {})
    col_names = []
    for tr in regions.get("table_regions", []):
        col_names.extend(tr.get("column_names", []))
    col_names = list(dict.fromkeys(col_names))
    if not col_names:
        # Fallback: read from layout directly
        layout = template_data.get("layout", {})
        cells = layout.get("cells", {})
        headers = []
        for key, cell in cells.items():
            if not isinstance(cell, dict): continue
            val = cell.get("value","").strip()
            if not val: continue
            parts = key.split(",")
            if len(parts) == 2 and int(parts[0]) == 0:
                c = int(parts[1])
                headers.append((c, val))
        col_names = [v for _, v in sorted(headers)]
    normalised = [_normalise_values(row, doc_type) for row in rows]
    r = DocumentExtractionResult(filename=filename)
    r.document_type = doc_type
    r.extracted_data = {
        "document_type": doc_type, "overall_confidence": confidence,
        "table_mode": True, "extraction_method": method,
        "table_rows": normalised, "column_headers": col_names,
        "row_count": len(normalised),
        "extracted_data": {
            **({col: {"value": normalised[0].get(col,""),"confidence":"high"} for col in col_names} if normalised else {}),
            "_table_row_count": {"value": str(len(normalised)),"confidence":"high"},
        },
    }
    r.extraction_response = extraction
    r.processing_time_ms = elapsed
    r.success = True
    print(f"[EXTRACT] {method}: {filename} -> {len(normalised)} rows @ {confidence}", flush=True)
    return r

def _fail(filename, error):
    from orchestrator import DocumentExtractionResult
    r = DocumentExtractionResult(filename=filename)
    r.error = error; r.processing_time_ms = 0; r.success = False
    return r


# ==============================================================================
# MAIN EXTRACTION ENGINE - VISION FIRST
# ==============================================================================

def _extract_with_template(orchestrator, file_path: Path, template_data: dict,
                            selected_pages: Optional[list] = None):
    """
    Vision-First extraction engine - safety-wrapped version.
    All errors are caught and returned as failed DocumentResult objects
    so the job always completes with meaningful error messages.
    """
    try:
        return _extract_with_template_inner(orchestrator, file_path, template_data,
                                            selected_pages=selected_pages)
    except Exception as e:
        print(f"[EXTRACT] FATAL {file_path.name}: {e}", flush=True)
        traceback.print_exc()
        r = _fail(file_path.name, f"Fatal extraction error: {str(e)[:200]}")
        r.processing_time_ms = 0
        return [r]


def _extract_image_with_template(orchestrator, file_path: Path,
                                  template_data: Optional[dict]) -> list:
    """
    Extract from a single image file (JPG/PNG/WEBP/TIFF/BMP).
    Sends the image directly to Gemini Vision — no pdfplumber layer.
    All field confidences are forced to 'medium'.
    Always sets needs_review=True with a note about manual verification.
    """
    import base64 as _b64
    import time as t

    start = t.time()
    doc_type = (template_data or {}).get("doc_type", "other")

    try:
        image_data = _load_image_as_png_bytes(file_path)
        image_b64  = _b64.b64encode(image_data).decode("utf-8")

        if template_data:
            system_instruction, prompt = _build_vision_prompt(template_data, "")
        else:
            system_instruction = _get_unguided_prompt()
            prompt = (
                "Extract every key data field visible in this document image.\n"
                "Return ONLY JSON:\n"
                '{"document_type": "...", "overall_confidence": "medium", '
                '"extracted_fields": {"field_name": "value"}, "table_rows": []}'
            )

        extraction = orchestrator.llm.extract(
            image_b64=image_b64,
            prompt=prompt,
            system_instruction=system_instruction,
        )
        elapsed = (t.time() - start) * 1000

        if not extraction.success or not extraction.parsed_json:
            r = _fail(file_path.name,
                      f"Image extraction failed: {extraction.error}")
            r.processing_time_ms = elapsed
            return [r]

        raw = extraction.parsed_json

        if template_data:
            result = _process_vision_result(
                raw, template_data, file_path.name, doc_type,
                elapsed, extraction,
                "",          # no doc_text — image has no text layer
                "image upload", 0,
            )
        else:
            from orchestrator import DocumentExtractionResult
            result = DocumentExtractionResult(filename=file_path.name)
            result.document_type = raw.get("document_type", doc_type)
            result.extracted_data = {
                "document_type":    raw.get("document_type", doc_type),
                "overall_confidence": "medium",
                "extraction_method": "image_upload",
                "extracted_data": {
                    k: {"value": str(v) if v is not None else "", "confidence": "medium"}
                    for k, v in raw.get("extracted_fields", {}).items()
                },
                "table_rows": raw.get("table_rows", []),
                "validation": {
                    "flagged_count": 1,
                    "flagged_fields": [],
                    "confidence_map": {},
                    "alignment_misaligned": False,
                    "alignment_warning": "",
                },
            }
            result.extraction_response = extraction
            result.processing_time_ms  = elapsed
            result.success             = True

        # Force medium confidence + needs_review for ALL image uploads
        if result.success and result.extracted_data:
            result.extracted_data["overall_confidence"] = "medium"
            result.extracted_data["image_upload"] = True
            inner = result.extracted_data.get("extracted_data", {})
            for key in inner:
                if isinstance(inner[key], dict):
                    inner[key]["confidence"] = "medium"
            # Inject a validation flag so _run_extraction_sync sets needs_review=True
            val = result.extracted_data.setdefault("validation", {})
            val["flagged_count"] = max(1, val.get("flagged_count", 0))
            val.setdefault("flagged_fields", []).append({
                "ref": "image_upload",
                "value": "",
                "issue": (
                    "Image upload — no text validation available, "
                    "please verify extracted values manually"
                ),
            })

        print(
            f"[IMAGE] {file_path.name}: extracted "
            f"{len(result.extracted_data.get('extracted_data', {}))} fields",
            flush=True,
        )
        return [result]

    except Exception as e:
        print(f"[IMAGE] Error {file_path.name}: {e}", flush=True)
        traceback.print_exc()
        r = _fail(file_path.name, str(e))
        r.processing_time_ms = (t.time() - start) * 1000
        return [r]


def _extract_with_template_inner(orchestrator, file_path: Path, template_data: dict,
                                  selected_pages: Optional[list] = None):
    """Inner extraction logic - called by the safety wrapper."""
    import time as t
    from core.preprocessor import preprocess_file

    # ── Feature flag: clean v3 extraction engine (engine/extractor.py) ───────
    # When USE_NEW_EXTRACTOR is on, delegate to the rewritten vision-first
    # pipeline. Any failure falls back to the legacy path below (safe rollback).
    if getattr(settings, "USE_NEW_EXTRACTOR", False):
        try:
            from extractor import run_extraction as _v3_run
            print(f"[EXTRACT] {file_path.name}: using v3 engine (USE_NEW_EXTRACTOR=true)", flush=True)
            return _v3_run(orchestrator, file_path, template_data,
                           selected_pages=selected_pages)
        except Exception as _v3_err:
            print(f"[EXTRACT] v3 engine failed ({_v3_err}) — falling back to legacy", flush=True)
            traceback.print_exc()

    doc_type = template_data.get("doc_type", "other")
    mode = template_data.get("mode", "columns")
    regions = template_data.get("regions", {})
    start = t.time()
    results = []

    # H4: log a diagnostic of the template before extraction begins
    _diagnose_template(template_data)

    try:
        doc = preprocess_file(file_path)
        doc_text = doc.extracted_text or ""
        page_images = doc.page_images_b64 or []

        # Filter to user-selected pages (1-based indices)
        if selected_pages and page_images:
            filtered = [page_images[i - 1] for i in selected_pages if 0 < i <= len(page_images)]
            if filtered:
                page_images = filtered
                print(f"[EXTRACT] {file_path.name}: page filter applied → "
                      f"{selected_pages} of {len(doc.page_images_b64)} pages", flush=True)

        print(f"[EXTRACT] {file_path.name}: text_len={len(doc_text)} "
              f"pages={len(page_images)} has_vision={bool(page_images)}", flush=True)

        # Auto-classify if unknown
        if doc_type in ("other", "", None) and doc_text:
            hint = _classify_by_hints(doc_text)
            if hint:
                doc_type = hint
                template_data = {**template_data, "doc_type": doc_type}
                print(f"[EXTRACT] {file_path.name}: auto-classified -> {doc_type}", flush=True)

        # -- BINDING MAP (additive) -------------------------------------------
        # Compute the neighbor-matrix binding map BEFORE prompt building. If it
        # detects unlabeled table_data cells, extraction switches to layout mode
        # inside _build_vision_prompt. Falls back silently to region analysis.
        _use_layout = False
        try:
            _bm = compute_binding_map(template_data, template_data.get("layout", {}))
            if _bm:
                template_data = {**template_data, "binding_map": _bm}
                _use_layout = bool(_bm.get("_meta", {}).get("has_table_data"))
                print(f"[BINDING] {file_path.name}: binding map computed "
                      f"(layout_mode={_use_layout})", flush=True)
        except Exception as _bm_e:
            print(f"[BINDING] {file_path.name}: skipped ({_bm_e})", flush=True)

        # -- LAYOUT MODE ------------------------------------------------------
        if mode == "layout":
            primary_mode = regions.get("primary_mode", "form_kv")

            print(f"[EXTRACT] {file_path.name}: mode={primary_mode} doc_type={doc_type} "
                  f"pages={len(page_images)}", flush=True)

            # -- TABLE-ONLY MODE ----------------------------------------------
            # Layout-mode templates (unlabeled rows) must go to vision so the AI
            # reads and places content — skip the pdfplumber table-direct path.
            if primary_mode == "table" and not _use_layout:
                # Try pdfplumber direct first
                direct_rows = _extract_pdf_table_direct(file_path, template_data)
                elapsed = (t.time() - start) * 1000

                if direct_rows and len(direct_rows) > 0:
                    # pdfplumber succeeded - validate with AI
                    print(f"[EXTRACT] {file_path.name}: direct table {len(direct_rows)} rows", flush=True)

                    # Quick AI validation pass
                    validation = _validate_with_pdfplumber(
                        {}, doc_text, direct_rows
                    )
                    results.append(_make_table_result(
                        direct_rows, template_data, file_path.name,
                        doc_type, elapsed, "direct_pdf_validated"
                    ))
                else:
                    # pdfplumber failed - use vision extraction
                    print(f"[EXTRACT] {file_path.name}: direct failed -> vision extraction", flush=True)
                    results.extend(
                        _vision_extract_all_documents(
                            orchestrator, file_path, template_data,
                            doc_type, doc_text, page_images, start
                        )
                    )

            # -- FORM, MIXED, TWO-COLUMN MODE ---------------------------------
            else:
                results.extend(
                    _vision_extract_all_documents(
                        orchestrator, file_path, template_data,
                        doc_type, doc_text, page_images, start
                    )
                )

        # -- COLUMNS MODE (legacy) ---------------------------------------------
        else:
            results.extend(
                _legacy_columns_extract(
                    orchestrator, file_path, template_data,
                    doc_type, doc_text, page_images, start
                )
            )

    except Exception as e:
        print(f"[EXTRACT] Error {file_path.name}: {e}", flush=True)
        traceback.print_exc()
        r = _fail(file_path.name, str(e))
        r.processing_time_ms = (time.time() - start) * 1000
        results.append(r)

    return results if results else [_fail(file_path.name, "No data extracted")]


# ==============================================================================
# SECTIONED MULTI-PASS EXTRACTION (parallel column templates)
# ==============================================================================

def _detect_vertical_sections(items: list) -> list:
    """
    Split a parallel group's items into vertical sections by detecting row gaps.
    A gap of more than 2 rows between consecutive items means a section header row
    sits between them (e.g. "Current Assets" rows 2-9, gap at rows 10-13,
    "Current Liabilities" rows 14-19).
    Returns: list of sections, each a list of items.
    """
    if not items:
        return []
    sorted_items = sorted(items, key=lambda x: x["row"])
    sections: list = [[sorted_items[0]]]
    for i in range(1, len(sorted_items)):
        gap = sorted_items[i]["row"] - sorted_items[i - 1]["row"]
        if gap > 2:
            sections.append([sorted_items[i]])
        else:
            sections[-1].append(sorted_items[i])
    return sections


def _find_section_header_in_gap(layout_cells: dict, label_col: int,
                                 row_start: int, row_end: int) -> str:
    """
    Find a section header label in label_col between row_start (inclusive) and
    row_end (exclusive). Returns the last non-empty, non-extract label found.
    """
    found = ""
    for r in range(row_start, row_end):
        cell = layout_cells.get(f"{r},{label_col}")
        if not cell or not isinstance(cell, dict):
            continue
        val = str(cell.get("value") or "").strip()
        if val and not cell.get("extractTarget"):
            found = val
    return found


def _extract_parallel_groups_sectioned(
    orchestrator, template_data: dict, page_img, doc_text: str,
    filename: str, doc_type: str, start_time: float,
):
    """
    Multi-pass extraction for parallel column templates that have multiple vertical
    sections within the same column band (e.g. a balance sheet where col B is the
    value column for BOTH Current Assets rows 2-9 AND Current Liabilities rows 14-19).

    A single LLM call conflates both sections; this function makes one call per
    vertical section so each pass only sees the cells it should fill.

    Merged extracted_fields are written back via the existing T1 FieldBinding system
    (_write_form_excel / write_template_row both read from extracted_fields).

    Returns a DocumentExtractionResult on success, or None to fall back to single-pass.
    """
    regions = template_data.get("regions", {})
    layout_cells = template_data.get("layout", {}).get("cells", {})
    parallel_groups = regions.get("parallel_column_groups", [])

    if not parallel_groups:
        return None

    # Split each group's items into vertical sections
    group_sections = [_detect_vertical_sections(pg["items"]) for pg in parallel_groups]
    n_sections = max(len(s) for s in group_sections)

    if n_sections <= 1:
        return None  # single vertical section — single-pass is fine

    print(
        f"[SECTIONED] {filename}: {n_sections} vertical sections across "
        f"{len(parallel_groups)} parallel groups — using multi-pass extraction",
        flush=True,
    )

    merged_extracted_fields: dict = {}
    last_extraction = None

    for sec_idx in range(n_sections):
        # Build modified parallel groups containing only this section's items
        sectioned_groups = []
        section_name_parts = []

        for gi, pg in enumerate(parallel_groups):
            sec_list = group_sections[gi]
            sec_items = sec_list[sec_idx] if sec_idx < len(sec_list) else []
            if not sec_items:
                continue

            if sec_idx == 0:
                sec_label = pg["section_label"]
            else:
                # Find section header label in the gap rows above this section
                prev_end_row = sec_list[sec_idx - 1][-1]["row"] if (sec_idx - 1) < len(sec_list) and sec_list[sec_idx - 1] else -1
                this_start_row = sec_items[0]["row"]
                sec_label = _find_section_header_in_gap(
                    layout_cells, pg["label_col"],
                    prev_end_row + 1, this_start_row,
                )
                if not sec_label:
                    sec_label = (
                        f"Section {sec_idx + 1} "
                        f"(col {pg['label_col_letter']}-{pg['value_col_letter']})"
                    )

            sectioned_groups.append({**pg, "items": sec_items, "section_label": sec_label})
            section_name_parts.append(sec_label)

        if not sectioned_groups:
            continue

        section_display = " / ".join(section_name_parts)
        print(f"[SECTIONED] Pass {sec_idx + 1}/{n_sections}: {section_display}", flush=True)

        modified_regions = {**regions, "parallel_column_groups": sectioned_groups}
        modified_template = {**template_data, "regions": modified_regions}
        system_instruction, prompt = _build_vision_prompt(modified_template, doc_text)

        # Prepend a scope fence to prevent cross-section bleeding
        scope_prefix = (
            f"=== SECTION SCOPE (Pass {sec_idx + 1} of {n_sections}) ===\n"
            f"Extract ONLY the values for: {section_display}\n"
            f"Ignore all other sections of this document in this pass.\n"
            f"Only fill the cell references listed below — nothing else.\n\n"
        )
        prompt = scope_prefix + prompt

        extraction = None
        _base_prompt = prompt
        for attempt in range(3):
            if attempt > 0:
                prompt = (
                    "IMPORTANT: Return ONLY valid JSON with no markdown fences.\n\n"
                ) + _base_prompt
            try:
                if page_img:
                    extraction = orchestrator.llm.extract(
                        image_b64=page_img, prompt=prompt,
                        system_instruction=system_instruction,
                    )
                    if not extraction.success and doc_text:
                        extraction = orchestrator.llm.extract(
                            text=doc_text, prompt=prompt,
                            system_instruction=system_instruction,
                        )
                elif doc_text:
                    extraction = orchestrator.llm.extract(
                        text=doc_text, prompt=prompt,
                        system_instruction=system_instruction,
                    )
                else:
                    break
                if extraction and extraction.success and extraction.parsed_json:
                    break
                if attempt < 2:
                    time.sleep(5 * (3 ** attempt))
            except Exception as e:
                print(f"[SECTIONED] Pass {sec_idx + 1} attempt {attempt + 1}: {e}", flush=True)
                if attempt < 2:
                    time.sleep(5 * (3 ** attempt))

        if extraction and extraction.success and extraction.parsed_json:
            raw = extraction.parsed_json
            # Handle documents[] wrapper (LLM wraps output in documents array)
            docs = raw.get("documents", [raw])
            doc_raw = docs[0] if docs else raw
            ef = doc_raw.get("extracted_fields", {})
            merged_extracted_fields.update(ef)
            last_extraction = extraction
            print(
                f"[SECTIONED] Pass {sec_idx + 1} OK: {len(ef)} fields "
                f"(sample: {list(ef.keys())[:4]})",
                flush=True,
            )
        else:
            err = (extraction.error if extraction else "no response")[:80]
            print(f"[SECTIONED] Pass {sec_idx + 1} FAILED: {err}", flush=True)

        if sec_idx < n_sections - 1:
            time.sleep(2.0)

    if not merged_extracted_fields:
        print(f"[SECTIONED] {filename}: all passes failed — falling back to single-pass", flush=True)
        return None

    import time as _t
    elapsed = (_t.time() - start_time) * 1000
    merged_raw = {
        "extracted_fields": merged_extracted_fields,
        "overall_confidence": "high",
        "document_type": doc_type,
        "table_rows": [],
    }
    return _process_vision_result(
        merged_raw, template_data, filename, doc_type,
        elapsed, last_extraction, doc_text, "sectioned_parallel", 0,
    )



def _pdfplumber_spatial_extract(file_path: Path, regions: dict, layout: dict,
                                 page_num: int = 0) -> dict:
    """
    Spatial parallel-column extraction using pdfplumber.extract_words() instead of
    extract_text(). extract_words() preserves 2D x/y word positions; extract_text()
    linearises multi-column layouts and interleaves columns.

    Algorithm:
      1. Get every word on the page with its (x0, top, x1, bottom) coordinates.
      2. Group words into rows by y-coordinate (4-pt buckets).
      3. For each parallel column group, compute the template column's x-band
         from the layout's colWidths array.
      4. For each PDF row: collect label words (label x-band) + value words (value x-band).
      5. Match label text against template items using exact + partial Jaccard matching.
      6. Write matched value to the correct template cell ref.
      7. Write section total (row starting with "Total …") to the next_label_row cell.
      8. Clear section header value cells (they should not contain data).

    Returns {cell_ref: {"value": str, "confidence": "high"}} or {} on failure.
    """
    import re as _re
    from collections import defaultdict

    para_groups = regions.get("parallel_column_groups", [])
    if not para_groups:
        return {}

    cells      = layout.get("cells", {})
    col_widths = layout.get("colWidths", [])

    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            if page_num >= len(pdf.pages):
                page_num = 0
            page   = pdf.pages[page_num]
            pw     = float(page.width)
            words  = page.extract_words(x_tolerance=3, y_tolerance=3,
                                         keep_blank_chars=False) or []
    except Exception as exc:
        print(f"[SPATIAL] extract_words failed: {exc}", flush=True)
        return {}

    if not words:
        print(f"[SPATIAL] {file_path.name}: no words on page {page_num}", flush=True)
        return {}

    print(f"[SPATIAL] {file_path.name}: {len(words)} words, page_w={pw:.0f}", flush=True)

    # ── Row grouping ─────────────────────────────────────────────────────────
    row_map: dict = defaultdict(list)
    for w in words:
        y_key = round(float(w.get("top", 0)) / 4) * 4
        row_map[y_key].append(w)

    pdf_rows = []
    for y in sorted(row_map.keys()):
        ws = sorted(row_map[y], key=lambda w: float(w.get("x0", 0)))
        pdf_rows.append({"y": y, "words": ws, "text": " ".join(ww["text"] for ww in ws)})

    # ── FIX 2: derive column x-bands from the PDF's OWN word clusters ──────────
    # Template colWidths reflect the editor's visual layout, NOT where values
    # actually sit in the source PDF. Cluster the page's words by x-position
    # (gap-based: a gap > 20pt starts a new column) and map template columns to
    # PDF clusters left-to-right. Falls back to colWidths geometry if the cluster
    # count doesn't match the template's column count.
    def _pdf_x_clusters(all_words, gap: float = 20.0):
        clusters = []  # list of [x0, x1]
        for w in sorted(all_words, key=lambda w: float(w.get("x0", 0))):
            wx0 = float(w.get("x0", 0))
            wx1 = float(w.get("x1", wx0))
            if clusters and wx0 - clusters[-1][1] <= gap:
                clusters[-1][1] = max(clusters[-1][1], wx1)
                clusters[-1][0] = min(clusters[-1][0], wx0)
            else:
                clusters.append([wx0, wx1])
        return clusters

    # FIX B: exclude page-spanning words (title/header rows) before clustering —
    # a word wider than 60% of the page width bridges column gaps and collapses
    # distinct columns into one wide cluster.
    _cluster_words = [
        w for w in words
        if (float(w.get("x1", 0)) - float(w.get("x0", 0))) <= pw * 0.6
    ]
    _clusters = _pdf_x_clusters(_cluster_words)
    _tpl_cols = sorted({c for pg in para_groups for c in (pg["label_col"], pg["value_col"])})
    _pdf_col_bands: dict = {}
    if _tpl_cols and len(_clusters) == len(_tpl_cols):
        for _i, _tc in enumerate(_tpl_cols):
            _pdf_col_bands[_tc] = (_clusters[_i][0], _clusters[_i][1])
        print(f"[SPATIAL] PDF word-clusters → template cols {_tpl_cols}: "
              f"{[(round(_pdf_col_bands[c][0]), round(_pdf_col_bands[c][1])) for c in _tpl_cols]}",
              flush=True)
    else:
        print(f"[SPATIAL] cluster count {len(_clusters)} != template cols "
              f"{len(_tpl_cols)} — falling back to colWidths geometry", flush=True)

    # ── Template column → PDF x-band ─────────────────────────────────────────
    def col_x_band(col_idx: int):
        """Return (x_start, x_end) for a template column index.
        FIX 2: prefer PDF-derived word clusters; fall back to template colWidths."""
        if col_idx in _pdf_col_bands:
            return _pdf_col_bands[col_idx]
        if col_widths and col_idx < len(col_widths):
            total = sum(col_widths) or 1
            x0 = sum(col_widths[:col_idx]) / total * pw
            x1 = sum(col_widths[:col_idx + 1]) / total * pw
            return x0, x1
        # Fallback: divide page equally by number of columns in template
        n_cols = max((c for (r, c) in
                      (tuple(map(int, k.split(","))) for k in cells
                       if "," in k)), default=3) + 1
        w_each = pw / max(n_cols, 1)
        return col_idx * w_each, (col_idx + 1) * w_each

    def words_in_band(row_words, x0, x1, expand: float = 12.0):
        return [w for w in row_words
                if float(w.get("x0", 0)) >= x0 - expand
                and float(w.get("x1", pw)) <= x1 + expand]

    # ── Helpers ───────────────────────────────────────────────────────────────
    _NUM_RE = _re.compile(r"^[\$\-\(\)]?[\d,]+(?:\.\d{1,2})?[\)]?$")

    def is_num(text: str) -> bool:
        return bool(_NUM_RE.match(text.replace(" ", "")))

    def to_signed(text: str) -> str:
        t = text.strip()
        if t.startswith("(") and t.endswith(")"):
            inner = t[1:-1].replace(",", "").replace("$", "").strip()
            return f"-{inner}"
        return t.replace(",", "").replace("$", "").strip()

    def norm(s: str) -> str:
        return " ".join(s.upper().split())

    def jaccard(a: str, b: str) -> float:
        sa, sb = set(a.split()), set(b.split())
        if not sa or not sb:
            return 0.0
        return len(sa & sb) / len(sa | sb)

    # ── Build compact grid for section-header lookup ─────────────────────────
    compact_grid: dict = {}
    max_r = 0
    for key, cell in cells.items():
        if not isinstance(cell, dict) or "," not in key:
            continue
        try:
            r, c = map(int, key.split(","))
        except ValueError:
            continue
        max_r = max(max_r, r)
        v = (cell.get("value") or "")
        if str(v).strip():
            compact_grid[(r, c)] = str(v).strip()

    extracted_fields: dict = {}

    # ══════════════════════════════════════════════════════════════════════════
    # Per-group extraction
    # ══════════════════════════════════════════════════════════════════════════
    for pg in para_groups:
        label_col = pg["label_col"]
        value_col = pg["value_col"]
        pg_items  = pg.get("items", [])

        # Don't skip groups with no pre-defined items — dynamic fill handles blank-row templates
        lx0, lx1 = col_x_band(label_col)
        vx0, vx1 = col_x_band(value_col)

        print(f"[SPATIAL] group '{pg.get('section_label','')}'"
              f"  label_col={label_col} x={lx0:.0f}-{lx1:.0f}"
              f"  value_col={value_col} x={vx0:.0f}-{vx1:.0f}", flush=True)

        # Build expected-label map: norm(label_text) → (row, col, value_ref)
        expected: dict = {}
        for item in pg_items:
            r   = item.get("row")
            col = item.get("col")
            if r is None:
                continue
            # label text is in the template cell at (r, label_col)
            lbl = compact_grid.get((r, label_col), "")
            if not lbl:
                # also try item.label if available
                lbl = item.get("label", "")
            if lbl:
                expected[norm(lbl)] = (r, label_col)

        # Identify the Total row for this group (next non-item row below items)
        item_rows = sorted({item["row"] for item in pg_items if item.get("row") is not None})
        last_item_row = max(item_rows) if item_rows else 0
        total_template_row = None
        for r in range(last_item_row + 1, max_r + 1):
            v = compact_grid.get((r, label_col), "")
            if v and "total" in v.lower():
                total_template_row = r
                break
            if v:  # non-empty non-total row — stop looking
                break

        # Clear the section-header value cell (it's a header, not data)
        section_header_row = None
        for r in range(0, (item_rows[0] if item_rows else 0)):
            v = compact_grid.get((r, label_col), "")
            if v:
                section_header_row = r
        if section_header_row is not None:
            hdr_ref = _cell_ref(section_header_row, value_col)
            extracted_fields[hdr_ref] = {"value": "", "confidence": "high"}

        # ── Per-row matching ──────────────────────────────────────────────────
        total_found = None
        items_filled = 0

        for pdf_row in pdf_rows:
            l_words = words_in_band(pdf_row["words"], lx0, lx1)
            v_words = words_in_band(pdf_row["words"], vx0, vx1)

            # Label: non-numeric words in label band
            lbl_parts = [w["text"] for w in l_words if not is_num(w["text"])]
            if not lbl_parts:
                continue
            lbl_text = " ".join(lbl_parts).strip()
            lbl_norm = norm(lbl_text)

            # Value: numeric words in value band (take rightmost = the actual value)
            num_parts = [w["text"] for w in v_words if is_num(w["text"])]
            if not num_parts:
                continue
            raw_val = num_parts[-1]
            signed_val = to_signed(raw_val)

            # Check if this is the Total row
            if lbl_norm.startswith("TOTAL") and total_template_row is not None:
                total_ref = _cell_ref(total_template_row, value_col)
                extracted_fields[total_ref] = {"value": signed_val, "confidence": "high"}
                total_found = signed_val
                print(f"[SPATIAL]   total → {total_ref} = {signed_val}", flush=True)
                continue

            # Match label to template
            matched_row = None

            # 1. Exact match
            if lbl_norm in expected:
                matched_row = expected[lbl_norm][0]
            else:
                # 2. Jaccard similarity ≥ 0.5
                best_score = 0.0
                for exp_norm, (r, _) in expected.items():
                    score = jaccard(lbl_norm, exp_norm)
                    if score >= 0.5 and score > best_score:
                        best_score = score
                        matched_row = r

            if matched_row is not None:
                cell_ref = _cell_ref(matched_row, value_col)
                extracted_fields[cell_ref] = {"value": signed_val, "confidence": "high"}
                items_filled += 1
                print(f"[SPATIAL]   '{lbl_text}' → {cell_ref} = {signed_val}", flush=True)

        print(f"[SPATIAL] group '{pg.get('section_label','')}': "
              f"{items_filled} items, total={'found' if total_found else 'not found'}", flush=True)

        # ── Dynamic fill: blank-row slots between section headers and Total rows ──
        # When template rows are intentionally blank (user expects dynamic discovery),
        # the items list has no blank-row entries. We detect those spans here and
        # fill them with items extracted via spatial coordinates from the PDF.
        # Find all non-blank rows in this group's label column
        non_blank_label_rows = sorted(
            r for r in range(max_r + 1)
            if compact_grid.get((r, label_col))
        )
        for i in range(len(non_blank_label_rows) - 1):
            curr_r = non_blank_label_rows[i]
            next_r = non_blank_label_rows[i + 1]
            gap = next_r - curr_r - 1
            if gap < 1:
                continue
            curr_label = compact_grid.get((curr_r, label_col), "")
            next_label = compact_grid.get((next_r, label_col), "")
            # Only fill blank spans that sit BELOW a section header (not below a Total)
            if "total" in curr_label.lower():
                continue
            # Clear the section-header's value cell — it's a label, not data
            hdr_val_ref = _cell_ref(curr_r, value_col)
            extracted_fields[hdr_val_ref] = {"value": "", "confidence": "high"}
            # Check all rows in the gap are genuinely blank in label col
            all_blank = all(
                not compact_grid.get((r, label_col), "")
                for r in range(curr_r + 1, next_r)
            )
            if not all_blank:
                continue
            fill_rows = list(range(curr_r + 1, next_r))
            # Find the section header's y-position in the PDF
            section_header_y = None
            for pdf_row in pdf_rows:
                l_words = words_in_band(pdf_row["words"], lx0, lx1)
                lbl_text_s = " ".join(w["text"] for w in l_words if not is_num(w["text"])).strip()
                if lbl_text_s and jaccard(norm(lbl_text_s), norm(curr_label)) >= 0.45:
                    section_header_y = pdf_row["y"]
                    print(f"[SPATIAL-DYN] section header '{curr_label}' found at y={section_header_y:.0f}", flush=True)
                    break
            # Find the terminating row (Total or next section header) y-position
            next_label_y = None
            for pdf_row in pdf_rows:
                l_words = words_in_band(pdf_row["words"], lx0, lx1)
                lbl_text_s = " ".join(w["text"] for w in l_words if not is_num(w["text"])).strip()
                if lbl_text_s and jaccard(norm(lbl_text_s), norm(next_label)) >= 0.45:
                    next_label_y = pdf_row["y"]
                    break
            if section_header_y is None:
                print(f"[SPATIAL-DYN] section header '{curr_label}' not found in PDF — skipping", flush=True)
                continue
            # Extract all label+value pairs between section header and total
            dyn_items: list = []
            for pdf_row in pdf_rows:
                y = pdf_row["y"]
                if y <= section_header_y:
                    continue
                if next_label_y is not None and y >= next_label_y:
                    break
                l_words = words_in_band(pdf_row["words"], lx0, lx1)
                v_words = words_in_band(pdf_row["words"], vx0, vx1)
                lbl_parts = [w["text"] for w in l_words if not is_num(w["text"])]
                num_parts = [w["text"] for w in v_words if is_num(w["text"])]
                if not lbl_parts or not num_parts:
                    continue
                lbl_text_d = " ".join(lbl_parts).strip()
                val_text_d = to_signed(num_parts[-1])
                # Skip subtotal/total lines inside the section
                if norm(lbl_text_d).startswith("TOTAL") or norm(lbl_text_d).startswith("SUBTOTAL"):
                    continue
                dyn_items.append((lbl_text_d, val_text_d))
            print(
                f"[SPATIAL-DYN] '{curr_label}' → {len(dyn_items)} items "
                f"for {len(fill_rows)} slots",
                flush=True,
            )
            for slot_idx, (item_label, item_value) in enumerate(dyn_items):
                if slot_idx >= len(fill_rows):
                    print(f"[SPATIAL-DYN] '{curr_label}': overflow — {len(dyn_items)} items but only {len(fill_rows)} slots", flush=True)
                    break
                row = fill_rows[slot_idx]
                label_ref = _cell_ref(row, label_col)
                value_ref = _cell_ref(row, value_col)
                extracted_fields[label_ref] = {"value": item_label,  "confidence": "high"}
                extracted_fields[value_ref] = {"value": item_value,  "confidence": "high"}
                print(f"[SPATIAL-DYN]   '{item_label}' → {value_ref} = {item_value}", flush=True)

    print(f"[SPATIAL] total extracted: {len(extracted_fields)} fields", flush=True)
    return extracted_fields

def _pdfplumber_extract_dynamic_parallel(doc_text: str, regions: dict, layout: dict) -> dict:
    """
    Extract values for dynamic-fill rows in parallel-column templates.

    When a parallel-column template has empty data rows between section headers and
    Total rows (e.g. a balance sheet where rows 1-8 are blank placeholders for
    Current Assets line items), this function:
      1. Scans each group's label column for contiguous empty-row spans (≥ 2 rows).
      2. Splits the pdfplumber text into named sections using those header labels.
      3. Maps extracted (item_label, amount) pairs into the template cells.

    Returns {cell_ref: {"value": str, "confidence": "high"}} for all filled cells,
    or {} when no dynamic zones can be matched.
    """
    para_groups = regions.get("parallel_column_groups", [])
    if not para_groups:
        return {}

    cells = layout.get("cells", {})
    if not cells:
        return {}

    # Build compact grid (row, col) -> value_str from layout cells
    compact_grid: dict = {}
    max_row = 0
    for key, cell in cells.items():
        if not isinstance(cell, dict):
            continue
        parts = key.split(",")
        if len(parts) != 2:
            continue
        try:
            r, c = int(parts[0]), int(parts[1])
        except (ValueError, TypeError):
            continue
        max_row = max(max_row, r)
        raw_val = cell.get("value")
        val_str = str(raw_val).strip() if raw_val is not None else ""
        if val_str:
            compact_grid[(r, c)] = val_str

    # Find dynamic fill zones for each group.
    # A dynamic fill zone = ≥2 consecutive empty rows in the label column
    # sandwiched between two non-empty rows.
    dynamic_zones = []  # list of zone dicts

    for pg in para_groups:
        label_col = pg["label_col"]
        value_col = pg["value_col"]

        # All non-empty rows in this group's label column, sorted ascending
        label_rows = [
            (r, compact_grid[(r, label_col)])
            for r in range(max_row + 1)
            if compact_grid.get((r, label_col), "")
        ]

        if len(label_rows) < 2:
            continue

        for i in range(len(label_rows) - 1):
            curr_r, curr_val = label_rows[i]
            next_r, next_val = label_rows[i + 1]

            empty_span = next_r - curr_r - 1
            if empty_span < 2:
                continue  # no meaningful dynamic zone here

            curr_lc = curr_val.lower().strip()
            is_total = curr_lc in ("total", "grand total", "subtotal", "final total")

            if is_total:
                # curr_r is a Total row (end of previous section).
                # The empty span that follows belongs to the next section whose
                # header is next_val — but only if next_val is itself non-Total.
                next_lc = next_val.lower().strip()
                if next_lc in ("total", "grand total", "subtotal", "final total"):
                    continue
                zone_label = next_val
                zone_header_row = next_r   # the section label IS next_r
            else:
                # curr_val is a section header; the empty span is its data zone.
                zone_label = curr_val
                zone_header_row = curr_r   # section label is curr_r

            fill_rows = list(range(curr_r + 1, next_r))
            dynamic_zones.append({
                "group":           pg,
                "zone_label":      zone_label,
                "label_col":       label_col,
                "value_col":       value_col,
                "fill_rows":       fill_rows,
                "zone_header_row": zone_header_row,  # row of section header in template
                "next_label_row":  next_r,           # Total row that terminates this zone
            })

    if not dynamic_zones:
        return {}

    print(
        f"[PLUMBER-DYN] {len(dynamic_zones)} dynamic fill zones: "
        + ", ".join(
            f"'{z['zone_label']}' col {chr(65+z['label_col'])}-{chr(65+z['value_col'])} "
            f"({len(z['fill_rows'])} slots)"
            for z in dynamic_zones
        ),
        flush=True,
    )

    # --- Parse pdfplumber text into named sections ----------------------------
    zone_labels = list({z["zone_label"] for z in dynamic_zones})

    def _norm(s: str) -> str:
        n = re.sub(r'[^\w\s]', ' ', s)
        return re.sub(r'\s+', ' ', n).lower().strip()

    norm_zone_map = {_norm(lbl): lbl for lbl in zone_labels if lbl}


    val_re_dyn = re.compile(
        r'^(.*?)\s+\(?\$?([-]?[0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{0,2})?)\)?$'
    )
    tab_re_dyn = re.compile(
        r'^(.*?)\t\$?([-]?[0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{0,2})?)$'
    )

    # Semantic type-keywords: used so "LONG-TERM LIABILITIES" maps to the
    # "Non current liabilities" zone even though Jaccard is too low for a strict match.
    _TYPE_STEMS = ["asset", "liabilit", "equity", "stockholder", "shareholder"]

    def _type_stem(label: str):
        n = _norm(label)
        for stem in _TYPE_STEMS:
            if stem in n:
                return stem
        return None

    current_section: str = None
    pdf_sections: dict = {}        # zone_label → [(item_label, cleaned_value)]
    pdf_section_totals: dict = {}  # zone_label → total_value_string

    for raw_line in doc_text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("---"):
            continue

        # --- Section header detection (primary + type-stem fallback) ---
        # MUST run before the `current_section is None` guard so that headings
        # like "LONG-TERM LIABILITIES" can open a zone even when current_section
        # was just closed by a preceding Total line.
        norm_line = _norm(line)
        norm_line_words = set(norm_line.split())
        matched = None
        best_score = 0.0
        for norm_lbl, orig_lbl in norm_zone_map.items():
            if not norm_lbl:
                continue
            norm_lbl_words = set(norm_lbl.split())
            if not norm_lbl_words or not norm_line_words:
                continue
            wc_ratio = (min(len(norm_lbl_words), len(norm_line_words)) /
                        max(len(norm_lbl_words), len(norm_line_words)))
            if wc_ratio < 0.7:
                continue  # word counts too different → skip
            intersection = norm_lbl_words & norm_line_words
            union = norm_lbl_words | norm_line_words
            score = len(intersection) / len(union) if union else 0.0
            if score >= 0.5 and score > best_score:
                best_score = score
                matched = orig_lbl

        # Type-stem fallback: ALL-CAPS headings that fail primary Jaccard
        # (e.g. "LONG-TERM LIABILITIES" → "Non current liabilities")
        # Guard: compound titles like "LIABILITIES & EQUITY" are section dividers,
        # not zone headers — skip them.
        if not matched and "&" not in line:
            alpha_chars = [ch for ch in line if ch.isalpha()]
            if alpha_chars and (sum(1 for ch in alpha_chars if ch.isupper()) /
                                len(alpha_chars)) > 0.5:
                line_stem = _type_stem(norm_line)
                if line_stem:
                    for norm_lbl, orig_lbl in norm_zone_map.items():
                        # Only skip a zone that's already been opened AND has items.
                        # An empty section (0 items) can be re-opened by a better match.
                        if orig_lbl in pdf_sections and pdf_sections[orig_lbl]:
                            continue
                        if line_stem in _norm(orig_lbl):
                            matched = orig_lbl
                            print(f"[PLUMBER-DYN] type-open: '{line}' → '{orig_lbl}'",
                                  flush=True)
                            break

        if matched:
            current_section = matched
            pdf_sections.setdefault(current_section, [])
            continue

        # Skip non-header lines when no section is open
        if current_section is None:
            continue

        m = tab_re_dyn.match(line) or val_re_dyn.match(line)
        if not m:
            # Non-numeric line inside an open section — if it looks like another
            # all-caps heading that wasn't matched, close the current section.
            alpha_chars = [ch for ch in line if ch.isalpha()]
            if alpha_chars and (sum(1 for ch in alpha_chars if ch.isupper()) /
                                len(alpha_chars)) > 0.5:
                current_section = None
            continue

        lbl_raw = m.group(1).strip().rstrip(":").strip()
        val_raw = m.group(2).strip().lstrip("$").strip()
        if not lbl_raw or not val_raw:
            continue

        # Total/subtotal lines — CAPTURE value and close section.
        # This lets us write the total to the template's Total row (B10, D10, etc.).
        norm_item = _norm(lbl_raw)
        if (norm_item in ("total", "grand total", "subtotal")
                or norm_item.startswith("total ")
                or "subtotal" in norm_item):
            val_clean_tot = val_raw.replace(",", "")
            if re.search(r'\(\$?[\d,]+(?:\.\d{1,2})?\)', line) and not val_clean_tot.startswith("-"):
                val_clean_tot = "-" + val_clean_tot
            pdf_section_totals[current_section] = val_clean_tot
            current_section = None  # close section after its total row
            continue

        # Detect accounting-negative parentheses in the original line
        val_clean = val_raw.replace(",", "")
        if re.search(r'\(\$?[\d,]+(?:\.\d{1,2})?\)', line) and not val_clean.startswith("-"):
            val_clean = "-" + val_clean

        # "Less: ..." items represent deductions — store as negative
        if lbl_raw.lower().startswith(("less:", "less ", "less-")):
            if not val_clean.startswith("-"):
                val_clean = "-" + val_clean

        pdf_sections[current_section].append((lbl_raw, val_clean))

    if not pdf_sections:
        print("[PLUMBER-DYN] no PDF sections parsed — skipping dynamic fill", flush=True)
        return {}

    print(
        "[PLUMBER-DYN] PDF sections: "
        + ", ".join(f"'{k}': {len(v)} items" for k, v in pdf_sections.items()),
        flush=True,
    )

    # --- Map PDF sections to dynamic zones and fill cells --------------------
    extracted_fields: dict = {}

    matched_pdf_sections: set = set()  # track which PDF sections already used

    for zone in dynamic_zones:
        zone_label      = zone["zone_label"]
        label_col       = zone["label_col"]
        value_col       = zone["value_col"]
        fill_rows       = zone["fill_rows"]
        zone_header_row = zone.get("zone_header_row")
        next_label_row  = zone.get("next_label_row")

        # Clear the zone header's value cell so form-extraction's wrong guess
        # (e.g. total amount sitting next to the section header label) is overridden.
        if zone_header_row is not None:
            hdr_val_ref = _cell_ref(zone_header_row, value_col)
            extracted_fields[hdr_val_ref] = {"value": "", "confidence": "high"}

        # Find the best-matching PDF section for this zone.
        # Primary: word-level Jaccard with word-count-ratio guard.
        norm_zone = _norm(zone_label)
        norm_zone_words = set(norm_zone.split())
        pdf_items = None
        matched_pdf_key = None
        best_score = 0.0
        for pdf_sec_name, items in pdf_sections.items():
            if pdf_sec_name in matched_pdf_sections:
                continue
            norm_pdf = _norm(pdf_sec_name)
            norm_pdf_words = set(norm_pdf.split())
            if not norm_zone_words or not norm_pdf_words:
                continue
            wc_ratio = (min(len(norm_zone_words), len(norm_pdf_words)) /
                        max(len(norm_zone_words), len(norm_pdf_words)))
            if wc_ratio < 0.7:
                continue
            intersection = norm_zone_words & norm_pdf_words
            union = norm_zone_words | norm_pdf_words
            score = len(intersection) / len(union) if union else 0.0
            if score >= 0.6 and score > best_score:
                best_score = score
                pdf_items = items
                matched_pdf_key = pdf_sec_name

        if not pdf_items:
            print(f"[PLUMBER-DYN] no PDF match for zone '{zone_label}'", flush=True)
        else:
            matched_pdf_sections.add(matched_pdf_key)
            print(
                f"[PLUMBER-DYN] zone '{zone_label}': "
                f"{len(pdf_items)} PDF items → {len(fill_rows)} template slots",
                flush=True,
            )

            for slot_idx, (item_label, item_value) in enumerate(pdf_items):
                if slot_idx >= len(fill_rows):
                    break
                row = fill_rows[slot_idx]
                label_ref = _cell_ref(row, label_col)
                value_ref = _cell_ref(row, value_col)
                extracted_fields[label_ref] = {"value": item_label,  "confidence": "high"}
                extracted_fields[value_ref] = {"value": item_value,  "confidence": "high"}

        # Write the section total to the template's Total row value cell.
        # This fills B10, D10, B21, D21 (empty cells next to "Total" labels).
        section_total = (pdf_section_totals.get(zone_label)
                         or pdf_section_totals.get(matched_pdf_key))
        if section_total and next_label_row is not None:
            total_val_ref = _cell_ref(next_label_row, value_col)
            extracted_fields[total_val_ref] = {"value": section_total, "confidence": "high"}
            print(f"[PLUMBER-DYN] total '{zone_label}': {section_total} → {total_val_ref}",
                  flush=True)

    # --- Grand totals: read from PDF, map semantically to Final Total cells -----
    # Scan doc_text for ALL-CAPS lines that carry a total-type value
    # (e.g. "TOTAL ASSETS $1,608,600", "TOTAL LIABILITIES $557,400").
    # These are explicitly in the PDF — we don't compute them.
    grand_total_lines = []   # [(norm_label, raw_value_str), ...]
    for raw_line in doc_text.splitlines():
        line = raw_line.strip()
        gm = tab_re_dyn.match(line) or val_re_dyn.match(line)
        if not gm:
            continue
        lbl_raw_g = gm.group(1).strip()
        val_raw_g = gm.group(2).strip().lstrip("$").replace(",", "")
        alpha_g = [c for c in lbl_raw_g if c.isalpha()]
        if not alpha_g:
            continue
        # Must be predominantly uppercase (grand-total headers are ALL-CAPS)
        if sum(1 for c in alpha_g if c.isupper()) / len(alpha_g) < 0.6:
            continue
        norm_g = _norm(lbl_raw_g)
        # Skip lines already captured as zone section headers or items
        if norm_g in norm_zone_map:
            continue
        # Must look like a total (contains "total", "sum", or similar)
        if not any(w in norm_g for w in ("total", "sum", "aggregate")):
            continue
        grand_total_lines.append((norm_g, val_raw_g))

    # Find "Final Total" / "Grand Total" label cells in the template.
    # The value cell sits one column to the right of the label cell.
    ft_cells = []    # [(row_0idx, label_col, value_col), ...]
    for (r, c), v in compact_grid.items():
        if _norm(str(v)) in ("final total", "grand total"):
            ft_cells.append((r, c, c + 1))
    ft_cells.sort(key=lambda x: x[0])   # ascending row order

    if ft_cells and grand_total_lines:
        # Semantic matching: for each grand total line, determine its semantic type
        # ("asset", "liabilit", "equity", …) using the same _type_stem helper.
        # Then find the unique next_label_row among zones that share that type,
        # and write the value to the Final Total cell at that row.
        written_ft = set()   # prevent double-writing the same cell

        for norm_g, val_g in grand_total_lines:
            gt_stem = _type_stem(norm_g)
            target_nlr = None

            if gt_stem:
                # Collect all next_label_rows for zones whose label contains this stem
                candidate_nlrs = {
                    z["next_label_row"]
                    for z in dynamic_zones
                    if gt_stem in _norm(z["zone_label"])
                    and z.get("next_label_row") is not None
                }
                if len(candidate_nlrs) == 1:
                    target_nlr = next(iter(candidate_nlrs))

            if target_nlr is None:
                continue

            for (r, lc, vc) in ft_cells:
                if r == target_nlr:
                    ft_ref = _cell_ref(r, vc)
                    if ft_ref not in written_ft:   # first match wins
                        extracted_fields[ft_ref] = {"value": val_g, "confidence": "high"}
                        written_ft.add(ft_ref)
                        print(
                            f"[PLUMBER-DYN] grand_total '{norm_g}' "
                            f"(stem={gt_stem}) → {ft_ref}: {val_g}",
                            flush=True,
                        )
                    break

    n_filled = sum(1 for v in extracted_fields.values()
                   if isinstance(v, dict) and v.get("value"))
    print(f"[PLUMBER-DYN] filled {n_filled} cells across {len(dynamic_zones)} zones", flush=True)
    return extracted_fields


def _pdfplumber_extract_form_fields(doc_text: str, regions: dict) -> dict:
    """
    Directly extract form field values from pdfplumber text using label matching.

    For text-based PDFs (not scanned) with kv_pair or parallel_group templates,
    this avoids calling the LLM entirely:
      1. Parse the document text into (label, value) pairs via regex.
      2. Match each template kv_pair label to a document label (fuzzy).
      3. Return {cell_ref: {"value": str, "confidence": str}} for every matched field.

    Returns {} if extraction fails or doc_text is empty.
    """
    if not doc_text or not doc_text.strip():
        return {}

    kv_pairs    = regions.get("kv_pairs", [])
    para_groups = regions.get("parallel_column_groups", [])

    # Collect all template (label → value_ref) mappings
    tpl_items = []  # list of (label_str, value_ref)
    if para_groups:
        for pg in para_groups:
            for item in pg.get("items", []):
                lbl = item.get("label", "").strip()
                ref = item.get("value_ref", "")
                if lbl and ref:
                    tpl_items.append((lbl, ref))
    elif kv_pairs:
        for kv in kv_pairs:
            lbl = kv.get("label", "").strip()
            ref = kv.get("value_ref", "")
            if lbl and ref:
                tpl_items.append((lbl, ref))
    else:
        # form_with_targets: cells marked "Extract here" with no adjacent label column
        for et in regions.get("explicit_targets", []):
            lbl = et.get("label", "").strip()
            ref = et.get("ref", "")
            if lbl and ref and not lbl.startswith("field at "):
                tpl_items.append((lbl, ref))

    if not tpl_items:
        return {}

    # --- Parse document text into {normalized_label: numeric_string} ---
    # Matches lines like:
    #   "Cash & Cash Equivalents $184,320"
    #   "Less: Accum. Depreciation $139,800"
    #   "Total Non-Current Assets 248,800"   (no $ sign)
    doc_pairs: dict = {}
    # tab-separated (pdfplumber tables)
    tab_re = re.compile(
        r'^(.*?)\t\$?([-]?[0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{0,2})?)$'
    )
    # space-separated with $, optional accounting parens
    dollar_re = re.compile(
        r'^(.*?)\s+\$\(?([-]?[0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{0,2})?)\)?$'
    )
    # fallback: any whitespace separator, optional $, optional parens
    # tried last so dollar_re takes priority for the $ case
    val_re = re.compile(
        r'^(.*?)\s+\(?\$?([-]?[0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{0,2})?)\)?$'
    )

    for raw_line in doc_text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("---"):
            continue
        m = tab_re.match(line) or dollar_re.match(line) or val_re.match(line)
        if not m:
            continue
        label_raw = m.group(1).strip().rstrip(":").strip()
        value_raw = m.group(2).strip().lstrip("$").strip()
        if not label_raw or not value_raw:
            continue
        # Normalize whitespace and lowercase for matching
        norm = re.sub(r'[^\w\s&.()/]', ' ', label_raw)
        norm = re.sub(r'\s+', ' ', norm).lower().strip()
        # Handle accounting negatives in original line: "(139,800)" → negate
        stripped = value_raw.replace(",", "")
        if re.search(r'\(\$?' + re.escape(stripped) + r'\)', line):
            value_raw = '-' + value_raw
        # Remove commas so downstream numeric conversion is simple
        value_raw = value_raw.replace(",", "")
        if norm and value_raw:
            doc_pairs[norm] = value_raw

    if not doc_pairs:
        return {}

    def _norm_label(s: str) -> str:
        n = re.sub(r'[^\w\s&.()/]', ' ', s)
        return re.sub(r'\s+', ' ', n).lower().strip()

    def _should_negate(label: str) -> bool:
        """Labels that represent deductions from a total should be stored negative."""
        ln = label.lower().strip()
        return ln.startswith(("less:", "less ", "less-"))

    # Terms that are too generic to fuzzy-match — they appear in many labels
    # (e.g. "Total" would match "Total Current Assets", "Total Liabilities", etc.)
    # Short labels NOT in this set (like "Cash", "Inventory") get fuzzy matching.
    _GENERIC_NO_FUZZY = frozenset({
        "total", "amount", "balance", "net", "gross", "subtotal", "value",
        "debit", "credit", "description", "item", "quantity", "rate",
        "date", "name", "number", "ref", "other", "summary", "remarks",
        "assets", "liabilities", "equity",
    })

    extracted_fields: dict = {}
    matched_refs: set = set()

    for tpl_label, value_ref in tpl_items:
        if value_ref in matched_refs:
            continue
        norm_tpl = _norm_label(tpl_label)

        matched_val  = None
        matched_conf = None

        tpl_words_list = norm_tpl.split()
        # Only skip fuzzy matching for known-ambiguous generic financial terms.
        # Short but specific labels (Cash, Inventory, Depreciation) get fuzzy matching.
        is_generic = norm_tpl in _GENERIC_NO_FUZZY

        # 1. Exact match (always tried, even for generic labels)
        if norm_tpl in doc_pairs:
            matched_val  = doc_pairs[norm_tpl]
            matched_conf = "high"
        elif not is_generic:
            # 2. Substring containment — template label is a substring of doc label
            #    or doc label is a substring of template label.
            #    Length-biased: prefer longer doc labels (more specific).
            best_label = None
            for doc_label in doc_pairs:
                if norm_tpl in doc_label or doc_label in norm_tpl:
                    if best_label is None or len(doc_label) > len(best_label):
                        best_label = doc_label
            if best_label:
                matched_val  = doc_pairs[best_label]
                matched_conf = "high"
            else:
                # 3. Word-overlap match (>= 75% of template words found in doc label)
                tpl_words   = set(tpl_words_list)
                best_score  = 0.0
                best_label  = None
                for doc_label in doc_pairs:
                    doc_words = set(doc_label.split())
                    if not tpl_words:
                        continue
                    overlap = len(tpl_words & doc_words) / len(tpl_words)
                    if overlap > best_score:
                        best_score = overlap
                        best_label = doc_label
                if best_score >= 0.75 and best_label:
                    matched_val  = doc_pairs[best_label]
                    matched_conf = "medium"

        if matched_val is not None:
            # Apply "Less: ..." negation rule (deduction line items stored as negative)
            if _should_negate(tpl_label) and matched_val and not matched_val.startswith("-"):
                matched_val = "-" + matched_val
            extracted_fields[value_ref] = {
                "value":      matched_val,
                "confidence": matched_conf,
            }
            matched_refs.add(value_ref)

    n_filled = sum(1 for v in extracted_fields.values()
                   if isinstance(v, dict) and v.get("value"))
    n_total  = len(tpl_items)
    print(
        f"[PLUMBER] pdfplumber form extraction: {n_filled}/{n_total} fields matched",
        flush=True,
    )
    return extracted_fields


def _vision_extract_all_documents(orchestrator, file_path, template_data,
                                   doc_type, doc_text, page_images, start):
    """
    Extract all documents from a PDF.
    Uses vision (images) when available, falls back to text when not.
    Handles: single doc, multi-page, multiple docs per page, 100 cheques in one PDF.
    """
    import time as t
    results = []
    regions = template_data.get("regions", {})
    has_images = bool(page_images)

    # ── Document boundary detection ──────────────────────────────────────────────
    # Multi-page PDFs: each page IS one document (one invoice per page, one statement
    # per page, etc.) — no LLM detection needed. Skip the vision call and build N
    # segments directly, one per page. This saves one expensive LLM call per batch.
    #
    # Single-page PDFs: may contain multiple documents on one page (e.g. two cheques
    # on one sheet). Use LLM vision to detect boundaries in that case only.
    #
    # Exception: financial docs that naturally span multiple pages (audit_report,
    # balance_sheet, income_statement) are kept as ONE segment even if multi-page.
    _MULTI_PAGE_DOCS = {"audit_report", "balance_sheet", "income_statement", "tax_form"}
    _is_multi_page = has_images and len(page_images) > 1
    _is_spanning_doc = (doc_type or "").lower() in _MULTI_PAGE_DOCS

    if _is_multi_page and not _is_spanning_doc:
        # One segment per page — no LLM boundary detection needed
        doc_segments = [
            {"index": i, "page_indices": [i], "hint": f"page {i+1}", "sub_index": None, "total_on_page": 1}
            for i in range(len(page_images))
        ]
        print(f"[EXTRACT] {file_path.name}: {len(doc_segments)} pages → {len(doc_segments)} doc segments (no LLM detection)", flush=True)
    elif has_images:
        # Single page OR spanning doc type: use LLM to detect boundaries
        doc_segments = _detect_document_boundaries_vision(
            page_images, orchestrator, file_path.name, doc_type
        )
    else:
        # No images - treat as single document, use text extraction
        print(f"[EXTRACT] {file_path.name}: no page images, using text-only extraction", flush=True)
        doc_segments = [{"index": 0, "page_indices": [], "hint": "full document"}]

    total_docs = len(doc_segments)
    print(f"[EXTRACT] {file_path.name}: {total_docs} document(s) to process", flush=True)

    for seg in doc_segments:
        seg_index = seg["index"]
        seg_hint = seg.get("hint", f"doc {seg_index+1}")
        page_indices = seg.get("page_indices", [0])
        sub_index = seg.get("sub_index", None)
        total_on_page = seg.get("total_on_page", 1)

        # Get page image for this segment if available
        page_img = None
        if has_images and page_indices:
            idx = page_indices[0] if page_indices[0] < len(page_images) else 0
            page_img = page_images[idx]
        elif has_images and page_images:
            page_img = page_images[0]

        # ── pdfplumber-first path for text-based documents ───────────────────────
        # For form/parallel_groups templates where the PDF has extractable text,
        # match template labels directly to document values without any LLM call.
        # Falls through to LLM only when coverage is insufficient.
        # IMPORTANT: Skip pdfplumber entirely for scanned/image-only PDFs.
        # pdfplumber returns empty/garbage for scanned docs, which causes all
        # correctly LLM-extracted fields to be flagged as low confidence.
        primary_mode = regions.get("primary_mode", "form_kv")
        _pdf_has_text = bool(doc_text) and _is_text_pdf(file_path)
        # Layout-mode templates (unlabeled line-item rows, table_data cells) MUST go
        # to the layout-aware LLM call — pdfplumber can only produce cell-ref
        # extracted_fields and cannot reconstruct layout_sections, so letting it win
        # here silently drops the layout output. Skip the pdfplumber-first path.
        _is_layout_mode = bool(
            template_data.get("binding_map", {}).get("_meta", {}).get("has_table_data")
        )
        if not _pdf_has_text and doc_text:
            print(f"[PLUMBER] {file_path.name}: no extractable text — skipping pdfplumber, LLM only", flush=True)
        if _is_layout_mode:
            print(f"[PLUMBER] {file_path.name}: layout-mode template — skipping pdfplumber-first, using layout LLM", flush=True)
        if (_pdf_has_text and not _is_layout_mode
                and primary_mode in ("form_kv", "form_with_targets", "parallel_groups", "mixed")):
            _plumber_ef = _pdfplumber_extract_form_fields(doc_text, regions)

            # For parallel_groups AND mixed-mode templates that have parallel column groups
            # (e.g. balance sheet classified as mixed because header rows look like table headers),
            # also extract dynamic fill zones — empty row spans between section headers.
            _dyn_ef: dict = {}
            if primary_mode in ("parallel_groups", "mixed") and regions.get("parallel_column_groups"):
                # Try spatial extraction first (uses extract_words() + x/y coordinates —
                # much more accurate for multi-column layouts like balance sheets).
                # Falls back to regex text-based extraction if spatial returns nothing.
                _dyn_ef = _pdfplumber_spatial_extract(
                    file_path, regions, template_data.get("layout", {}),
                    page_num=(seg.get("page_indices", [0]) or [0])[0],
                )
                if not _dyn_ef:
                    print(f"[SPATIAL] {file_path.name}: no spatial results, falling back to text-based", flush=True)
                    _dyn_ef = _pdfplumber_extract_dynamic_parallel(
                        doc_text, regions, template_data.get("layout", {})
                    )
                if _dyn_ef:
                    # Merge: dynamic fill wins for any cell it covers (more specific)
                    _plumber_ef = {**_plumber_ef, **_dyn_ef}

            n_plumber   = sum(1 for v in _plumber_ef.values()
                              if isinstance(v, dict) and v.get("value"))
            # Count total template fields to compute coverage ratio
            _all_items  = regions.get("kv_pairs", [])
            if regions.get("parallel_column_groups"):
                _all_items = [
                    item
                    for pg in regions["parallel_column_groups"]
                    for item in pg.get("items", [])
                ]
            if not _all_items:
                _all_items = regions.get("explicit_targets", [])
            n_tpl = len(_all_items)
            coverage = n_plumber / n_tpl if n_tpl else 0
            # Use pdfplumber result when label-matched coverage is high enough
            # OR when dynamic fill found REAL values (not just empty header-clears).
            # FIX A: an all-empty _dyn_ef (e.g. spatial clustering failed) must NOT
            # count as success — otherwise the empty result is used and the LLM is
            # skipped, producing a template with all value cells blank.
            has_real_values = any(
                str(v.get("value", "")).strip()
                for v in _dyn_ef.values()
                if isinstance(v, dict)
            )
            use_plumber = coverage >= 0.5 or has_real_values
            print(
                f"[PLUMBER] {file_path.name}: coverage {n_plumber}/{n_tpl} "
                f"({coverage:.0%}) dyn={len(_dyn_ef)} — "
                + ("using pdfplumber result" if use_plumber else "falling back to LLM"),
                flush=True,
            )
            if use_plumber:
                import time as _t2
                elapsed = (_t2.time() - start) * 1000
                raw_plumber = {
                    "extracted_fields": _plumber_ef,
                    "overall_confidence": "high",
                    "document_type":     doc_type,
                    "table_rows":        [],
                }
                result = _process_vision_result(
                    raw_plumber, template_data, file_path.name, doc_type,
                    elapsed, None, doc_text, "pdfplumber_form", seg_index,
                )
                seg_fn = (file_path.name if total_docs == 1
                          else f"{file_path.stem}_doc{seg_index+1}{file_path.suffix}")
                result.filename = seg_fn
                results.append(result)
                continue  # no LLM call needed

        # Sectioned multi-pass extraction for parallel column templates with multiple
        # vertical sections (e.g. balance sheet where col B covers both Current Assets
        # rows 2-9 AND Current Liabilities rows 14-19 in one column band).
        # Each vertical section gets its own LLM call to prevent cross-section misrouting.
        if primary_mode == "parallel_groups":
            sectioned = _extract_parallel_groups_sectioned(
                orchestrator, template_data, page_img, doc_text,
                file_path.name, doc_type, start,
            )
            if sectioned is not None:
                seg_fn = (file_path.name if total_docs == 1
                          else f"{file_path.stem}_doc{seg_index+1}{file_path.suffix}")
                sectioned.filename = seg_fn
                results.append(sectioned)
                continue  # skip the single-pass extraction for this segment

        # Build extraction prompt — returns (system_instruction, user_prompt)
        # system_instruction = registry expert persona (stable, cached by Gemini)
        # user_prompt = template fields + doc text (variable per document)
        system_instruction, prompt = _build_vision_prompt(template_data, doc_text)

        # Add sub-document context for multiple docs on same page
        if total_on_page > 1 and sub_index is not None:
            prompt = (
                f"=== DOCUMENT CONTEXT ===\n"
                f"This page contains {total_on_page} separate documents.\n"
                f"Extract ONLY document #{sub_index+1} (index {sub_index}).\n"
                f"Description: {seg_hint}\n\n"
            ) + prompt

        # Extract - use image if available, text otherwise
        # Retry up to 3 times with exponential backoff to handle 429 rate limits
        extraction = None
        last_error = ""
        _base_prompt = prompt
        for attempt in range(3):
            # H3: on retry, prepend a JSON correction instruction to the prompt
            if attempt > 0:
                prompt = (
                    "IMPORTANT: Your previous response could not be parsed as JSON. "
                    "Return ONLY valid JSON with no markdown fences, no explanation, "
                    "no text before or after the JSON object.\n\n"
                ) + _base_prompt
            try:
                if page_img:
                    print(f"[EXTRACT] {file_path.name}: sending image to AI "
                          f"(attempt {attempt+1})", flush=True)
                    extraction = orchestrator.llm.extract(
                        image_b64=page_img, prompt=prompt,
                        system_instruction=system_instruction)
                    if not extraction.success and doc_text:
                        print(f"[EXTRACT] {file_path.name}: vision failed -> text fallback", flush=True)
                        extraction = orchestrator.llm.extract(
                            text=doc_text, prompt=prompt,
                            system_instruction=system_instruction)
                elif doc_text:
                    print(f"[EXTRACT] {file_path.name}: sending text to AI "
                          f"(attempt {attempt+1})", flush=True)
                    extraction = orchestrator.llm.extract(
                        text=doc_text, prompt=prompt,
                        system_instruction=system_instruction)
                else:
                    seg_fn = (file_path.name if total_docs == 1
                              else f"{file_path.stem}_doc{seg_index+1}{file_path.suffix}")
                    r = _fail(seg_fn, "No content - PDF has no extractable text or images")
                    results.append(r)
                    extraction = None
                    break

                if extraction and extraction.success and extraction.parsed_json:
                    break  # success

                last_error = (extraction.error if extraction else "no response")[:100]
                print(f"[EXTRACT] {file_path.name}: attempt {attempt+1} failed "
                      f"({last_error}) - "
                      f"{'retrying' if attempt < 2 else 'giving up'}", flush=True)
                if attempt < 2:
                    wait = 5 * (3 ** attempt)  # 5s, 15s
                    print(f"[EXTRACT] waiting {wait}s before retry", flush=True)
                    time.sleep(wait)

            except Exception as e:
                last_error = str(e)[:100]
                print(f"[EXTRACT] LLM error attempt {attempt+1}: {e}", flush=True)
                if attempt < 2:
                    time.sleep(5 * (3 ** attempt))

        if extraction is None:
            continue  # already added fail result above

        if not extraction.success or not extraction.parsed_json:
            seg_fn = (file_path.name if total_docs == 1
                      else f"{file_path.stem}_doc{seg_index+1}{file_path.suffix}")
            r = _fail(seg_fn, f"All retries failed: {last_error}")
            r.processing_time_ms = (t.time() - start) * 1000
            results.append(r)
            continue

        raw = extraction.parsed_json
        elapsed = (t.time() - start) * 1000

        # ── ISSUE 1: layout-mode format enforcement ──────────────────────────────
        # A weaker fallback model (e.g. gemini-flash-lite-latest after 503s) may
        # return extracted_fields instead of the required layout_sections format.
        # Retry once with a stronger directive, then fall back to converting the
        # extracted_fields into layout_sections via the binding map.
        if _is_layout_mode:
            def _unwrap_doc(_r):
                if (isinstance(_r, dict) and isinstance(_r.get("documents"), list)
                        and _r["documents"] and isinstance(_r["documents"][0], dict)):
                    return _r["documents"][0]
                return _r if isinstance(_r, dict) else {}
            _d0 = _unwrap_doc(raw)
            if not _d0.get("layout_sections"):
                print(f"[LAYOUT] {file_path.name}: response missing layout_sections "
                      f"— retrying with stronger format enforcement", flush=True)
                _enforce = (
                    "YOU MUST respond using the layout_sections format specified. "
                    "Do NOT use extracted_fields format. The layout_sections key is "
                    "REQUIRED in your response.\n\n"
                )
                try:
                    if page_img:
                        _re_ext = orchestrator.llm.extract(
                            image_b64=page_img, prompt=_enforce + _base_prompt,
                            system_instruction=system_instruction)
                    else:
                        _re_ext = orchestrator.llm.extract(
                            text=doc_text, prompt=_enforce + _base_prompt,
                            system_instruction=system_instruction)
                    if (_re_ext and _re_ext.success and _re_ext.parsed_json
                            and _unwrap_doc(_re_ext.parsed_json).get("layout_sections")):
                        raw = _re_ext.parsed_json
                        extraction = _re_ext
                        _d0 = _unwrap_doc(raw)
                        print(f"[LAYOUT] {file_path.name}: enforcement retry produced layout_sections", flush=True)
                except Exception as _le:
                    print(f"[LAYOUT] enforcement retry error: {_le}", flush=True)

            if not _d0.get("layout_sections") and _d0.get("extracted_fields"):
                _conv = _convert_extracted_fields_to_layout(
                    _d0.get("extracted_fields", {}), template_data.get("binding_map", {})
                )
                if _conv:
                    _d0["layout_sections"] = _conv     # mutates the doc inside raw
                    print(f"[LAYOUT] {file_path.name}: converted extracted_fields -> "
                          f"layout_sections ({sum(len(s.get('rows', [])) for s in _conv.values())} rows)",
                          flush=True)

        # Handle documents array in response
        documents = raw.get("documents", [raw])
        if not documents:
            documents = [raw]

        for doc_result_raw in documents:
            doc_idx = doc_result_raw.get("doc_index", seg_index)
            doc_hint = doc_result_raw.get("doc_hint", seg_hint)

            seg_fn = (file_path.name if total_docs == 1 and len(documents) == 1
                      else f"{file_path.stem}_doc{doc_idx+1}{file_path.suffix}")

            result = _process_vision_result(
                doc_result_raw, template_data, seg_fn, doc_type,
                elapsed, extraction, doc_text, doc_hint, doc_idx
            )
            results.append(result)

    return results if results else [_fail(file_path.name, "No documents extracted")]


def _legacy_columns_extract(orchestrator, file_path, template_data,
                             doc_type, doc_text, page_images, start):
    """Legacy flat-column extraction for old-format templates."""
    import time as t
    results = []
    header_cols = template_data.get("header_cols", [])
    system_prompt = _get_system_prompt(doc_type)

    def col_hint(col):
        return {"Number": "number only", "Currency": "number only",
                "Date": "YYYY-MM-DD", "Text": "text"}.get(col.get("type","Text"), "text")

    header_lines = "\n".join(
        f'  - "{c["name"]}": {col_hint(c)}'
        for c in sorted(header_cols, key=lambda x: x.get("order",0))
        if c.get("name","").strip()
    )
    prompt = f"""{system_prompt}

Extract these fields from the document:
{header_lines}

Return ONLY JSON:
{{"document_type": "{doc_type}", "overall_confidence": "high|medium|low", "header": {{}}}}"""

    try:
        if page_images:
            extraction = orchestrator.llm.extract(image_b64=page_images[0], prompt=prompt)
        else:
            extraction = orchestrator.llm.extract(text=doc_text, prompt=prompt)
    except Exception as e:
        results.append(_fail(file_path.name, str(e)))
        return results

    elapsed = (t.time() - start) * 1000

    if not extraction.success or not extraction.parsed_json:
        r = _fail(file_path.name, extraction.error)
        r.processing_time_ms = elapsed
        results.append(r)
        return results

    from orchestrator import DocumentExtractionResult
    raw = extraction.parsed_json
    header_data = raw.get("header", {})
    normalised = {}
    for col in header_cols:
        name = col.get("name","").strip()
        if not name: continue
        fd = header_data.get(name)
        if fd is None:
            normalised[name] = {"value":"","confidence":"high"}
        elif isinstance(fd, dict):
            normalised[name] = {"value": fd.get("value",""), "confidence": fd.get("confidence","high")}
        else:
            normalised[name] = {"value": str(fd) if fd is not None else "", "confidence":"high"}

    r = DocumentExtractionResult(filename=file_path.name)
    r.document_type = doc_type
    r.extracted_data = {
        "document_type": doc_type,
        "overall_confidence": raw.get("overall_confidence","medium"),
        "extraction_method": "legacy_columns",
        "extracted_data": normalised,
    }
    r.extraction_response = extraction
    r.processing_time_ms = elapsed
    r.success = True
    results.append(r)
    return results


# ==============================================================================
# BACKGROUND THREAD
# ==============================================================================

def _post_categorize(extracted: dict, orchestrator, doc_type: str) -> dict:
    """
    Second-pass categorization when the extraction didn't produce Category values.
    Sends only the table rows to Gemini for categorization — minimal tokens.
    """
    try:
        table_rows = extracted.get("table_rows", [])
        if not table_rows:
            return extracted

        rows_text = "\n".join(
            f"Row {i+1}: {json.dumps(row)}"
            for i, row in enumerate(table_rows[:50])  # max 50 rows
        )
        prompt = (
            f"Assign a business category to each row below.\n"
            f"Use these categories: Travel, Meals, Lodging, Office Supplies, "
            f"Software, Marketing, Professional Fees, Utilities, Salary, "
            f"Rent, Insurance, Equipment, Tax, Bank Charges, Other.\n\n"
            f"Return ONLY a JSON object: {{\"categories\": [\"cat1\", \"cat2\", ...]}}\n"
            f"One category per row in the same order.\n\n"
            f"Rows:\n{rows_text}"
        )
        result = orchestrator.llm.extract(text=rows_text, prompt=prompt)
        if result.success and result.parsed_json:
            cats = result.parsed_json.get("categories", [])
            for i, row in enumerate(table_rows):
                if i < len(cats) and not row.get("Category"):
                    row["Category"] = cats[i]
            extracted["table_rows"] = table_rows
            print(f"[OPTIONS] categorized {len(cats)} rows", flush=True)
    except Exception as e:
        print(f"[OPTIONS] categorize failed: {e}", flush=True)
    return extracted


def _post_summarize(extracted: dict, orchestrator, doc_type: str) -> dict:
    """
    Generate a 2-3 sentence plain English summary of the extracted document.
    Uses Gemini — stored in extracted['summary'].
    """
    try:
        inner = extracted.get("extracted_data", {})
        fields = {k: (v.get("value", "") if isinstance(v, dict) else v)
                  for k, v in inner.items() if not k.startswith("_label_")}
        if not fields:
            return extracted

        field_str = "\n".join(f"{k}: {v}" for k, v in list(fields.items())[:20])
        table_rows = extracted.get("table_rows", [])
        table_str = f"\n{len(table_rows)} line items." if table_rows else ""

        prompt = (
            f"Summarise this {doc_type} document in 2-3 plain English sentences. "
            f"Be specific about key values. No bullet points.\n\n"
            f"Fields:\n{field_str}{table_str}\n\n"
            f"Return ONLY a JSON object: {{\"summary\": \"your summary here\"}}"
        )
        result = orchestrator.llm.extract(text=field_str, prompt=prompt)
        if result.success and result.parsed_json:
            summary = result.parsed_json.get("summary", "")
            if summary:
                extracted["summary"] = summary
                print(f"[OPTIONS] summary generated: {summary[:80]}...", flush=True)
    except Exception as e:
        print(f"[OPTIONS] summary failed: {e}", flush=True)
    return extracted


def _post_anomaly(extracted: dict, orchestrator, doc_type: str) -> dict:
    """
    Detect anomalies in extracted data — duplicate values, unusual amounts,
    missing required fields, outliers.
    Stored in extracted['anomalies'] as a list of strings.
    """
    try:
        inner = extracted.get("extracted_data", {})
        fields = {k: (v.get("value", "") if isinstance(v, dict) else v)
                  for k, v in inner.items() if not k.startswith("_label_")}
        table_rows = extracted.get("table_rows", [])[:30]

        field_str = "\n".join(f"{k}: {v}" for k, v in list(fields.items())[:15])
        row_str   = "\n".join(f"Row {i+1}: {json.dumps(r)}"
                               for i, r in enumerate(table_rows))

        prompt = (
            f"Analyse this extracted {doc_type} data for anomalies: "
            f"unusual values, duplicates, outliers, or missing required fields.\n"
            f"Return ONLY a JSON object: {{\"anomalies\": [\"issue1\", \"issue2\"]}}\n"
            f"Maximum 5 items. If nothing unusual return {{\"anomalies\": []}}\n\n"
            f"Fields:\n{field_str}\n\nRows:\n{row_str}"
        )
        result = orchestrator.llm.extract(text=field_str, prompt=prompt)
        if result.success and result.parsed_json:
            anomalies = result.parsed_json.get("anomalies", [])
            extracted["anomalies"] = anomalies if isinstance(anomalies, list) else []
            print(f"[OPTIONS] anomalies: {len(extracted['anomalies'])} found", flush=True)
    except Exception as e:
        print(f"[OPTIONS] anomaly failed: {e}", flush=True)
    return extracted


_FINANCIAL_CRITICAL_TYPES = frozenset({
    "balance_sheet", "income_statement", "payslip", "sales_invoice",
})


def _cross_validate_section_totals(extracted_data: dict, doc_type: str) -> Optional[str]:
    """
    H2: For balance sheets and income statements, cross-check whether the sum of
    line-item amounts roughly matches any declared section total.
    Returns a human-readable warning string if a mismatch > 1% is found, else None.
    """
    if doc_type not in ("balance_sheet", "income_statement"):
        return None

    table_rows = extracted_data.get("table_rows", [])
    if len(table_rows) < 2:
        return None

    inner = extracted_data.get("extracted_data", {})

    # Collect declared totals from extracted_data (form fields)
    declared_totals = {}
    for label, field in inner.items():
        label_lower = label.lower().strip()
        if label_lower.startswith("_label_"):
            continue
        if any(t in label_lower for t in ("total", "grand total", "net total")):
            raw_val = field.get("value", "") if isinstance(field, dict) else str(field)
            try:
                declared_totals[label] = abs(float(
                    re.sub(r'[,$£€₹()\s]', '', str(raw_val)).replace(")", "")
                ))
            except (ValueError, TypeError):
                pass

    if not declared_totals:
        return None

    # Sum numeric columns across table rows
    numeric_cols = set()
    for row in table_rows[:5]:
        for k, v in row.items():
            if k.startswith("_"):
                continue
            try:
                float(re.sub(r'[,$£€₹()\s]', '', str(v)))
                numeric_cols.add(k)
            except (ValueError, TypeError):
                pass

    for col in numeric_cols:
        row_sum = 0.0
        for row in table_rows:
            try:
                row_sum += abs(float(
                    re.sub(r'[,$£€₹()\s]', '', str(row.get(col, "") or ""))
                ))
            except (ValueError, TypeError):
                pass

        for label, declared in declared_totals.items():
            if declared > 0 and abs(row_sum - declared) / declared > 0.01:
                return (
                    f"H2 section-total mismatch: sum of '{col}' column = {row_sum:.2f}, "
                    f"declared '{label}' = {declared:.2f} "
                    f"(diff {abs(row_sum - declared) / declared * 100:.1f}%)"
                )
    return None


def _diagnose_template(template_data: dict) -> None:
    """H4: Log a diagnostic summary of the parsed template so we can verify detection."""
    if not template_data:
        print("[DIAG] template_data is None — no template in use", flush=True)
        return
    regions = template_data.get("regions", {})
    mode = regions.get("primary_mode", "?")
    targets = len(regions.get("explicit_targets", []))
    kv = len(regions.get("kv_pairs", []))
    two_col = len(regions.get("two_col_pairs", []))
    tables = regions.get("table_regions", [])
    transposed = len(regions.get("transposed_tables", []))
    sec_ctx = regions.get("needs_section_context", False)
    has_plain = bool(regions.get("plain_text_description"))
    print(
        f"[DIAG] template='{template_data.get('name','?')}' "
        f"doc_type={template_data.get('doc_type','?')} "
        f"mode={mode} targets={targets} kv={kv} two_col={two_col} "
        f"tables={len(tables)} transposed={transposed} "
        f"section_ctx={sec_ctx} plain_text={has_plain}",
        flush=True,
    )
    for i, tr in enumerate(tables):
        print(
            f"[DIAG]   table[{i}]: '{tr.get('section_label','?')}' "
            f"cols={tr.get('column_names',[])} "
            f"header_row={tr.get('header_row','?')}",
            flush=True,
        )


def _run_extraction_sync(job_id, file_paths, schema_path, db_url, template_data,
                          project_dir, backend_dir, engine_dir, options=None,
                          selected_pages=None):
    import os
    os.environ["PYTHONUTF8"] = "1"
    for p in [engine_dir, backend_dir, project_dir]:
        if p not in sys.path:
            sys.path.insert(0, p)

    try:
        from sqlalchemy import create_engine as sa_engine
        from sqlalchemy.orm import sessionmaker
        from app.models.models import ExtractionJob, DocumentResult
    except Exception as e:
        print(f"[THREAD] DB import failed: {e}", flush=True); return

    try:
        connect_args = {"check_same_thread": False} if "sqlite" in db_url else {}
        _eng = sa_engine(db_url, connect_args=connect_args)
        Session = sessionmaker(bind=_eng)
        session = Session()
    except Exception as e:
        print(f"[THREAD] DB session failed: {e}", flush=True); return

    try:
        job = session.query(ExtractionJob).filter_by(id=job_id).first()
        if not job: return
        job.status = "processing"
        session.commit()

        from orchestrator import Orchestrator
        orchestrator = Orchestrator(client_schema_path=schema_path)

        successful = failed = needs_review = 0
        total_tokens_used = 0
        total_cost_usd = 0.0
        start_time = time.time()

        for i, fp in enumerate(file_paths):
            print(f"[THREAD] processing file {i+1}/{len(file_paths)}: {fp}", flush=True)

            # Issue 2: update job progress so the UI can poll live status
            try:
                job.progress_message = (
                    f"Processing document {i+1} of {len(file_paths)}..."
                )
                session.commit()
            except Exception:
                pass

            try:
                file_path = Path(fp)
                is_img = _is_image_file(file_path)  # Issue 3: route image files
                if template_data:
                    if is_img:
                        results = _extract_image_with_template(
                            orchestrator, file_path, template_data
                        )
                    else:
                        results = _extract_with_template(
                            orchestrator, file_path, template_data,
                            selected_pages=selected_pages,
                        )
                else:
                    if is_img:
                        results = _extract_image_with_template(
                            orchestrator, file_path, None
                        )
                    else:
                        result = orchestrator._process_single_document(file_path)
                        results = [result]

                for result in results:
                    try:
                        validation_data = (result.extracted_data or {}).get("validation", {})
                        has_flags      = validation_data.get("flagged_count", 0) > 0
                        error_msg      = result.error if hasattr(result, 'error') and result.error else ""

                        # H1: for financial doc types, any low-confidence numeric field
                        #     escalates the entire document to needs_review=True
                        doc_type_str = (result.document_type or "").lower()
                        if not has_flags and doc_type_str in _FINANCIAL_CRITICAL_TYPES:
                            conf_map = validation_data.get("confidence_map", {})
                            if "low" in conf_map.values():
                                has_flags = True
                                print(
                                    f"[H1] {result.filename}: financial doc "
                                    f"({doc_type_str}) has low-confidence field "
                                    f"— needs_review escalated",
                                    flush=True,
                                )

                        # Bug 6: misaligned rows also escalate to needs_review
                        if not has_flags and validation_data.get("alignment_misaligned"):
                            has_flags = True

                        # H2: section-total cross-validation for balance sheets / income statements
                        if result.success and doc_type_str in ("balance_sheet", "income_statement"):
                            section_warn = _cross_validate_section_totals(
                                result.extracted_data or {}, doc_type_str
                            )
                            if section_warn:
                                print(f"[H2] {result.filename}: {section_warn}", flush=True)
                                has_flags = True
                                error_msg = (error_msg + "; " if error_msg else "") + section_warn

                        # ── Post-extraction processing based on selected options ──
                        extracted = result.extracted_data or {}

                        if result.success and options:
                            # --- CATEGORIZATION ---
                            # Category is already embedded by the AI in table_rows
                            # via the extraction prompt (bank_statement taxonomy,
                            # expense_report Category column etc.)
                            # If categorize option is on and Category is missing from
                            # table rows, inject it via a second Gemini call
                            if "categorize" in options:
                                table_rows = extracted.get("table_rows", [])
                                missing_cat = table_rows and not table_rows[0].get("Category")
                                if missing_cat:
                                    extracted = _post_categorize(
                                        extracted, orchestrator, result.document_type or ""
                                    )

                            # --- AI SUMMARY ---
                            # Gemini generates a 2-3 sentence plain English summary
                            # stored in extraction_json.summary
                            if "summary" in options and not extracted.get("summary"):
                                extracted = _post_summarize(
                                    extracted, orchestrator, result.document_type or ""
                                )

                            # --- ANOMALY DETECTION ---
                            # Gemini flags unusual values, stored in extraction_json.anomalies
                            if "anomaly" in options and not extracted.get("anomalies"):
                                extracted = _post_anomaly(
                                    extracted, orchestrator, result.document_type or ""
                                )

                            result.extracted_data = extracted

                        # FIX 5: capture the raw LLM response (empty for pdfplumber-only docs)
                        _raw_llm = ""
                        if result.extraction_response is not None:
                            _raw_llm = getattr(result.extraction_response, "raw_text", "") or ""

                        doc = DocumentResult(
                            job_id=job_id,
                            filename=result.filename,
                            document_type=result.document_type if result.success else "unknown",
                            overall_confidence=extracted.get("overall_confidence"),
                            extraction_json=json.dumps(extracted, default=str) if extracted else None,
                            raw_llm_response=_raw_llm or None,
                            validation_errors=error_msg or ("; ".join(result.validation.errors) if result.validation else ""),
                            validation_warnings=(
                                "; ".join(f"{f['ref']}: {f['value']}" for f in validation_data.get("flagged_fields", []))
                                if has_flags else ""
                            ),
                            needs_review=has_flags or (result.validation.needs_review if result.validation else False),
                            model_used=(result.extraction_response.model_used
                                        if result.extraction_response else "vision_direct"),
                            tokens_used=(result.extraction_response.tokens_used
                                         if result.extraction_response else 0),
                            latency_ms=result.processing_time_ms,
                        )
                        session.add(doc)
                        session.flush()

                        doc_tokens = (result.extraction_response.tokens_used
                                      if result.extraction_response else 0) or 0
                        total_tokens_used += doc_tokens
                        doc_cost = (doc_tokens / 1_000_000) * 0.15
                        total_cost_usd += doc_cost

                        print(f"[THREAD] saved: {result.filename} "
                              f"tokens={doc_tokens} cost=${doc_cost:.5f} "
                              f"| job_total={total_tokens_used} tokens "
                              f"${total_cost_usd:.4f}", flush=True)

                        if result.success:
                            successful += 1
                            if doc.needs_review:
                                needs_review += 1
                        else:
                            failed += 1

                    except Exception as save_err:
                        print(f"[THREAD] SAVE ERROR for {result.filename}: {save_err}", flush=True)
                        traceback.print_exc()
                        failed += 1

            except Exception as doc_err:
                print(f"[THREAD] doc error: {doc_err}", flush=True)
                traceback.print_exc()
                failed += 1

            # Rate limit protection: delay between documents
            # Prevents Groq/Gemini 429 errors on batch uploads
            # Default 3s — enough for Groq free tier (30 req/min = 2s min gap)
            if i < len(file_paths) - 1:
                delay = float(getattr(settings, 'RATE_LIMIT_DELAY', 3.0))
                print(f"[THREAD] rate limit delay {delay}s before next doc", flush=True)
                time.sleep(delay)

        session.commit()
        job.status = "completed"
        job.successful = successful
        job.failed = failed
        job.needs_review = needs_review
        job.total_time_sec = time.time() - start_time
        job.completed_at = datetime.utcnow()
        session.commit()
        print(f"[THREAD] COMPLETE: {successful} ok, {failed} failed, "
              f"{needs_review} review | "
              f"TOKENS: {total_tokens_used:,} total | "
              f"COST: ${total_cost_usd:.4f} | "
              f"TIME: {job.total_time_sec:.1f}s", flush=True)

    except Exception as e:
        print(f"[THREAD] FAILED: {e}", flush=True)
        traceback.print_exc()
        try:
            job = session.query(ExtractionJob).filter_by(id=job_id).first()
            if job:
                job.status = "failed"
                job.error_message = str(e)
                job.completed_at = datetime.utcnow()
                session.commit()
        except Exception:
            pass
    finally:
        session.close()


# ==============================================================================
# EXCEL EXPORT
# ==============================================================================

@router.get("/jobs/{job_id}/export")
def export_job_excel(

    job_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    job = _get_job_or_404(job_id, current_user, db)
    doc_results = (
        db.query(DocumentResult)
        .filter(DocumentResult.job_id == job_id)
        .order_by(DocumentResult.id)
        .all()
    )
    if not doc_results:
        raise HTTPException(status_code=404, detail="No results found.")

    # FIX 4: single source of truth — prefer the region analysis saved at
    # extraction time so export cannot diverge from extraction.
    saved_regions = None
    for _d in doc_results:
        _ed = _d.get_extracted_data()
        if isinstance(_ed, dict) and _ed.get("template_regions"):
            saved_regions = _ed["template_regions"]
            break

    sheet_data = None
    template_regions = None
    if job.schema_id:
        try:
            tpl = db.query(ColumnTemplate).filter(
                ColumnTemplate.id == int(job.schema_id)
            ).first()
            if tpl and tpl.description:
                raw = json.loads(tpl.description)
                if isinstance(raw, dict) and "cells" in raw:
                    sheet_data = raw
                    template_regions = saved_regions or _analyse_template_regions(raw)
        except Exception as e:
            print(f"[EXPORT] Template load error: {e}", flush=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    if sheet_data:
        _write_excel(ws, doc_results, sheet_data, template_regions, openpyxl)
    else:
        # No template saved — write a basic structured export
        # This should rarely happen as jobs are always run with templates
        _write_flat_table(ws, doc_results, openpyxl)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(buf, headers={
        "Content-Disposition": f'attachment; filename="job_{job_id}_results.xlsx"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    })


@router.get("/jobs/{job_id}/export/zip")
def export_job_zip(
    job_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export each document as a separate Excel file, bundled into a ZIP."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    import zipfile

    job = _get_job_or_404(job_id, current_user, db)
    doc_results = (
        db.query(DocumentResult)
        .filter(DocumentResult.job_id == job_id)
        .order_by(DocumentResult.id)
        .all()
    )
    if not doc_results:
        raise HTTPException(status_code=404, detail="No results found.")

    # FIX 4: single source of truth — prefer the region analysis saved at
    # extraction time so export cannot diverge from extraction.
    saved_regions = None
    for _d in doc_results:
        _ed = _d.get_extracted_data()
        if isinstance(_ed, dict) and _ed.get("template_regions"):
            saved_regions = _ed["template_regions"]
            break

    sheet_data = None
    template_regions = None
    if job.schema_id:
        try:
            tpl = db.query(ColumnTemplate).filter(
                ColumnTemplate.id == int(job.schema_id)
            ).first()
            if tpl and tpl.description:
                raw = json.loads(tpl.description)
                if isinstance(raw, dict) and "cells" in raw:
                    sheet_data = raw
                    template_regions = saved_regions or _analyse_template_regions(raw)
        except Exception as e:
            print(f"[EXPORT] Template load error: {e}", flush=True)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for doc in doc_results:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"

            if sheet_data:
                _write_excel(ws, [doc], sheet_data, template_regions, openpyxl)
            else:
                _write_flat_table(ws, [doc], openpyxl)

            # Use the original filename without extension as the Excel filename
            base = doc.filename.rsplit(".", 1)[0] if "." in doc.filename else doc.filename
            excel_name = f"{base}_extracted.xlsx"

            excel_buf = io.BytesIO()
            wb.save(excel_buf)
            zf.writestr(excel_name, excel_buf.getvalue())

    zip_buf.seek(0)
    safe_name = f"job_{job_id}_documents.zip"
    return StreamingResponse(zip_buf, headers={
        "Content-Disposition": f'attachment; filename="{safe_name}"',
        "Content-Type": "application/zip",
    })


def _write_excel(ws, doc_results, sheet_data, template_regions, openpyxl_mod):
    """Route to the correct writer based on template regions primary_mode.
    Always uses the template structure for routing — never relies on AI output
    flags which can be stale or wrong from a previous extraction run."""
    from openpyxl.utils import get_column_letter
    cells_tpl = sheet_data.get("cells", {})
    col_widths = sheet_data.get("colWidths", [])
    max_r, max_c = _find_template_dimensions(cells_tpl)

    for c_idx, width_px in enumerate(col_widths):
        if width_px and c_idx <= max_c:
            ws.column_dimensions[get_column_letter(c_idx + 1)].width = max(8, round(width_px / 7))

    # COMPONENT 5: layout-mode output (extract & place) takes priority when present
    def _has_layout(d):
        ed = d.get_extracted_data()
        ls = ed.get("layout_sections") if isinstance(ed, dict) else None
        return isinstance(ls, dict) and any(
            isinstance(s, dict) and s.get("rows") for s in ls.values()
        )
    if any(_has_layout(d) for d in doc_results):
        print("[EXPORT] routing: layout-mode writer", flush=True)
        _write_layout_excel(ws, doc_results, sheet_data, cells_tpl, openpyxl_mod)
        return

    # Route based on template structure — authoritative, never stale
    primary_mode = (template_regions or {}).get("primary_mode", "form_kv")

    print(f"[EXPORT] routing: primary_mode={primary_mode}", flush=True)

    if primary_mode == "table":
        _write_table_excel(ws, doc_results, sheet_data, cells_tpl, template_regions, openpyxl_mod)
    elif primary_mode == "mixed":
        _write_mixed_excel(ws, doc_results, sheet_data, cells_tpl, max_r, max_c, template_regions, openpyxl_mod)
    else:
        # form_with_targets, form_kv, two_column — all use form writer
        _write_form_excel(ws, doc_results, sheet_data, cells_tpl, max_r, max_c, openpyxl_mod)


def _write_layout_excel(ws, doc_results, sheet_data, cells_tpl, openpyxl_mod):
    """
    COMPONENT 5 — layout-aware writer. Writes the static template cells (headers,
    totals labels), then places each layout_sections row's label/value into the
    designated columns at the AI-specified spreadsheet row, then writes fixed
    extracted_fields (e.g. totals) by cell reference.
    """
    from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

    # 1. Static template cells (column headers, section/total labels, etc.)
    for key, cell_def in cells_tpl.items():
        if not isinstance(cell_def, dict):
            continue
        parts = key.split(",")
        if len(parts) != 2:
            continue
        try:
            tr, tc = int(parts[0]), int(parts[1])
        except ValueError:
            continue
        v = str(cell_def.get("value") or "").strip()
        xl = ws.cell(row=tr + 1, column=tc + 1)
        if v:
            xl.value = v
        if cell_def.get("style"):
            _apply_cell_style(xl, cell_def["style"], openpyxl_mod)

    def _set_num(cell, raw):
        s = str(raw).replace(",", "").replace("$", "").strip()
        try:
            if s and re.match(r'^-?[0-9]+\.?[0-9]*$', s):
                cell.value = float(s) if "." in s else int(s)
                return
        except (ValueError, TypeError):
            pass
        cell.value = str(raw)

    # Template bounding-box width (for the column safety net)
    _max_tc = max((int(k.split(",")[1]) for k in cells_tpl
                   if "," in k and k.split(",")[1].isdigit()), default=0)

    def _in_bounds(letter):
        try:
            return bool(letter) and column_index_from_string(letter.upper()) <= (_max_tc + 1)
        except Exception:
            return False

    # 2. Dynamic content + 3. fixed cells, per document
    for doc in doc_results:
        ed = doc.get_extracted_data()
        ls = ed.get("layout_sections", {}) if isinstance(ed, dict) else {}
        groups = ed.get("binding_column_groups", []) if isinstance(ed, dict) else []

        def _nrm(s):
            return re.sub(r'[^a-z0-9]+', '', str(s or "").lower())

        def _fix_cols(rnum, lc, vc, sec_key):
            """SAFETY NET (FIX 2/3): if a column is empty/None or outside the
            template bounding box, remap to the column group covering this row.
            Disambiguate side-by-side groups by fuzzy-matching the section key;
            if no group covers the row, fall back to the last group."""
            need_lc, need_vc = not _in_bounds(lc), not _in_bounds(vc)
            if not groups or (not need_lc and not need_vc):
                return lc, vc
            r0 = rnum - 1
            covering = [g for g in groups
                        if int(g.get("start_row", 0)) <= r0 <= int(g.get("end_row", -1))]
            chosen = None
            if covering:
                # FIX 3: prefer the group whose section label matches the section key
                chosen = next((g for g in covering
                               if _nrm(g.get("section_label")) == _nrm(sec_key)), None)
                if chosen is None:           # else the group already matching an AI column
                    chosen = next((g for g in covering
                                   if (g.get("label_col_letter") or "").upper() == lc
                                   or (g.get("value_col_letter") or "").upper() == vc), covering[0])
            elif groups:
                chosen = groups[-1]          # no group covers this row -> last group
            if not chosen:
                return lc, vc
            # FIX 2: once we remap a row, correct BOTH columns to the chosen group's
            # columns — not just the label. (Previously new_vc was only corrected when
            # vc was out of bounds, so an in-bounds-but-wrong value column like "E"
            # was left as "A/E" instead of "A/B".)
            new_lc = (chosen.get("label_col_letter") or lc).upper()
            new_vc = (chosen.get("value_col_letter") or vc).upper()
            if (new_lc, new_vc) != (lc, vc):
                print(f"[LAYOUT-FIX] row {rnum}: label {lc or '<none>'} -> {new_lc}, "
                      f"value {vc or '<none>'} -> {new_vc} via binding_column_groups "
                      f"(group '{chosen.get('section_label', '')}')", flush=True)
            return new_lc, new_vc

        for sec_key, sec in (ls.items() if isinstance(ls, dict) else []):
            if not isinstance(sec, dict):
                continue
            for row in sec.get("rows", []) or []:
                if not isinstance(row, dict):
                    continue
                try:
                    rnum = int(row.get("row"))
                except (TypeError, ValueError):
                    continue
                lc = str(row.get("label_col") or "").strip().upper()
                vc = str(row.get("value_col") or "").strip().upper()
                lc, vc = _fix_cols(rnum, lc, vc, sec_key)
                lbl, vval = row.get("label"), row.get("value")
                if lc and lbl not in (None, ""):
                    try:
                        ws.cell(row=rnum, column=column_index_from_string(lc)).value = str(lbl)
                    except Exception:
                        pass
                if vc and vval not in (None, ""):
                    try:
                        _set_num(ws.cell(row=rnum, column=column_index_from_string(vc)), vval)
                    except Exception:
                        pass
        for ref, v in (ed.get("extracted_fields", {}) or {}).items():
            value = v.get("value", "") if isinstance(v, dict) else v
            if value in (None, ""):
                continue
            try:
                col_letter, rownum = coordinate_from_string(str(ref))
                _set_num(ws.cell(row=rownum, column=column_index_from_string(col_letter)), value)
            except Exception:
                continue

    print(f"[EXPORT] layout: wrote {len(doc_results)} document block(s)", flush=True)


def _write_table_excel(ws, doc_results, sheet_data, cells_tpl, template_regions, openpyxl_mod):
    """
    Table mode: write header row with original styles, then one data row per line item.
    Handles tables at any position in the grid (not just row 0).
    """
    table_regions = (template_regions or {}).get("table_regions", [])
    header_row_idx = table_regions[0]["header_row"] if table_regions else 0
    start_col_idx = table_regions[0]["start_col"] if table_regions else 0

    # Write ALL cells from the template (including any form fields above the table)
    for key, cell_def in cells_tpl.items():
        if not isinstance(cell_def, dict): continue
        parts = key.split(",")
        if len(parts) != 2: continue
        tr, tc = int(parts[0]), int(parts[1])
        xl_cell = ws.cell(row=tr + 1, column=tc + 1)
        xl_cell.value = cell_def.get("value", "").strip()
        if cell_def.get("style"):
            _apply_cell_style(xl_cell, cell_def["style"], openpyxl_mod)

    # Write data rows starting from the row after the header
    col_names = []
    for tr in table_regions:
        col_names.extend(tr.get("column_names", []))
    col_names = list(dict.fromkeys(col_names))
    col_indices = {name: (start_col_idx + i) for i, name in enumerate(col_names)}

    current_row = header_row_idx + 2  # +1 for 1-based, +1 to skip header

    for doc_result in doc_results:
        extracted = doc_result.get_extracted_data()
        table_rows = extracted.get("table_rows", [])

        for row_data in table_rows:
            for col_name, c_idx in col_indices.items():
                val = row_data.get(col_name, "")
                if isinstance(val, dict):
                    val = val.get("value", "")
                xl_cell = ws.cell(row=current_row, column=c_idx + 1)
                try:
                    xl_cell.value = (float(val) if "." in str(val) else int(val)) if val else val
                except (ValueError, TypeError):
                    xl_cell.value = val
            current_row += 1

    print(f"[EXPORT] table: header at row {header_row_idx+1}, "
          f"{current_row - header_row_idx - 2} data rows", flush=True)


def _write_form_excel(ws, doc_results, sheet_data, cells_tpl, max_r, max_c, openpyxl_mod):
    """
    Form mode: one filled template block per document result.
    Multi-doc: 3 blank rows + grey divider between documents.
    """
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    merges_tpl   = sheet_data.get("merges", {})
    # Extra rows per block for separator (only applied from block 2 onwards)
    SEPARATOR_HEIGHT = 5  # 3 blank + 1 divider + 1 blank
    block_height = max_r + 2

    def get_row_offset(block_idx):
        """Calculate row offset for a block, accounting for separators."""
        if block_idx == 0:
            return 0
        return block_idx * block_height + block_idx * SEPARATOR_HEIGHT

    for block_idx, doc_result in enumerate(doc_results):
        row_offset = get_row_offset(block_idx)
        extracted_data   = doc_result.get_extracted_data()
        extracted_fields = extracted_data.get("extracted_fields", {})
        validation       = extracted_data.get("validation", {})
        confidence_map   = validation.get("confidence_map", {})

        label_to_value = {
            k: (v.get("value","") if isinstance(v,dict) else str(v or ""))
            for k, v in extracted_data.get("extracted_data", {}).items()
            if not k.startswith("_label_")
        }

        # First pass: write all cells and collect numeric values by ref
        # so we can calculate formulas
        cell_values = {}  # ref -> numeric value for formula calculation

        for key, cell_def in cells_tpl.items():
            if not isinstance(cell_def, dict) or cell_def.get("mergeParent"):
                continue
            parts = key.split(",")
            if len(parts) != 2:
                continue
            tr, tc = int(parts[0]), int(parts[1])
            xl_cell = ws.cell(row=row_offset + tr + 1, column=tc + 1)
            tpl_value = cell_def.get("value","").strip()

            ref = f"{_col_letter(tc)}{tr+1}"
            if cell_def.get("extractTarget"):
                filled = extracted_fields.get(ref) or label_to_value.get(tpl_value, "")
                if isinstance(filled, dict):
                    filled = filled.get("value", "")
                filled = filled if filled is not None else ""

                # Try to store as number for formula calculation
                try:
                    num_val = float(filled.replace(",","")) if filled else None
                    if num_val is not None:
                        xl_cell.value = num_val
                        cell_values[ref] = num_val
                    else:
                        xl_cell.value = filled
                except (ValueError, AttributeError):
                    xl_cell.value = filled

                # Highlight low-confidence
                confidence = confidence_map.get(ref, "high")
                if confidence == "low":
                    try:
                        from openpyxl.styles import PatternFill
                        xl_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFF0AA")
                    except Exception:
                        pass

            elif not tpl_value and ref in extracted_fields:
                # kv_pair value cell: empty in template but AI extracted a value
                kv_filled = extracted_fields.get(ref)
                if isinstance(kv_filled, dict):
                    kv_filled = kv_filled.get("value", "")
                kv_filled = kv_filled or ""
                try:
                    num_val = float(str(kv_filled).replace(",", "")) if kv_filled else None
                    xl_cell.value = num_val if num_val is not None else kv_filled
                    if num_val is not None:
                        cell_values[ref] = num_val
                except (ValueError, AttributeError):
                    xl_cell.value = kv_filled

            elif tpl_value.startswith("="):
                # It's a formula — calculate it if possible
                calculated = _calculate_formula(tpl_value, cell_values, row_offset)
                if calculated is not None:
                    # Write the calculated value (Excel will recalc on open anyway)
                    xl_cell.value = calculated
                    # Also store for downstream formula references
                    ref = f"{_col_letter(tc)}{tr+1}"
                    cell_values[ref] = calculated
                else:
                    # Write the formula with adjusted row offset for this block
                    adjusted = _adjust_formula_for_block(tpl_value, row_offset)
                    xl_cell.value = adjusted
            else:
                xl_cell.value = tpl_value

            if cell_def.get("style"):
                _apply_cell_style(xl_cell, cell_def["style"], openpyxl_mod)

            merge_span = cell_def.get("mergeSpan") or merges_tpl.get(key)
            if merge_span:
                sr, sc = merge_span.get("rows",1), merge_span.get("cols",1)
                if sr > 1 or sc > 1:
                    try:
                        ws.merge_cells(
                            start_row=row_offset+tr+1, start_column=tc+1,
                            end_row=row_offset+tr+sr, end_column=tc+sc,
                        )
                    except Exception:
                        pass

        # Document separator between blocks — grey divider row
        if block_idx > 0:
            sep_base = row_offset - SEPARATOR_HEIGHT + 1  # 1-based
            # Grey shaded divider row
            for col_i in range(1, max_c + 3):
                ws.cell(row=sep_base + 3, column=col_i).fill = \
                    PatternFill(fill_type="solid", fgColor="FFF3F4F6")
            lc = ws.cell(row=sep_base + 3, column=1)
            lc.value = f"Document {block_idx + 1}  ·  {doc_result.filename}"
            lc.font = Font(bold=True, color="FF374151", size=10)

        # Document separator between blocks — placed AFTER writing this block
        # Use a separate counter so we don't corrupt row_offset for current block
        separator_row = row_offset + max_r + 2
        if block_idx > 0:
            # This separator was written at the start — skip (handled below)
            pass

        # ── Dynamic cells pass ────────────────────────────────────────────────────
        # Write any extracted_fields cells NOT already covered by cells_tpl entries.
        # This handles dynamic fill rows (e.g. balance sheet rows 1-8) where the
        # template has no cell definition but pdfplumber filled the values.
        written_refs = set()
        for key, cell_def in cells_tpl.items():
            if not isinstance(cell_def, dict) or cell_def.get("mergeParent"):
                continue
            parts = key.split(",")
            if len(parts) != 2:
                continue
            try:
                _tr2, _tc2 = int(parts[0]), int(parts[1])
            except ValueError:
                continue
            written_refs.add(f"{_col_letter(_tc2)}{_tr2 + 1}")

        _dyn_written = 0
        for ef_ref, ef_val in extracted_fields.items():
            if not ef_ref or ef_ref in written_refs:
                continue
            _col_str = "".join(ch for ch in ef_ref if ch.isalpha()).upper()
            _row_str = "".join(ch for ch in ef_ref if ch.isdigit())
            if not _col_str or not _row_str:
                continue
            try:
                _tr2 = int(_row_str) - 1
                _tc2 = sum(
                    (ord(ch) - 64) * (26 ** i)
                    for i, ch in enumerate(reversed(_col_str))
                ) - 1
                if _tr2 < 0 or _tc2 < 0:
                    continue
            except (ValueError, IndexError):
                continue
            _xl = ws.cell(row=row_offset + _tr2 + 1, column=_tc2 + 1)
            _val = ef_val.get("value", "") if isinstance(ef_val, dict) else str(ef_val or "")
            _val = _val if _val is not None else ""
            try:
                _clean = str(_val).replace(",", "").replace("$", "").strip() if _val else ""
                if _clean and re.match(r'^-?[0-9]+\.?[0-9]*$', _clean):
                    _xl.value = float(_clean)
                    cell_values[ef_ref] = float(_clean)
                else:
                    _xl.value = _val or ""
            except (ValueError, AttributeError):
                _xl.value = _val or ""
            _dyn_written += 1

        if _dyn_written:
            print(f"[EXPORT] form dynamic pass: {_dyn_written} extra cells written", flush=True)

        # Flag count indicator
        flag_count = validation.get("flagged_count", 0)
        if flag_count > 0:
            nc = ws.cell(row=row_offset + 1, column=max_c + 2)
            nc.value = f"! {flag_count} low-confidence fields"
            nc.font = Font(color="FFDC2626", size=9, italic=True)

    print(f"[EXPORT] form: {len(doc_results)} blocks written", flush=True)


def _calculate_formula(formula: str, cell_values: dict, row_offset: int) -> Optional[float]:
    """
    Calculate simple Excel formulas using known cell values.
    Handles: =SUM(B9:B11), =SUM(B3,B5,B7), basic arithmetic.
    Returns calculated value or None if cannot calculate.
    """
    import re
    f = formula.strip()

    # =SUM(range) e.g. =SUM(B9:B11)
    sum_range = re.match(r'^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', f, re.IGNORECASE)
    if sum_range:
        col1, row1, col2, row2 = sum_range.groups()
        total = 0.0
        found_any = False
        for r in range(int(row1), int(row2) + 1):
            ref = f"{col1.upper()}{r}"
            if ref in cell_values:
                total += cell_values[ref]
                found_any = True
        return round(total, 2) if found_any else None

    # =SUM(B3,B5,B7) comma-separated
    sum_list = re.match(r'^=SUM\(([^)]+)\)$', f, re.IGNORECASE)
    if sum_list:
        refs = [r.strip() for r in sum_list.group(1).split(",")]
        total = 0.0
        found_any = False
        for ref in refs:
            if ref.upper() in cell_values:
                total += cell_values[ref.upper()]
                found_any = True
        return round(total, 2) if found_any else None

    return None


def _adjust_formula_for_block(formula: str, row_offset: int) -> str:
    """
    Adjust cell references in a formula for a block offset.
    =SUM(B9:B11) with offset 20 becomes =SUM(B29:B31)
    """
    import re
    if row_offset == 0:
        return formula

    def adjust_ref(match):
        col = match.group(1)
        row = int(match.group(2))
        return f"{col}{row + row_offset}"

    return re.sub(r'([A-Z]+)(\d+)', adjust_ref, formula)


def _write_mixed_excel(ws, doc_results, sheet_data, cells_tpl, max_r, max_c,
                        template_regions, openpyxl_mod):
    """
    Mixed mode Excel writer.
    Handles: single table, multiple tables, tables anywhere in template.

    For each document block:
    1. Write form field rows (rows above the first table header)
    2. For each table region:
       a. Write the table header row from template
       b. Write all extracted data rows for that table
    3. Write any form rows that appear AFTER the last table (summary rows)
    4. Blank separator between document blocks
    """
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    merges_tpl   = sheet_data.get("merges", {})
    table_regions = (template_regions or {}).get("table_regions", [])

    # Get section_label_rows from region analyser (authoritative source)
    # These are rows identified during template analysis as structural labels
    region_section_label_rows = set(
        (template_regions or {}).get("section_label_rows", set())
    )

    # Sort tables by position
    tables_sorted = sorted(table_regions, key=lambda t: (t["header_row"], t["start_col"]))

    # Build set of table header rows
    table_header_rows = set(t["header_row"] for t in tables_sorted)

    # Section label rows — use region analyser result, fall back to heuristic
    if region_section_label_rows:
        section_label_rows = region_section_label_rows
    else:
        # Fallback: rows directly above table headers with text but no extract targets
        section_label_rows = set()
        for tbl in tables_sorted:
            hr = tbl["header_row"]
            for r_above in range(max(0, hr - 2), hr):
                has_label = any(
                    cells_tpl.get(f"{r_above},{c}", {}).get("value", "")
                    for c in range(max_c + 1)
                )
                has_extract = any(
                    cells_tpl.get(f"{r_above},{c}", {}).get("extractTarget", False)
                    for c in range(max_c + 1)
                )
                if has_label and not has_extract:
                    section_label_rows.add(r_above)

    # Blank rows in the template between sections — skip in output
    blank_between_tables = set()
    if tables_sorted:
        first_table_row = tables_sorted[0]["header_row"]
        last_table_row  = tables_sorted[-1]["header_row"]
        for r in range(first_table_row, last_table_row + 1):
            if r in table_header_rows or r in section_label_rows:
                continue
            row_has_content = any(
                cells_tpl.get(f"{r},{c}", {}).get("value", "") or
                cells_tpl.get(f"{r},{c}", {}).get("extractTarget", False)
                for c in range(max_c + 1)
            )
            if not row_has_content:
                blank_between_tables.add(r)

    # Determine which template rows are "form rows" (above first table)
    first_table_row = tables_sorted[0]["header_row"] if tables_sorted else max_r + 1
    last_table_row  = tables_sorted[-1]["header_row"] if tables_sorted else -1

    # Sort all template cells by row
    sorted_cells = sorted(
        [(key, cell_def) for key, cell_def in cells_tpl.items()
         if isinstance(cell_def, dict) and not cell_def.get("mergeParent")],
        key=lambda x: (int(x[0].split(",")[0]), int(x[0].split(",")[1]))
    )

    def write_template_row(tr, current_row, extracted_fields, label_to_value,
                           confidence_map):
        """Write one template row at the given output row."""
        from openpyxl.cell import MergedCell
        for key, cell_def in sorted_cells:
            parts = key.split(",")
            if int(parts[0]) != tr:
                continue
            tc = int(parts[1])
            tpl_value = cell_def.get("value", "").strip()
            xl_cell = ws.cell(row=current_row, column=tc + 1)

            # Skip cells that are already merged children — read-only in openpyxl
            if isinstance(xl_cell, MergedCell):
                continue

            ref = f"{_col_letter(tc)}{tr + 1}"
            if cell_def.get("extractTarget"):
                filled = extracted_fields.get(ref) or label_to_value.get(tpl_value, "")
                if isinstance(filled, dict):
                    filled = filled.get("value", "")
                try:
                    clean_filled = str(filled).replace(",", "").strip() if filled else ""
                    # Only convert to float for financial amount fields
                    # Never convert IDs, codes, reference numbers, routing numbers
                    amount_labels = {"amount","total","subtotal","balance","price",
                                    "cost","tax","freight","fee","charge","payment",
                                    "salary","gross","net","deduction","rate"}
                    label_lower = tpl_value.lower() if tpl_value else ""
                    is_amount = any(kw in label_lower for kw in amount_labels)
                    if filled and is_amount and re.match(r'^-?[0-9]+\.?[0-9]*$', clean_filled):
                        xl_cell.value = float(clean_filled)
                    else:
                        xl_cell.value = filled or ""
                except (ValueError, TypeError):
                    xl_cell.value = filled or ""
                conf = confidence_map.get(ref, "high")
                if conf == "low":
                    try:
                        xl_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFF0AA")
                    except Exception:
                        pass
            elif not tpl_value and ref in extracted_fields:
                # kv_pair value cell: empty in template but AI extracted a value
                kv_filled = extracted_fields.get(ref)
                if isinstance(kv_filled, dict):
                    kv_filled = kv_filled.get("value", "")
                kv_filled = kv_filled or ""
                try:
                    clean_kv = str(kv_filled).replace(",", "").strip() if kv_filled else ""
                    if clean_kv and re.match(r'^-?[0-9]+\.?[0-9]*$', clean_kv):
                        xl_cell.value = float(clean_kv)
                    else:
                        xl_cell.value = kv_filled
                except (ValueError, TypeError):
                    xl_cell.value = kv_filled
                conf = confidence_map.get(ref, "high")
                if conf == "low":
                    try:
                        xl_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFF0AA")
                    except Exception:
                        pass
            elif tpl_value.startswith("="):
                try:
                    xl_cell.value = tpl_value
                except AttributeError:
                    pass
            else:
                try:
                    xl_cell.value = tpl_value
                except AttributeError:
                    pass

            if cell_def.get("style"):
                _apply_cell_style(xl_cell, cell_def["style"], openpyxl_mod)

            merge_span = cell_def.get("mergeSpan") or merges_tpl.get(key)
            if merge_span:
                sr = merge_span.get("rows", 1)
                sc = merge_span.get("cols", 1)
                if sr > 1 or sc > 1:
                    try:
                        ws.merge_cells(
                            start_row=current_row, start_column=tc + 1,
                            end_row=current_row + sr - 1, end_column=tc + sc,
                        )
                    except Exception:
                        pass

    def write_table_data_rows(table_rows, tbl, current_row):
        """Write extracted data rows for one table."""
        col_names  = tbl.get("column_names", [])
        start_col  = tbl.get("start_col", 0)
        col_indices = {name: (start_col + i) for i, name in enumerate(col_names)}
        section    = tbl.get("section_label", "")

        # Build source key for this table (matches _table_source set during collection)
        import re as _re
        source_key = _re.sub(r'[^a-z0-9]', '_', section.lower()).strip('_')

        # Priority 1: rows tagged with _table_source matching this table
        rows_with_source = [
            r for r in table_rows
            if isinstance(r, dict) and r.get("_table_source", "") == source_key
        ]

        if rows_with_source:
            rows_for_table = rows_with_source
        elif len(tables_sorted) == 1:
            # Single table — use all rows
            rows_for_table = [r for r in table_rows if isinstance(r, dict)]
        else:
            # Fall back: rows that have this table's first column filled
            first_col = col_names[0] if col_names else None
            rows_for_table = [
                r for r in table_rows
                if isinstance(r, dict) and first_col and r.get(first_col, "")
                and not r.get("_table_source")
            ]

        written = 0
        for row_data in rows_for_table:
            # Skip rows that are headers or section labels
            row_vals = [str(v).strip() for v in row_data.values()
                       if v and not str(v).startswith("_")]
            if all(v in col_names or v == section for v in row_vals if v):
                continue

            for col_name, c_idx in col_indices.items():
                val = row_data.get(col_name, "")
                if isinstance(val, dict):
                    val = val.get("value", "")
                val = str(val).strip() if val is not None else ""
                if val in col_names:
                    val = ""
                xl_cell = ws.cell(row=current_row, column=c_idx + 1)
                try:
                    clean = val.replace(",", "").replace("$", "").replace("£", "").strip()
                    amount_labels = {"amount","total","subtotal","balance","price",
                                    "cost","tax","freight","fee","charge","payment",
                                    "salary","gross","net","deduction","rate","qty","quantity"}
                    col_lower = col_name.lower()
                    is_amount = any(kw in col_lower for kw in amount_labels)
                    if clean and clean not in ("", "-") and is_amount and re.match(r'^-?[0-9]+\.?[0-9]*$', clean):
                        xl_cell.value = float(clean)
                    else:
                        xl_cell.value = val or ""
                except (ValueError, TypeError):
                    xl_cell.value = val
            current_row += 1
            written += 1

        return current_row, written

    # ── Main write loop ───────────────────────────────────────────────────────
    current_output_row = 1

    for block_idx, doc_result in enumerate(doc_results):
        extracted_data   = doc_result.get_extracted_data()
        extracted_fields = extracted_data.get("extracted_fields", {})
        validation       = extracted_data.get("validation", {})
        confidence_map   = validation.get("confidence_map", {})
        table_rows       = extracted_data.get("table_rows", [])

        label_to_value = {
            k: (v.get("value", "") if isinstance(v, dict) else str(v or ""))
            for k, v in extracted_data.get("extracted_data", {}).items()
            if not k.startswith("_label_")
        }

        block_start_row = current_output_row

        # Document separator for multi-doc jobs
        # 3 blank rows + a clear divider line so document boundaries are obvious
        if block_idx > 0:
            # 3 blank rows to separate documents (vs 1 row between tables)
            current_output_row += 3

            # Divider row with filename — styled distinctly from table separators
            for col_i in range(1, max_c + 3):
                div_cell = ws.cell(row=current_output_row, column=col_i)
                div_cell.fill = PatternFill(fill_type="solid", fgColor="FFE5E7EB")
            lc = ws.cell(row=current_output_row, column=1)
            lc.value = f"Document: {doc_result.filename}"
            lc.font = Font(bold=True, color="FF1F2937", size=10)
            current_output_row += 1
            # 1 more blank row after the divider before content starts
            current_output_row += 1

        # Track which output row each template row maps to
        template_row_to_output = {}

        # Step 1: Write form rows ABOVE first table
        # Skip: section label rows (written in Step 2 before each table header)
        #        blank rows between sections
        for tr in range(min(first_table_row, max_r + 1)):
            if tr in section_label_rows or tr in blank_between_tables:
                # Skip — section labels are written in Step 2 before their table
                continue
            write_template_row(tr, current_output_row, extracted_fields,
                               label_to_value, confidence_map)
            template_row_to_output[tr] = current_output_row
            current_output_row += 1

        # Step 2: For each table — write section label, then header, then data rows
        for tbl_idx, tbl in enumerate(tables_sorted):
            hr = tbl["header_row"]

            # Fixed 1-row gap before each table (not the template's variable gap)
            # This ensures consistent spacing regardless of how many blank rows
            # the user left in the template between tables
            if tbl_idx > 0:
                current_output_row += 1  # exactly 1 blank row between tables

            # Write section label row(s) immediately before this table header
            # Only look 2 rows above — matching the detection window above
            for r_label in range(max(0, hr - 2), hr):
                if r_label in section_label_rows:
                    write_template_row(r_label, current_output_row, extracted_fields,
                                      label_to_value, confidence_map)
                    current_output_row += 1

            # Write table header row from template (column names row)
            write_template_row(hr, current_output_row, extracted_fields,
                               label_to_value, confidence_map)
            current_output_row += 1

            # Write data rows for this table
            current_output_row, n_rows = write_table_data_rows(
                table_rows, tbl, current_output_row
            )

            # If no rows extracted — write one blank placeholder row
            if n_rows == 0:
                current_output_row += 1

        # Step 3: Write form rows AFTER the last table (summary/totals)
        for tr in range(last_table_row + 1, max_r + 1):
            if tr in blank_between_tables:
                current_output_row += 1
                continue
            write_template_row(tr, current_output_row, extracted_fields,
                               label_to_value, confidence_map)
            template_row_to_output[tr] = current_output_row
            current_output_row += 1

        # ── Dynamic cells pass (mixed mode) ──────────────────────────────────────
        # Write extracted_fields cells that fell in blank_between_tables rows
        # and were therefore skipped by the main write loop.
        _written_m = set()
        for key, cell_def in cells_tpl.items():
            if not isinstance(cell_def, dict) or cell_def.get("mergeParent"):
                continue
            parts = key.split(",")
            if len(parts) != 2:
                continue
            try:
                _tr2, _tc2 = int(parts[0]), int(parts[1])
            except ValueError:
                continue
            _written_m.add(f"{_col_letter(_tc2)}{_tr2 + 1}")

        _dyn_written_m = 0
        for ef_ref, ef_val in extracted_fields.items():
            if not ef_ref or ef_ref in _written_m:
                continue
            _col_str = "".join(ch for ch in ef_ref if ch.isalpha()).upper()
            _row_str = "".join(ch for ch in ef_ref if ch.isdigit())
            if not _col_str or not _row_str:
                continue
            try:
                _tr2 = int(_row_str) - 1
                _tc2 = sum(
                    (ord(ch) - 64) * (26 ** i)
                    for i, ch in enumerate(reversed(_col_str))
                ) - 1
                if _tr2 < 0 or _tc2 < 0:
                    continue
            except (ValueError, IndexError):
                continue
            # Find nearest written template row above to anchor output position
            _preceding = [r for r in template_row_to_output if r < _tr2]
            if not _preceding:
                continue
            _anchor_tpl = max(_preceding)
            _anchor_out = template_row_to_output[_anchor_tpl]
            _out_row2 = _anchor_out + (_tr2 - _anchor_tpl)
            _xl = ws.cell(row=_out_row2, column=_tc2 + 1)
            _val = ef_val.get("value", "") if isinstance(ef_val, dict) else str(ef_val or "")
            _val = _val if _val is not None else ""
            try:
                _clean = str(_val).replace(",", "").replace("$", "").strip() if _val else ""
                if _clean and re.match(r'^-?[0-9]+\.?[0-9]*$', _clean):
                    _xl.value = float(_clean)
                else:
                    _xl.value = _val or ""
            except (ValueError, AttributeError):
                _xl.value = _val or ""
            _dyn_written_m += 1

        if _dyn_written_m:
            print(f"[EXPORT] mixed dynamic pass: {_dyn_written_m} extra cells written", flush=True)

        # Flag indicator
        flag_count = validation.get("flagged_count", 0)
        if flag_count > 0:
            nc = ws.cell(row=block_start_row, column=max_c + 2)
            nc.value = f"! {flag_count} low-confidence"
            nc.font = Font(color="FFDC2626", size=9, italic=True)

        # No trailing blank row here — spacing is added at the START of the next block
        # (3 blank rows for new document, 1 blank row between tables within same doc)

    print(f"[EXPORT] mixed: {len(doc_results)} blocks, "
          f"{current_output_row} total rows written", flush=True)


def _write_flat_table(ws, doc_results, openpyxl_mod):
    """Fallback flat table when no template."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    all_keys, seen = [], set()
    for dr in doc_results:
        for k in (dr.get_extracted_data().get("extracted_data") or {}):
            if k not in seen and not k.startswith("_label_"):
                seen.add(k); all_keys.append(k)
    hf = PatternFill(fill_type="solid", fgColor="FF4F46E5")
    hfont = Font(bold=True, color="FFFFFFFF", size=11)
    c = ws.cell(row=1, column=1, value="Filename"); c.font=hfont; c.fill=hf
    for ci, key in enumerate(all_keys, 2):
        c = ws.cell(row=1, column=ci, value=key); c.font=hfont; c.fill=hf
    for ri, dr in enumerate(doc_results, 2):
        ws.cell(row=ri, column=1, value=dr.filename)
        inner = dr.get_extracted_data().get("extracted_data") or {}
        for ci, key in enumerate(all_keys, 2):
            v = inner.get(key)
            ws.cell(row=ri, column=ci, value=(v.get("value","") if isinstance(v,dict) else (v or "")))
    ws.column_dimensions["A"].width = 30
    for ci in range(2, len(all_keys)+2):
        ws.column_dimensions[get_column_letter(ci)].width = 20


# ==============================================================================
# STYLE HELPERS
# ==============================================================================

def _parse_hex_color(hex_color):
    if not hex_color: return None
    h = hex_color.lstrip("#")
    if len(h)==3: h="".join(ch*2 for ch in h)
    return f"FF{h.upper()}" if len(h)==6 else None

def _apply_cell_style(xl_cell, style, _openpyxl_mod):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    fk = {}
    if style.get("bold"): fk["bold"]=True
    if style.get("italic"): fk["italic"]=True
    if style.get("underline"): fk["underline"]="single"
    if style.get("fontSize"): fk["size"]=style["fontSize"]
    if style.get("fontFamily"): fk["name"]=style["fontFamily"]
    fc = _parse_hex_color(style.get("fontColor"))
    if fc: fk["color"]=fc
    if fk:
        try: xl_cell.font = Font(**fk)
        except Exception: pass
    bg = _parse_hex_color(style.get("bgColor"))
    if bg:
        try: xl_cell.fill = PatternFill(fill_type="solid", fgColor=bg)
        except Exception: pass
    try:
        xl_cell.alignment = Alignment(
            horizontal={"left":"left","center":"center","right":"right"}.get(style.get("align",""),"left"),
            wrap_text=bool(style.get("wrap")), vertical="center"
        )
    except Exception: pass
    if style.get("borderAll"):
        t = Side(style="thin")
        try: xl_cell.border = Border(left=t,right=t,top=t,bottom=t)
        except Exception: pass

def _find_template_dimensions(cells):
    max_r, max_c = 0, 0
    for key in cells:
        parts = key.split(",")
        if len(parts)==2:
            r,c = int(parts[0]),int(parts[1])
            max_r=max(max_r,r); max_c=max(max_c,c)
    return max_r, max_c

def _get_table_headers(layout):
    cells = layout.get("cells", {})
    headers = []
    for key, cell in cells.items():
        if not isinstance(cell, dict): continue
        val = cell.get("value","").strip()
        if not val: continue
        parts = key.split(",")
        if len(parts)==2 and int(parts[0])==0:
            c = int(parts[1])
            headers.append({"col":c,"label":val,"ref":_cell_ref(0,c),"style":cell.get("style",{})})
    return sorted(headers, key=lambda x: x["col"])


# ==============================================================================
# JOB ROUTES
# ==============================================================================

@router.get("/jobs", response_model=list[JobListItem])
def list_jobs(limit: int=50, offset: int=0, status_filter: Optional[str]=None,
              db: Session=Depends(get_db), current_user: User=Depends(get_current_user)):
    q = db.query(ExtractionJob).order_by(ExtractionJob.created_at.desc())
    # Super admin sees all jobs
    if current_user.role == "admin" and not current_user.client_id:
        pass
    # Company admin sees all jobs within their company
    elif current_user.role in ("admin", "company_admin") and current_user.client_id:
        q = q.filter(ExtractionJob.client_id == current_user.client_id)
    # Regular user sees only their own jobs
    else:
        q = q.filter(ExtractionJob.user_id == current_user.id)
    if status_filter:
        q = q.filter(ExtractionJob.status == status_filter)
    return q.offset(offset).limit(limit).all()

@router.get("/jobs/{job_id}", response_model=JobStatus)
def get_job(job_id: int, db: Session=Depends(get_db), current_user: User=Depends(get_current_user)):
    return _get_job_or_404(job_id, current_user, db)

@router.get("/jobs/{job_id}/results", response_model=list[DocumentResultResponse])
def get_job_results(job_id: int, doc_type: Optional[str]=None, needs_review: Optional[bool]=None,
                    db: Session=Depends(get_db), current_user: User=Depends(get_current_user)):
    _get_job_or_404(job_id, current_user, db)
    q = db.query(DocumentResult).filter(DocumentResult.job_id==job_id)
    if doc_type: q = q.filter(DocumentResult.document_type==doc_type)
    if needs_review is not None: q = q.filter(DocumentResult.needs_review==needs_review)
    docs = q.order_by(DocumentResult.id).all()
    return [DocumentResultResponse(
        id=d.id, job_id=d.job_id, filename=d.filename,
        document_type=d.document_type, overall_confidence=d.overall_confidence,
        extracted_data=d.get_extracted_data(),
        validation_errors=d.validation_errors, validation_warnings=d.validation_warnings,
        needs_review=d.needs_review, reviewed=d.reviewed, reviewed_by=d.reviewed_by,
        model_used=d.model_used, tokens_used=d.tokens_used or 0,
        latency_ms=d.latency_ms or 0, created_at=d.created_at,
    ) for d in docs]

@router.put("/jobs/{job_id}/docs/{doc_id}")
def update_document(job_id: int, doc_id: int, payload: DocumentUpdateRequest,
                    db: Session=Depends(get_db), current_user: User=Depends(get_current_user)):
    _get_job_or_404(job_id, current_user, db)
    doc = db.query(DocumentResult).filter(DocumentResult.id==doc_id, DocumentResult.job_id==job_id).first()
    if not doc: raise HTTPException(status_code=404, detail="Document not found")
    doc.set_extracted_data(payload.extracted_data)
    doc.reviewed=True; doc.reviewed_by=current_user.username; doc.needs_review=False
    db.commit()
    return {"message": "Updated", "doc_id": doc_id}

@router.post("/jobs/{job_id}/docs/{doc_id}/approve")
def approve_document(job_id: int, doc_id: int,
                     db: Session=Depends(get_db), current_user: User=Depends(get_current_user)):
    _get_job_or_404(job_id, current_user, db)
    doc = db.query(DocumentResult).filter(DocumentResult.id==doc_id, DocumentResult.job_id==job_id).first()
    if not doc: raise HTTPException(status_code=404, detail="Document not found")
    doc.reviewed=True; doc.reviewed_by=current_user.username; doc.needs_review=False
    db.commit()
    return {"message": "Approved", "doc_id": doc_id}

@router.delete("/jobs/{job_id}")
def cancel_job(job_id: int, db: Session=Depends(get_db), current_user: User=Depends(get_current_user)):
    job = _get_job_or_404(job_id, current_user, db)
    if job.status not in ("pending","processing"):
        raise HTTPException(status_code=400, detail=f"Cannot cancel job with status '{job.status}'")
    job.status="cancelled"; job.completed_at=datetime.utcnow(); db.commit()
    return {"message": "Cancelled", "job_id": job_id}

def _get_job_or_404(job_id, current_user, db):
    job = db.query(ExtractionJob).filter(ExtractionJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    # Super admin (no client_id) sees everything
    if current_user.role == "admin" and not current_user.client_id:
        return job
    # Company admin sees all jobs within their company
    if current_user.role in ("admin", "company_admin") and current_user.client_id:
        if job.client_id == current_user.client_id:
            return job
    # Regular user sees only their own jobs
    if job.user_id == current_user.id:
        return job
    raise HTTPException(status_code=403, detail="Access denied")

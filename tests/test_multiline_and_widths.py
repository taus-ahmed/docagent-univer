"""
D9 — one join rule for multi-line values.
D14 — every written column gets a width.

Both were reported as NON-DETERMINISTIC: three different joins for one field
across runs, and auto-fit applying on some runs and not others. Neither was
random. The join was whatever the model returned, and the widths were applied
across an extent computed from cells carrying TEXT — so which columns got one
depended on where the template happened to have headings.

Both now have a stated rule and a geometry answer.
"""
import json

import pytest

from tests.harness import bootstrap as _bs

_bs.bootstrap()

from text_layer import canonical_value, read_page  # noqa: E402


# ══════════════════════════════════════════════════════════════════════════
# D9 — the join
# ══════════════════════════════════════════════════════════════════════════

#: The four joins actually observed for INV-2024-0031's Notes field, across
#: cached runs of the SAME document. Three of them are wrong and one of the
#: wrong ones merges two side-by-side columns.
NOTES = "Payment received via wire transfer Feb 10, 2024. Ref: WT-20240210-4421."


@pytest.fixture(scope="module")
def invoice_lines(pdf_dir):
    import pdfplumber
    with pdfplumber.open(pdf_dir / "INV-2024-0031.pdf") as pdf:
        _text, lines, _r = read_page(pdf.pages[0])
    return lines


class TestOneJoinRule:
    def test_a_newline_join_becomes_a_space_join(self, invoice_lines):
        got, gutter = canonical_value(
            "Payment received via wire transfer Feb 10, 2024. Ref:\n"
            "WT-20240210-4421.", invoice_lines)
        assert got == NOTES
        assert gutter is False

    def test_words_fused_together_are_separated(self, invoice_lines):
        got, _g = canonical_value(
            "Paymentreceived via wire transfer Feb 10, 2024. Ref: "
            "WT-20240210-4421.", invoice_lines)
        assert got == NOTES

    def test_an_already_correct_join_is_unchanged(self, invoice_lines):
        """Idempotent, or it would churn every correct value it touched."""
        got, _g = canonical_value(NOTES, invoice_lines)
        assert got == NOTES

    def test_the_value_continues_within_its_own_column(self, invoice_lines):
        """Reading order interleaves side-by-side columns: after "Ref:" the
        next word in the flat text is the LEFT column's "ABA:", not this
        value's own continuation. The join steps over anything outside its
        block."""
        got, _g = canonical_value(
            "Payment received via wire transfer Feb 10, 2024. Ref:\n"
            "WT-20240210-4421.", invoice_lines)
        assert "ABA" not in got and "021000021" not in got


class TestItNeitherAddsNorDrops:
    def test_a_truncated_value_is_not_extended(self, invoice_lines):
        """THE DESIGN ANSWER, stated: a field does not absorb the next line
        just because one is there. Absorbing is how one field swallows
        another's value, and a short value is a visible fault where a
        swallowed one is not."""
        first_line = "Payment received via wire transfer Feb 10, 2024."
        got, _g = canonical_value(first_line, invoice_lines)
        assert got == first_line

    def test_a_value_that_matches_no_words_is_left_alone(self, invoice_lines):
        """Anything the model derived or reformatted — a MICR field, a
        renotated number — has no run of words to re-derive from."""
        assert canonical_value("A VALUE THAT IS NOT PRINTED HERE",
                               invoice_lines) is None

    def test_a_short_value_is_never_touched(self, invoice_lines):
        assert canonical_value("40", invoice_lines) is None

    def test_an_ordinary_single_line_value_is_unchanged(self, invoice_lines):
        got, _g = canonical_value("500 Commerce Drive, Chicago, IL 60601",
                                  invoice_lines)
        assert got == "500 Commerce Drive, Chicago, IL 60601"


class TestTwoColumnsMergedIntoOneValueIsReported:
    def test_crossing_a_gutter_is_detected(self, invoice_lines):
        """INV-2024-0031 lays Payment Instructions and Notes side by side.
        Ordinary word gaps on those lines are 2.3-2.6pt; the gutter is 122pt.
        A value spanning it has merged two blocks."""
        _got, gutter = canonical_value(
            "Wire: First National Bank of New York Payment received via wire "
            "transfer Feb 10, 2024. Ref:\nABA: 021000021 Account: 7743882201 "
            "WT-20240210-4421.", invoice_lines)
        assert gutter is True

    def test_the_merged_value_is_reported_not_trimmed(self):
        """Which half was wanted is not knowable here, so it is kept, marked
        low and flagged — the same treatment as any other value the engine
        cannot stand behind."""
        from slot_extractor import run_slot_extraction
        from template_shape import compute_shape
        import pdfplumber

        with pdfplumber.open(_bs.PDF_DIR / "INV-2024-0031.pdf") as pdf:
            text, lines, _r = read_page(pdf.pages[0])
        grid = {"cells": {"0,0": {"value": "Notes"}, "0,1": {"value": ""}},
                "colWidths": [], "merges": {}, "repeatRows": [], "regions": []}
        shape = compute_shape(grid, log=lambda _m: None)
        merged = ("Wire: First National Bank of New York Payment received via "
                  "wire transfer Feb 10, 2024. Ref: ABA: 021000021 Account: "
                  "7743882201 WT-20240210-4421.")
        payload = {"fields": {"F1": {"value": merged, "source": merged,
                                     "page": 1}}, "tables": {}}

        class _Resp:
            success = True
            parsed_json = payload
            raw_text = json.dumps(payload)

        class _LLM:
            def extract(self, **kw):
                return _Resp()

        orch = type("O", (), {"llm": _LLM()})()
        ed = run_slot_extraction(
            orch, "INV.pdf", {"layout": grid, "shape": shape}, None,
            page_images=[], doc_text=text, doc_text_pages=[text],
            file_type="digital_pdf", default_doc_type="sales_invoice",
            start=0.0, page_lines=[lines])[0].extracted_data

        assert ed["validation"]["merged_column_count"] == 1
        assert ed["validation"]["confidence_map"]["B1"] == "low"
        assert ed["needs_review"] is True
        assert any("column gutter" in f["reason"]
                   for f in ed["validation"]["flagged_fields"])
        assert ed["extracted_fields"]["B1"], "the value must be kept, not dropped"


# ══════════════════════════════════════════════════════════════════════════
# D14 — the widths
# ══════════════════════════════════════════════════════════════════════════

class TestEveryWrittenColumnGetsAWidth:
    def _sheet(self, col_widths):
        import openpyxl
        from openpyxl import Workbook
        from app.api.routes.extract import _fit_columns
        ws = Workbook().active
        ws["A1"] = "Item / Description"
        ws["B1"] = "Qty"
        ws["C1"] = "A rather long extracted value that would be truncated"
        ws["D1"] = "Total"
        _fit_columns(ws, col_widths)
        return ws

    def test_no_written_column_is_left_at_the_excel_default(self):
        """The defect: widths were applied across the TEXT-carrying extent, so
        columns the writer went on to fill got none and fell back to 8.43 —
        truncating extracted values and the user's own labels."""
        ws = self._sheet([])
        for col in "ABCD":
            assert ws.column_dimensions[col].width is not None, col
            assert ws.column_dimensions[col].width >= 8

    def test_a_column_with_no_stored_width_is_fitted_to_its_content(self):
        ws = self._sheet([])
        assert ws.column_dimensions["C"].width > ws.column_dimensions["B"].width

    def test_the_users_own_width_wins_where_there_is_one(self):
        """It is the width they dragged."""
        ws = self._sheet([700, 0, 0, 0])
        assert ws.column_dimensions["A"].width == 100

    def test_a_zero_or_missing_stored_width_falls_back_to_fitting(self):
        ws = self._sheet([0, None])
        assert ws.column_dimensions["A"].width >= len("Item / Description")

    def test_one_long_note_cannot_push_the_sheet_off_screen(self):
        import openpyxl
        from openpyxl import Workbook
        from app.api.routes.extract import _fit_columns
        ws = Workbook().active
        ws["A1"] = "x" * 500
        _fit_columns(ws, [])
        assert ws.column_dimensions["A"].width == 60

    def test_it_is_deterministic(self):
        """The reported symptom was 'applies on some runs and not others'."""
        widths = [tuple(self._sheet([]).column_dimensions[c].width
                        for c in "ABCD") for _ in range(3)]
        assert len(set(widths)) == 1, widths


# ══════════════════════════════════════════════════════════════════════════
# The join across a line break — a wrap is not always a space
# ══════════════════════════════════════════════════════════════════════════

@pytest.fixture(scope="module")
def cd_contact_lines(pdf_dir):
    """The Closing Disclosure's contact page: five parties across, nine rows
    down, and values wrapped inside columns 10-26pt apart."""
    import pdfplumber
    with pdfplumber.open(
            pdf_dir / "201403_cfpb_closing-disclosure_cover-H25B.pdf") as pdf:
        _t, lines, _r = read_page(pdf.pages[5])
    return lines


class TestAWrappedTokenIsNotGluedWithASpace:
    @pytest.mark.parametrize("given,expected", [
        ("sarah@\nepsilontitle.com", "sarah@epsilontitle.com"),
        ("joesmith@\nficusbank.com", "joesmith@ficusbank.com"),
    ])
    def test_a_wrapped_email_comes_back_whole(self, cd_contact_lines,
                                              given, expected):
        """The regression this file's own join rule introduced. An address
        glued with a space — `sarah@ epsilontitle.com` — looks perfectly
        ordinary in a spreadsheet and bounces when anyone uses it, which is
        precisely the category of failure the text layer exists to remove."""
        got, _crossed = canonical_value(given, cd_contact_lines)
        assert got == expected

    @pytest.mark.parametrize("given,expected", [
        ("Omega Real Estate\nBroker Inc.", "Omega Real Estate Broker Inc."),
        ("987 Suburb Ct.\nSomeplace, ST 12340",
         "987 Suburb Ct. Someplace, ST 12340"),
        ("123 Commerce Pl.\nSomecity, ST 12344",
         "123 Commerce Pl. Somecity, ST 12344"),
    ])
    def test_wrapped_prose_keeps_its_space(self, cd_contact_lines,
                                           given, expected):
        """The other half. A full stop before a capital ends an abbreviation,
        so `Pl.` + `Somecity` must not become `Pl.Somecity`."""
        got, _crossed = canonical_value(given, cd_contact_lines)
        assert got == expected

    def test_the_seam_decides_not_the_line_break(self):
        from text_layer import _joiner
        assert _joiner("joesmith@", "ficusbank.com", same_line=False) == ""
        assert _joiner("ficusbank.", "com", same_line=False) == ""
        assert _joiner("Non-", "Negotiable", same_line=False) == ""
        assert _joiner("joesmith", "@ficusbank.com", same_line=False) == ""
        assert _joiner("Estate", "Broker", same_line=False) == " "
        assert _joiner("Pl.", "Somecity,", same_line=False) == " "

    def test_within_a_line_it_is_always_a_space(self):
        """The PDF really did put whitespace there."""
        from text_layer import _joiner
        assert _joiner("joesmith@", "ficusbank.com", same_line=True) == " "


class TestTheGutterIsMeasuredFromThePage:
    def test_a_tightly_packed_matrix_still_has_columns(self, cd_contact_lines):
        """A fixed 24pt threshold read this matrix as one continuous line —
        its columns are 10-26pt apart, where INV-2024-0031's are 122pt."""
        from text_layer import gutter_for_page
        assert 6.0 < gutter_for_page(cd_contact_lines) < 15.0

    def test_a_widely_spaced_page_gets_a_wider_threshold(self, pdf_dir):
        import pdfplumber
        from text_layer import gutter_for_page
        with pdfplumber.open(pdf_dir / "INV-2024-0031.pdf") as pdf:
            _t, lines, _r = read_page(pdf.pages[0])
        assert gutter_for_page(lines) > 6.0

    def test_a_row_of_one_word_cells_does_not_set_the_scale(self,
                                                            cd_contact_lines):
        """Per LINE fails here and the way it fails is instructive: on a line
        that is nothing but one-word cells — five email addresses — EVERY gap
        is a gutter, so the line's own median gap IS one. The page is the
        right scale because most gaps on it are ordinary word spacing."""
        got, _c = canonical_value("joesmith@\nficusbank.com", cd_contact_lines)
        assert got == "joesmith@ficusbank.com"


# ══════════════════════════════════════════════════════════════════════════
# An identifier is not a quantity — and a quantity is not an identifier
# ══════════════════════════════════════════════════════════════════════════

class TestTheLabelDecidesWhatADigitRunIs:
    @pytest.mark.parametrize("value,label", [
        ("222222", "NMLS ID"),
        ("12345", "Contact NMLS ID"),
        ("7743882201", "Account No"),
        ("445120", "Invoice Number"),
        ("123456789", "Loan ID"),
        ("98765", "Policy"),
        ("60601", "Zip"),
        ("1234567890", "Phone"),
    ])
    def test_an_identifier_stays_text(self, value, label):
        """`NMLS ID 222222` exported as 222222.0. Six digits and no leading
        zero, so no rule about the SHAPE of the digits could save it — 222222
        is a perfectly ordinary number. The column it sits in is headed "NMLS
        ID", and that is evidence the value itself does not carry."""
        from app.api.routes.extract import coerce_cell_value
        assert coerce_cell_value(value, label) == value

    @pytest.mark.parametrize("value,label,expected", [
        ("40", "Qty", 40.0),
        ("40", "Quantity", 40.0),
        ("500", "Units", 500.0),
        ("180000", "Total", 180000.0),
        ("180000", "Sale Price", 180000.0),
        ("1234567", "Account Balance", 1234567.0),
        ("12570", "Invoice Total", 12570.0),
    ])
    def test_a_quantity_stays_a_number(self, value, label, expected):
        """The mirror-image failure, and the more damaging one: a rule that
        turns quantities into text breaks every sum in the sheet. Quantity
        words WIN over identifier words, because the overlap is real and always
        resolves the same way — "Account Balance" is money, "Account No" is
        not."""
        from app.api.routes.extract import coerce_cell_value
        assert coerce_cell_value(value, label) == expected

    def test_an_unlabelled_digit_run_falls_back_to_its_shape(self):
        """The cost, stated: a 6-8 digit identifier in a cell whose label says
        nothing is still read as a number. The label is the only evidence
        available, and without it 222222 is just a number."""
        from app.api.routes.extract import coerce_cell_value
        assert coerce_cell_value("222222", "") == 222222.0
        assert coerce_cell_value("021000021", "") == "021000021"
        assert coerce_cell_value("123456789", "") == "123456789"

    def test_an_identifier_gets_no_number_format(self):
        from app.api.routes.extract import cell_format
        assert cell_format("222222", "NMLS ID") is None
        assert cell_format("180000", "Total") == "#,##0"

    def test_the_writer_passes_the_label_through(self):
        """End to end: the value reaches the sheet as text, not 222222.0."""
        import openpyxl
        from openpyxl import Workbook
        from app.api.routes.extract import _write_slot_excel
        grid = {"cells": {"0,0": {"value": "NMLS ID"}, "0,1": {"value": ""}},
                "colWidths": [], "merges": {}, "repeatRows": [], "regions": []}
        ed = {"extracted_fields": {"B1": "222222"},
              "slot_map": {"fields": [{"ref": "B1", "row_label": "NMLS ID"}],
                           "tables": []}}

        class _Doc:
            def get_extracted_data(self):
                return ed

        ws = Workbook().active
        _write_slot_excel(ws, [_Doc()], grid, grid["cells"], openpyxl)
        assert ws["B1"].value == "222222"

"""
A template can declare its own tables instead of being guessed at.

Detection looks for headings with empty rows beneath them. That describes one
shape: a table running downwards with a solid heading line. A transposed table —
headings down the left, one record per column — has no empty rows beneath
anything, so detection finds no band at all and reads the headings as unrelated
single fields. The document extracts, the sheet fills, and the answer is quietly
wrong. That is the same failure class as the 16-word router: a narrow rule
applied to a shape it cannot describe, with no signal that it did not fit.

A declaration says what the region IS. The user selects the table as they see
it — heading line included — and names which way the records run. Detection
still runs everywhere nothing is declared, so existing templates are untouched.
"""
from tests.harness import bootstrap as bs

bs.bootstrap()

from template_shape import compute_shape, choose_path  # noqa: E402


def _grid(cells, regions=None, cols=8, rows=16):
    g = {f"{r},{c}": {"value": "", "style": {}}
         for r in range(rows) for c in range(cols)}
    for k, v in cells.items():
        g[k] = {"value": v, "style": {}}
    out = {"cells": g, "colWidths": [140] * cols, "merges": {}, "repeatRows": []}
    if regions is not None:
        out["regions"] = regions
    return out


# A transposed table: the field names run DOWN column A, and each employee is a
# COLUMN. Nothing about it looks like a band to the detector.
TRANSPOSED = {
    "0,0": "Payslip Summary",
    "2,0": "Employee Name",
    "3,0": "Employee ID",
    "4,0": "Gross Pay",
    "5,0": "Tax",
    "6,0": "Net Pay",
}
TRANSPOSED_REGION = [{"type": "table", "r1": 2, "c1": 0, "r2": 6, "c2": 4,
                      "orientation": "columns", "name": "Employees"}]


class TestDetectionCannotSeeATransposedTable:
    def test_undeclared_it_finds_no_table(self):
        """The failure being fixed: not an error, just a wrong answer."""
        shape = compute_shape(_grid(TRANSPOSED))
        assert shape["repeat_bands"] == []
        # five table headings read as five unrelated single fields, plus the
        # sheet title — six answers where the document has one table
        assert len(shape["field_slots"]) == 6

    def test_declared_it_becomes_one_table(self):
        shape = compute_shape(_grid(TRANSPOSED, TRANSPOSED_REGION))
        assert len(shape["repeat_bands"]) == 1
        b = shape["repeat_bands"][0]
        assert b["orientation"] == "columns"
        assert b["declared"] is True
        assert [f["header"] for f in b["fields"]] == [
            "Employee Name", "Employee ID", "Gross Pay", "Tax", "Net Pay"]
        assert (b["header_col"], b["start_col"], b["end_col"]) == (0, 1, 4)

    def test_the_declared_rows_are_no_longer_loose_fields(self):
        """A cell cannot be both a table cell and a standalone field — that is
        two answers for one address."""
        shape = compute_shape(_grid(TRANSPOSED, TRANSPOSED_REGION))
        # only the sheet title is left; the five table headings are gone
        assert [f["row_label"] for f in shape["field_slots"]] == ["Payslip Summary"]

    def test_the_heading_column_is_not_a_value_column(self):
        shape = compute_shape(_grid(TRANSPOSED, TRANSPOSED_REGION))
        assert 0 in shape["label_columns"]
        assert 0 not in shape["value_columns"]
        assert {1, 2, 3, 4} <= set(shape["value_columns"])

    def test_it_asks_for_four_columns_not_two(self):
        shape = compute_shape(_grid(TRANSPOSED, TRANSPOSED_REGION))
        assert shape["required_columns"] == 5      # heading + 4 record columns
        assert choose_path(shape, slot_doc_types=None)["error"] is None


# A normal downward table, declared rather than detected.
DOWNWARD = {"5,0": "Date", "5,1": "Description", "5,2": "Amount"}
DOWNWARD_REGION = [{"type": "table", "r1": 5, "c1": 0, "r2": 11,
                    "orientation": "rows", "c2": 2, "name": "Charges"}]


class TestDeclarationBeatsDetection:
    def test_a_declared_downward_table_matches_the_detected_one(self):
        """Declaring a table the detector would have found anyway must not
        change the answer, or declaring becomes a risk."""
        detected = compute_shape(_grid(DOWNWARD))["repeat_bands"][0]
        declared = compute_shape(_grid(DOWNWARD, DOWNWARD_REGION))["repeat_bands"][0]
        assert [c["header"] for c in declared["columns"]] == \
               [c["header"] for c in detected["columns"]]
        assert declared["header_row"] == detected["header_row"]
        assert declared["start_row"] == detected["start_row"]

    def test_a_declaration_bounds_the_table_where_detection_runs_on(self):
        """Detection ends a band at the next row with any text. A declaration
        ends it exactly where the user drew it."""
        cells = dict(DOWNWARD)
        detected = compute_shape(_grid(cells))["repeat_bands"][0]
        declared = compute_shape(_grid(cells, DOWNWARD_REGION))["repeat_bands"][0]
        assert declared["end_row"] == 11            # exactly where it was drawn
        assert detected["end_row"] == 15           # detection runs to the grid's end

    def test_detection_still_runs_outside_declared_regions(self):
        cells = dict(DOWNWARD)
        cells["0,0"] = "Statement Number"
        cells["1,0"] = "Account Holder"
        shape = compute_shape(_grid(cells, DOWNWARD_REGION))
        assert len(shape["repeat_bands"]) == 1
        assert [f["row_label"] for f in shape["field_slots"]] == \
               ["Statement Number", "Account Holder"]


class TestBadDeclarationsAreRefusedNotObeyed:
    def _shape(self, regions):
        logs = []
        return compute_shape(_grid(TRANSPOSED, regions), log=logs.append), logs

    def test_a_region_with_no_data_area_is_ignored_with_a_reason(self):
        shape, logs = self._shape([{"type": "table", "r1": 2, "c1": 0,
                                    "r2": 6, "c2": 0, "orientation": "columns"}])
        assert shape["repeat_bands"] == []
        assert any("no data columns" in m for m in logs), logs

    def test_a_heading_row_with_nothing_beneath_is_ignored(self):
        shape, logs = self._shape([{"type": "table", "r1": 5, "c1": 0,
                                    "r2": 5, "c2": 3, "orientation": "rows"}])
        assert shape["repeat_bands"] == []
        assert any("no data rows" in m for m in logs), logs

    def test_a_malformed_region_does_not_crash_the_shape(self):
        for bad in [[{"type": "table"}], [{"type": "table", "r1": "x", "c1": 0,
                                           "r2": 4, "c2": 3}], ["nonsense"],
                    [{"type": "note", "r1": 0, "c1": 0, "r2": 4, "c2": 3}],
                    "not a list", None]:
            shape = compute_shape(_grid(TRANSPOSED, bad))
            assert shape["repeat_bands"] == []
            assert len(shape["field_slots"]) == 6    # detection unaffected

    def test_reversed_corners_are_normalised(self):
        """Selecting bottom-right to top-left is the same selection."""
        a = compute_shape(_grid(TRANSPOSED, TRANSPOSED_REGION))["repeat_bands"][0]
        b = compute_shape(_grid(TRANSPOSED, [{"type": "table", "r1": 6, "c1": 4,
                                              "r2": 2, "c2": 0,
                                              "orientation": "columns"}]
                                ))["repeat_bands"][0]
        assert (a["header_col"], a["start_col"], a["end_col"]) == \
               (b["header_col"], b["start_col"], b["end_col"])
        assert [f["row"] for f in a["fields"]] == [f["row"] for f in b["fields"]]


class TestTemplatesWithNoDeclarationsAreUntouched:
    def test_absent_and_empty_regions_give_the_same_shape_as_before(self):
        base = compute_shape(_grid(DOWNWARD))
        for regions in ([], None):
            other = compute_shape(_grid(DOWNWARD, regions))
            assert other["summary"] == base["summary"]
            assert other["field_slots"] == base["field_slots"]


class TestTheSheetIsFilledSideways:
    """A declaration is only worth having if the writer honours it. The
    transposed table must come out of the export with records running ACROSS,
    the heading column intact, and nothing written over the labels."""

    def _sheet(self, records):
        import contextlib
        import io as _io
        import json

        import openpyxl
        from app.api.routes.extract import _write_slot_excel
        from app.models.models import DocumentResult
        from slot_extractor import _table_map

        grid = _grid(TRANSPOSED, TRANSPOSED_REGION)
        band = compute_shape(grid)["repeat_bands"][0]
        ed = {"extracted_fields": {},
              "slot_map": {"fields": [], "tables": [_table_map(band)]},
              f"{band['name']}_rows": records}
        doc = DocumentResult(filename="P.pdf", document_type="payslip",
                             extraction_json=json.dumps(ed, default=str))
        wb = openpyxl.Workbook()
        ws = wb.active
        with contextlib.redirect_stdout(_io.StringIO()):
            _write_slot_excel(ws, [doc], grid, grid["cells"], openpyxl)
        return ws

    RECORDS = [
        {"Employee Name": "A Khan", "Employee ID": "E-01",
         "Gross Pay": "5000", "Tax": "800", "Net Pay": "4200"},
        {"Employee Name": "B Silva", "Employee ID": "E-02",
         "Gross Pay": "6100", "Tax": "1000", "Net Pay": "5100"},
    ]

    def test_each_record_is_a_column(self):
        ws = self._sheet(self.RECORDS)
        assert [ws["B3"].value, ws["B4"].value, ws["B7"].value] == \
               ["A Khan", "E-01", 4200]
        assert [ws["C3"].value, ws["C4"].value, ws["C7"].value] == \
               ["B Silva", "E-02", 5100]

    def test_the_heading_column_survives(self):
        """The transposed table's first column is the template's own labels —
        the writer must not treat those rows as the table's to overwrite."""
        ws = self._sheet(self.RECORDS)
        assert [ws[f"A{r}"].value for r in range(3, 8)] == [
            "Employee Name", "Employee ID", "Gross Pay", "Tax", "Net Pay"]

    def test_more_records_than_columns_widen_the_sheet(self):
        """Overflow in a transposed table goes sideways. Pushing rows down
        would scatter one record across two places."""
        ws = self._sheet(self.RECORDS * 3)          # 6 records, 4 drawn columns
        assert ws["F3"].value == "A Khan"           # 5th record, past the template
        assert ws["G3"].value == "B Silva"          # 6th
        assert ws["A3"].value == "Employee Name"    # still not overwritten

    def test_no_records_leaves_the_template_alone(self):
        ws = self._sheet([])
        assert ws["A3"].value == "Employee Name"
        assert ws["B3"].value is None

"""
Unit tests for no-template shape inference (Phase 3). Offline, always green.

The point of Phase 3 is that no-template is NOT a second engine: inference
produces a real template, and everything downstream treats it exactly like a
user's. These tests hold that boundary — the inferred grid must satisfy the
same one rule, produce the same kind of shape, and group into sheets by shape
rather than one sheet per document.
"""
import json

import pytest

from tests.harness import bootstrap as bs

bs.bootstrap()

from shape_inference import (  # noqa: E402
    _clean, build_grid, saveable_template, signature,
)
from template_shape import compute_shape, is_usable  # noqa: E402

STATEMENT = {
    "document_type": "bank_statement",
    "title": "Account Statement",
    "fields": ["Bank Name", "Account Holder", "Statement No"],
    "tables": [{"name": "Transactions",
                "columns": ["Date", "Description", "Debit", "Credit", "Balance"],
                "row_count": 12}],
    "totals": ["Closing Balance"],
}


class TestClean:
    def test_drops_one_column_tables(self):
        """A single-column 'table' is a list of fields, not a table."""
        out = _clean({"tables": [{"name": "X", "columns": ["Only"]}]})
        assert out["tables"] == []

    def test_dedupes_and_trims_labels(self):
        out = _clean({"fields": ["  Invoice No :", "Invoice No", "", None, "Date"]})
        assert out["fields"] == ["Invoice No", "Date"]

    def test_accepts_labels_given_as_objects(self):
        out = _clean({"fields": [{"label": "Bank Name"}, {"name": "Period"}]})
        assert out["fields"] == ["Bank Name", "Period"]

    def test_row_count_is_clamped_to_something_sane(self):
        # unspecified -> a workable default; the writer expands the band anyway
        assert _clean({"tables": [{"columns": ["A", "B"], "row_count": 0}]}
                      )["tables"][0]["row_count"] == 8
        # a floor, so a one-row table still has somewhere to grow
        assert _clean({"tables": [{"columns": ["A", "B"], "row_count": 1}]}
                      )["tables"][0]["row_count"] == 3
        assert _clean({"tables": [{"columns": ["A", "B"], "row_count": 99999}]}
                      )["tables"][0]["row_count"] == 200

    def test_garbage_yields_an_empty_but_valid_template(self):
        out = _clean({})
        assert out["document_type"] == "other"
        assert out["fields"] == [] and out["tables"] == [] and out["totals"] == []


class TestGrid:
    def test_inferred_grid_obeys_the_one_rule(self):
        """The whole design rests on this: the inferred grid must be readable
        by the SAME shape rule as a user's grid, with no special casing."""
        shape = compute_shape(build_grid(STATEMENT))
        assert is_usable(shape)
        assert [f["row_label"] for f in shape["field_slots"]] == [
            "Bank Name", "Account Holder", "Statement No", "Closing Balance"]
        assert len(shape["repeat_bands"]) == 1
        band = shape["repeat_bands"][0]
        assert [c["header"] for c in band["columns"]] == [
            "Date", "Description", "Debit", "Credit", "Balance"]
        assert shape["required_columns"] == 5

    def test_band_gets_the_rows_inference_asked_for(self):
        band = compute_shape(build_grid(STATEMENT))["repeat_bands"][0]
        assert band["end_row"] - band["start_row"] + 1 == 12

    def test_band_is_named_after_its_section(self):
        assert compute_shape(build_grid(STATEMENT))["repeat_bands"][0]["name"] == \
            "Transactions"

    def test_totals_are_field_slots_not_band_rows(self):
        shape = compute_shape(build_grid(STATEMENT))
        band_rows = range(shape["repeat_bands"][0]["start_row"],
                          shape["repeat_bands"][0]["end_row"] + 1)
        total = next(f for f in shape["field_slots"]
                     if f["row_label"] == "Closing Balance")
        assert total["row"] not in band_rows

    def test_fields_only_document_still_works(self):
        shape = compute_shape(build_grid(
            _clean({"fields": ["Payee", "Amount"], "tables": []})))
        assert is_usable(shape)
        assert shape["repeat_bands"] == []

    def test_grid_is_the_editor_format(self):
        grid = build_grid(STATEMENT)
        assert set(grid) >= {"cells", "colWidths", "merges", "repeatRows"}
        assert all("," in k for k in grid["cells"])
        assert all("value" in v for v in grid["cells"].values())


class TestSignature:
    def test_same_structure_same_signature(self):
        assert signature(STATEMENT) == signature(json.loads(json.dumps(STATEMENT)))

    def test_signature_ignores_case_and_title(self):
        other = json.loads(json.dumps(STATEMENT))
        other["title"] = "Completely Different Title"
        other["fields"] = [f.upper() for f in other["fields"]]
        assert signature(other) == signature(STATEMENT)

    def test_different_columns_different_signature(self):
        other = json.loads(json.dumps(STATEMENT))
        other["tables"][0]["columns"].append("Reference")
        assert signature(other) != signature(STATEMENT)

    def test_different_document_type_different_signature(self):
        other = json.loads(json.dumps(STATEMENT))
        other["document_type"] = "sales_invoice"
        assert signature(other) != signature(STATEMENT)


class TestSaveableArtifact:
    def test_is_the_payload_the_template_api_accepts(self):
        art = saveable_template(STATEMENT, build_grid(STATEMENT))
        assert art["name"] == "Account Statement"
        assert art["document_type"] == "bank_statement"
        assert json.loads(art["description"])["cells"]        # a real grid
        names = [c["name"] for c in art["columns"]]
        assert "Bank Name" in names and "Date" in names and "Closing Balance" in names
        assert all({"name", "type", "order"} <= set(c) for c in art["columns"])

    def test_table_columns_are_marked_as_line_items(self):
        art = saveable_template(STATEMENT, build_grid(STATEMENT))
        by_name = {c["name"]: c for c in art["columns"]}
        assert by_name["Date"]["extraction_type"] == "lineitem"
        assert "extraction_type" not in by_name["Bank Name"]

    def test_saved_template_reproduces_the_same_shape(self):
        """Round trip: the artifact a user saves must extract identically to
        the inferred template it came from."""
        art = saveable_template(STATEMENT, build_grid(STATEMENT))
        reloaded = compute_shape(json.loads(art["description"]))
        assert reloaded["field_slots"] == compute_shape(build_grid(STATEMENT))["field_slots"]


def _doc(grid, sig, title):
    from app.models.models import DocumentResult
    return DocumentResult(
        filename=f"{title}.pdf", document_type="x",
        extraction_json=json.dumps({"inferred_grid": grid, "shape_signature": sig,
                                    "inferred_template": {"title": title},
                                    "extracted_fields": {}, "slot_map": {}}))


class TestBatchSheets:
    """Documents sharing an inferred shape share ONE sheet. Never one sheet per
    document when the shapes match."""

    def _write(self, docs):
        import openpyxl
        from app.api.routes.extract import _write_inferred_sheets
        import contextlib, io as _io
        wb = openpyxl.Workbook()
        with contextlib.redirect_stdout(_io.StringIO()):
            ok = _write_inferred_sheets(wb, wb.active, docs, openpyxl)
        return ok, wb

    def test_matching_shapes_stack_into_one_sheet(self):
        grid = build_grid(STATEMENT)
        sig = signature(STATEMENT)
        ok, wb = self._write([_doc(grid, sig, "Statement"),
                              _doc(grid, sig, "Statement"),
                              _doc(grid, sig, "Statement")])
        assert ok is True
        assert len(wb.sheetnames) == 1

    def test_different_shapes_get_their_own_sheet(self):
        other = json.loads(json.dumps(STATEMENT))
        other["document_type"] = "sales_invoice"
        other["title"] = "Invoice"
        ok, wb = self._write([_doc(build_grid(STATEMENT), signature(STATEMENT), "Statement"),
                              _doc(build_grid(other), signature(other), "Invoice")])
        assert ok is True
        assert len(wb.sheetnames) == 2

    def test_sheet_is_named_after_the_document(self):
        ok, wb = self._write([_doc(build_grid(STATEMENT), signature(STATEMENT),
                                   "Account Statement")])
        assert wb.sheetnames == ["Account Statement"]

    def test_duplicate_titles_do_not_collide(self):
        a = json.loads(json.dumps(STATEMENT))
        b = json.loads(json.dumps(STATEMENT))
        b["tables"][0]["columns"].append("Reference")   # different shape, same title
        ok, wb = self._write([_doc(build_grid(a), signature(a), "Statement"),
                              _doc(build_grid(b), signature(b), "Statement")])
        assert len(wb.sheetnames) == 2
        assert len(set(wb.sheetnames)) == 2

    def test_nothing_inferred_means_no_sheets_claimed(self):
        """So the export falls through to the flat writer rather than
        producing an empty workbook."""
        from app.models.models import DocumentResult
        d = DocumentResult(filename="x.pdf", document_type="x",
                           extraction_json=json.dumps({"extracted_fields": {}}))
        ok, _wb = self._write([d])
        assert ok is False

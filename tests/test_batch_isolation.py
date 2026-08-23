"""
Phase 4 — batch isolation.

Nothing before this tested cross-document contamination: the harness scores
documents one at a time, so a batch where document 2's values bled into
document 1's sheet would score perfectly and ship. That is the failure that
ends a client relationship fastest — a client's invoice showing another
client's totals — so it gets its own tests.

Three places contamination could enter, all covered here:

  1. EXTRACTION   a batch shares one template_data dict across documents; if
                  anything mutates it per-document, later documents extract
                  against an altered shape.
  2. EXPORT       one worksheet holds every document of a batch, stacked by a
                  computed row offset. A wrong offset overwrites a neighbour.
  3. GROUPING     with no template, documents are grouped into sheets by
                  inferred shape. A wrong group puts one document's rows under
                  another's headings.

The fixtures are five real bank statements whose identifiers, periods and
closing balances are all distinct, so contamination is directly observable
rather than inferred.
"""
import contextlib
import io
import json

import pytest

from tests.harness import bootstrap as bs

bs.bootstrap()

from tests.harness.llm_cache import LLMCache  # noqa: E402
from tests.harness.runner import _schema_path  # noqa: E402

# Each statement's own values, read from the PDFs independently of the engine.
SIBLINGS = {
    "STMT-2024-01": {"stmt": "STMT-2024-01", "period": "January 2024",
                     "closing": "125,357.26"},
    "STMT-2024-02": {"stmt": "STMT-2024-02", "period": "February 2024",
                     "closing": "43,703.39"},
    "STMT-2024-03": {"stmt": "STMT-2024-03", "period": "March 2024",
                     "closing": "112,273.46"},
    "STMT-2024-04": {"stmt": "STMT-2024-04", "period": "April 2024",
                     "closing": "71,801.10"},
    "STMT-2024-05": {"stmt": "STMT-2024-05", "period": "May 2024",
                     "closing": "21,011.93"},
}


def _numbers(text):
    """The set of numeric values in a blob of text, as discrete tokens.

    Deliberately NOT a digit-substring search: concatenating every digit in a
    block and looking for a subsequence produces false contamination reports,
    because two adjacent unrelated numbers can spell a third. A value is only
    present if it appears as a number in its own right.
    """
    import re as _re
    out = set()
    for tok in _re.findall(r"\d[\d,]*\.?\d*", str(text)):
        t = tok.replace(",", "").rstrip(".")
        if not t:
            continue
        try:
            out.add(round(float(t), 2))
        except ValueError:
            continue
    return out


def _num(v):
    return round(float(str(v).replace(",", "")), 2)


def _own_text(stem):
    """The document's own source text, read independently of the engine."""
    import pdfplumber
    with pdfplumber.open(str(bs.PDF_DIR / f"{stem}.pdf")) as pdf:
        return "\n".join((pg.extract_text() or "") for pg in pdf.pages)


def _template_data(template_name, doc_type):
    from app.api.routes.extract import _parse_template
    from app.models.models import ColumnTemplate
    grid = json.loads((bs.TEMPLATES_DIR / template_name).read_text(encoding="utf-8"))
    tpl = ColumnTemplate(name=template_name, document_type=doc_type,
                         description=json.dumps(grid), columns_json="[]")
    return _parse_template(tpl), grid


def _extract(pdf_stem, template_data):
    from app.api.routes.extract import _extract_with_template
    from orchestrator import Orchestrator
    bs.chdir_backend()
    with contextlib.redirect_stdout(io.StringIO()):
        orch = Orchestrator(client_schema_path=_schema_path())
        return _extract_with_template(orch, bs.PDF_DIR / f"{pdf_stem}.pdf",
                                      template_data)


def _doc_result(r):
    from app.models.models import DocumentResult
    ed = getattr(r, "extracted_data", None) or {}
    return DocumentResult(filename=getattr(r, "filename", "x.pdf"),
                          document_type=getattr(r, "document_type", "") or "x",
                          extraction_json=json.dumps(ed, default=str))


def _all_text(ed):
    """Every value the engine produced for one document, as one string."""
    parts = []

    def walk(o):
        if isinstance(o, dict):
            for k, v in o.items():
                if str(k).startswith("_") or k in ("raw_llm_responses",
                                                   "slot_map", "inferred_grid",
                                                   "inferred_template"):
                    continue
                walk(v)
        elif isinstance(o, list):
            for v in o:
                walk(v)
        elif o is not None:
            parts.append(str(o))
    walk(ed)
    return " | ".join(parts)


@pytest.fixture(scope="module")
def batch():
    """Five statements through ONE template, extracted as one batch would be."""
    cache = LLMCache(mode="replay")
    cache.install()
    try:
        td, grid = _template_data("bank_statement.json", "bank_statement")
        shape_before = json.dumps(td.get("shape"), sort_keys=True, default=str)
        out = {}
        for stem in SIBLINGS:
            cache.context = stem
            out[stem] = _extract(stem, td)
        shape_after = json.dumps(td.get("shape"), sort_keys=True, default=str)
    finally:
        cache.uninstall()
    return {"results": out, "grid": grid,
            "shape_before": shape_before, "shape_after": shape_after}


class TestExtractionIsolation:
    def test_each_document_returns_its_own_identifier(self, batch):
        for stem, results in batch["results"].items():
            text = _all_text(results[0].extracted_data)
            assert SIBLINGS[stem]["stmt"] in text, (
                f"{stem} did not return its own statement number")

    def test_no_document_returns_a_siblings_identifier(self, batch):
        for stem, results in batch["results"].items():
            text = _all_text(results[0].extracted_data)
            for other, vals in SIBLINGS.items():
                if other == stem:
                    continue
                assert vals["stmt"] not in text, (
                    f"CONTAMINATION: {stem} contains {other}'s statement "
                    f"number {vals['stmt']!r}")

    def test_no_document_returns_a_siblings_period(self, batch):
        for stem, results in batch["results"].items():
            text = _all_text(results[0].extracted_data).casefold()
            for other, vals in SIBLINGS.items():
                if other == stem:
                    continue
                assert vals["period"].casefold() not in text, (
                    f"CONTAMINATION: {stem} contains {other}'s period "
                    f"{vals['period']!r}")

    def test_every_number_a_document_returns_is_in_its_own_pdf(self, batch):
        """The general form of "only its own values": every number in a
        document's output must appear in THAT document's source text.

        Note these statements are consecutive months of one account, so a
        sibling's closing balance legitimately appears as this month's opening
        balance. Comparing values between siblings would call that
        contamination; grounding each value in its own document does not.
        """
        for stem, results in batch["results"].items():
            own = _numbers(_own_text(stem))
            got = _numbers(_all_text(results[0].extracted_data))
            strays = sorted(n for n in got if n not in own and n > 100)
            assert not strays, (
                f"CONTAMINATION: {stem} returned {strays}, which do not "
                f"appear anywhere in its own PDF")

    def test_each_document_returns_its_own_closing_balance(self, batch):
        for stem, results in batch["results"].items():
            nums = _numbers(_all_text(results[0].extracted_data))
            assert _num(SIBLINGS[stem]["closing"]) in nums, (
                f"{stem} did not return its own closing balance")

    def test_the_shared_template_is_not_mutated_by_the_batch(self, batch):
        """One template_data dict is reused for every document in a job. If a
        document mutated its shape, every later document would extract against
        an altered template."""
        assert batch["shape_before"] == batch["shape_after"]


@pytest.fixture(scope="module")
def sheet(batch):
    """One worksheet holding all five documents, written by the real writer."""
    import openpyxl
    from app.api.routes.extract import _analyse_template_regions, _write_excel
    docs = [_doc_result(batch["results"][s][0]) for s in SIBLINGS]
    wb = openpyxl.Workbook()
    ws = wb.active
    with contextlib.redirect_stdout(io.StringIO()):
        _write_excel(ws, docs, batch["grid"],
                     _analyse_template_regions(batch["grid"]), openpyxl)
    return ws


class TestExportIsolation:
    """One worksheet, five documents, stacked by computed row offset."""

    def test_every_document_appears_in_the_sheet(self, sheet):
        text = " | ".join(str(c.value) for row in sheet.iter_rows()
                          for c in row if c.value is not None)
        for stem, vals in SIBLINGS.items():
            assert vals["stmt"] in text, f"{stem} is missing from the export"

    def test_each_documents_block_holds_only_its_own_values(self, sheet):
        """Find the row of each statement number, then check the closing
        balance nearest below it belongs to that same statement."""
        rows_of = {}
        for row in sheet.iter_rows():
            for c in row:
                for stem, vals in SIBLINGS.items():
                    if c.value is not None and str(c.value).strip() == vals["stmt"]:
                        rows_of[stem] = c.row
        assert len(rows_of) == len(SIBLINGS), rows_of

        order = sorted(rows_of.items(), key=lambda kv: kv[1])
        bounds = {}
        for i, (stem, r0) in enumerate(order):
            r1 = order[i + 1][1] - 1 if i + 1 < len(order) else sheet.max_row
            bounds[stem] = (r0, r1)

        for stem, (r0, r1) in bounds.items():
            block = " ".join(
                str(c.value) for row in sheet.iter_rows(min_row=r0, max_row=r1)
                for c in row if c.value is not None)
            own = _numbers(_own_text(stem))
            strays = sorted(n for n in _numbers(block) if n not in own and n > 100)
            assert not strays, (
                f"CONTAMINATION in the export: {stem}'s block (rows {r0}-{r1}) "
                f"contains {strays}, which are not in its own PDF")

    def test_blocks_do_not_overlap(self, sheet):
        """Two documents writing to the same rows is the mechanism by which
        one would overwrite the other."""
        starts = []
        for row in sheet.iter_rows():
            for c in row:
                if c.value is not None and str(c.value).strip() in {
                        v["stmt"] for v in SIBLINGS.values()}:
                    starts.append(c.row)
        assert len(starts) == len(set(starts)), (
            f"two documents share a row: {sorted(starts)}")


MIXED = {
    "STMT-2024-01": "STMT-2024-01",
    "INV-2024-0031": "INV-2024-0031",
    "CHQ-001847": "CHQ-001847",
}


@pytest.fixture(scope="module")
def mixed():
    """Three different document types in one job, none with a template."""
    cache = LLMCache(mode="replay")
    cache.install()
    try:
        out = {}
        for stem in MIXED:
            cache.context = stem
            out[stem] = _extract(stem, None)
    finally:
        cache.uninstall()
    return out


class TestMixedBatchNoTemplate:
    """The harder case: different document types in one job, no template, so
    every document infers its own shape and the export groups by shape."""


    def test_each_type_infers_its_own_shape(self, mixed):
        sigs = {stem: (r[0].extracted_data or {}).get("shape_signature")
                for stem, r in mixed.items()}
        assert all(sigs.values()), sigs
        assert len(set(sigs.values())) == 3, (
            f"different document types must not share an inferred shape: {sigs}")

    def test_no_document_carries_a_siblings_identifier(self, mixed):
        for stem, results in mixed.items():
            text = _all_text(results[0].extracted_data)
            for other, ident in MIXED.items():
                if other == stem:
                    continue
                assert ident not in text, (
                    f"CONTAMINATION: {stem} contains {other}'s identifier")

    def test_each_shape_gets_its_own_sheet(self, mixed):
        import openpyxl
        from app.api.routes.extract import _write_inferred_sheets
        docs = [_doc_result(mixed[s][0]) for s in MIXED]
        wb = openpyxl.Workbook()
        with contextlib.redirect_stdout(io.StringIO()):
            ok = _write_inferred_sheets(wb, wb.active, docs, openpyxl)
        assert ok is True
        assert len(wb.sheetnames) == 3, wb.sheetnames

    def test_no_sheet_contains_another_documents_identifier(self, mixed):
        import openpyxl
        from app.api.routes.extract import _write_inferred_sheets
        docs = [_doc_result(mixed[s][0]) for s in MIXED]
        wb = openpyxl.Workbook()
        with contextlib.redirect_stdout(io.StringIO()):
            _write_inferred_sheets(wb, wb.active, docs, openpyxl)

        for sheet in wb.worksheets:
            text = " ".join(str(c.value) for row in sheet.iter_rows()
                            for c in row if c.value is not None)
            present = [ident for ident in MIXED.values() if ident in text]
            assert len(present) <= 1, (
                f"sheet {sheet.title!r} contains more than one document's "
                f"identifier: {present}")


# ══════════════════════════════════════════════════════════════════════════════
# BATCH SCHEMA REUSE (Phase 8)
# ══════════════════════════════════════════════════════════════════════════════

def _extract_batch(stems, batch_schemas, cache=None):
    """Run several documents through one job's schema cache, as the job thread
    does. Returns {stem: results} and the captured log."""
    from app.api.routes.extract import _extract_with_template
    from orchestrator import Orchestrator
    bs.chdir_backend()
    out, buf = {}, io.StringIO()
    with contextlib.redirect_stdout(buf):
        orch = Orchestrator(client_schema_path=_schema_path())
        for stem in stems:
            if cache is not None:
                cache.context = stem
            out[stem] = _extract_with_template(
                orch, bs.PDF_DIR / f"{stem}.pdf", None,
                batch_schemas=batch_schemas)
    return out, buf.getvalue()


class TestBatchSchemaReuse:
    """Inference names the columns and does not name them the same way twice.
    `signature()` hashes those names and the exporter groups sheets by it, so
    fifty invoices of one design produced up to fifty sheets with slightly
    different headings instead of one sheet with fifty rows. Inferring once per
    document KIND per job makes within-batch variance exactly zero — the schema
    is not re-derived, so it cannot differ."""

    SIBS = ["STMT-2024-01", "STMT-2024-02", "STMT-2024-03"]

    @pytest.fixture(scope="class")
    @classmethod
    def batch(cls):
        cache = LLMCache(mode="replay")
        cache.install()
        try:
            schemas = {}
            results, log = _extract_batch(cls.SIBS, schemas, cache)
            yield results, log, schemas
        finally:
            cache.uninstall()

    def test_same_kind_documents_share_one_shape(self, batch):
        """The headline defect: same layout, different sheets."""
        results, _log, _s = batch
        sigs = {s: (results[s][0].extracted_data or {}).get("shape_signature")
                for s in self.SIBS}
        assert all(sigs.values()), sigs
        assert len(set(sigs.values())) == 1, (
            f"documents of one kind must share ONE shape, got {sigs}")

    def test_the_schema_is_inferred_once_not_per_document(self, batch):
        _r, log, _s = batch
        assert log.count("schema for") == 1, log
        assert log.count("no inference call") == len(self.SIBS) - 1, log

    def test_the_cache_holds_one_entry_for_one_kind(self, batch):
        _r, _log, schemas = batch
        assert list(schemas) == ["bank_statement"], schemas

    def test_reuse_does_not_leak_values_between_documents(self, batch):
        """Sharing a SHAPE must never mean sharing VALUES — the sibling
        statements have distinct identifiers, periods and closing balances."""
        results, _log, _s = batch
        for stem in self.SIBS:
            text = _all_text(results[stem][0].extracted_data)
            for other in self.SIBS:
                if other == stem:
                    continue
                assert SIBLINGS[other]["stmt"] not in text, (
                    f"CONTAMINATION: {stem} contains {other}'s identifier")
                assert _num(SIBLINGS[other]["closing"]) not in _numbers(text), (
                    f"CONTAMINATION: {stem} contains {other}'s closing balance")

    def test_each_document_still_gets_its_own_values(self, batch):
        results, _log, _s = batch
        for stem in self.SIBS:
            text = _all_text(results[stem][0].extracted_data)
            assert SIBLINGS[stem]["stmt"] in text, stem
            assert _num(SIBLINGS[stem]["closing"]) in _numbers(text), stem

    def test_a_mixed_batch_still_gets_one_shape_per_kind(self):
        """Reuse is keyed by document type, so an invoice never inherits a
        statement's schema."""
        cache = LLMCache(mode="replay")
        cache.install()
        try:
            schemas = {}
            stems = list(MIXED)
            results, _log = _extract_batch(stems, schemas, cache)
        finally:
            cache.uninstall()
        sigs = {s: (results[s][0].extracted_data or {}).get("shape_signature")
                for s in stems}
        assert all(sigs.values()), sigs
        assert len(set(sigs.values())) == len(stems), (
            f"different document kinds must not share a shape: {sigs}")

    def test_no_cache_means_infer_every_time(self):
        """batch_schemas=None is the single-document call and must be
        unchanged — the harness and every one-off extraction rely on it."""
        cache = LLMCache(mode="replay")
        cache.install()
        try:
            _r, log = _extract_batch(self.SIBS[:2], None, cache)
        finally:
            cache.uninstall()
        assert "no inference call" not in log, log
        assert "cached for the rest of this job" not in log, log

"""
D13 — what the user builds is what they receive.

The slot writer wrote VALUES and nothing else, so merged cells, centring,
shading and borders were present in the editor and absent from the file. A
balance sheet whose section headings are merged and centred on screen and
left-aligned in the download does not look like the same document, and the
person who drew it has no way to tell whether the engine understood the layout
or ignored it.

Only what the editor can actually set is read — `CellStyle` in
DocAgentSpreadsheet.tsx. The one judgement call is merges: a merge is a RANGE,
so it is re-created only where the range still means the same thing.
"""
import json

import pytest

from tests.harness import bootstrap as _bs

_bs.bootstrap()

from template_shape import compute_shape  # noqa: E402


def _grid():
    """A heading merged across four columns, a shaded bordered label column,
    and a band beneath — the shape `bank_statement_101` actually has."""
    return {
        "cells": {
            "0,0": {"value": "ACCOUNT STATEMENT",
                    "style": {"bold": True, "align": "center",
                              "bgColor": "#f9cb9c", "borderAll": True,
                              "fontSize": 14}},
            "2,0": {"value": "Bank Name", "style": {"borderAll": True}},
            "2,1": {"value": "", "style": {"borderAll": True}},
            "4,0": {"value": "Date"}, "4,1": {"value": "Amount"},
        },
        "colWidths": [], "repeatRows": [],
        "merges": {"0,0": {"rows": 1, "cols": 4}},
        "regions": [{"type": "table", "r1": 4, "c1": 0, "r2": 9, "c2": 1,
                     "orientation": "rows", "name": "T"}],
    }


def _write(grid, rows=()):
    import openpyxl
    from openpyxl import Workbook
    from app.api.routes.extract import _write_slot_excel
    from slot_extractor import run_slot_extraction

    text = "ACCOUNT STATEMENT\nBank Name FIRST NATIONAL\nDate Amount\n"
    payload = {"fields": {"F1": {"value": "FIRST NATIONAL",
                                 "source": "Bank Name FIRST NATIONAL",
                                 "page": 1}},
               "tables": {"T": list(rows)}}

    class _Resp:
        success = True
        parsed_json = payload
        raw_text = json.dumps(payload)

    class _LLM:
        def extract(self, **kw):
            return _Resp()

    shape = compute_shape(grid, log=lambda _m: None)
    ed = run_slot_extraction(
        type("O", (), {"llm": _LLM()})(), "S.pdf",
        {"layout": grid, "shape": shape}, None, page_images=[],
        doc_text=text, doc_text_pages=[text], file_type="digital_pdf",
        default_doc_type="bank_statement", start=0.0)[0].extracted_data

    class _Doc:
        def get_extracted_data(self):
            return ed

    ws = Workbook().active
    _write_slot_excel(ws, [_Doc()], grid, grid["cells"], openpyxl)
    return ws


class TestTheLookSurvivesTheExport:
    @pytest.fixture(scope="class")
    def ws(self):
        return _write(_grid())

    def test_the_merged_heading_is_merged(self, ws):
        assert "A1:D1" in [str(r) for r in ws.merged_cells.ranges]

    def test_it_is_still_centred_and_bold(self, ws):
        assert ws["A1"].alignment.horizontal == "center"
        assert ws["A1"].font.bold is True
        assert ws["A1"].font.size == 14

    def test_the_shading_survives(self, ws):
        assert ws["A1"].fill.start_color.rgb == "FFF9CB9C"

    def test_the_borders_survive(self, ws):
        assert ws["A3"].border.left.style == "thin"
        assert ws["A3"].border.bottom.style == "thin"

    def test_a_bordered_EMPTY_cell_keeps_its_border(self, ws):
        """The box the user drew round a value cell is the box the value lands
        in. Those cells carry no text by construction — they are the slots."""
        assert ws["B3"].border.right.style == "thin"

    def test_an_unstyled_cell_is_left_alone(self, ws):
        assert ws["A5"].fill.fill_type in (None, "none")
        assert ws["A5"].font.bold in (None, False)

    def test_the_value_still_arrives(self, ws):
        """Formatting must not cost content."""
        assert ws["B3"].value == "FIRST NATIONAL"


class TestAMergeIsARangeAndIsTreatedLikeOne:
    def test_a_merge_over_a_band_that_moved_is_skipped(self):
        """The writer expands a band to the document's row count, so a merge
        crossing it would land on rows that are no longer the rows it
        described — hiding real values behind a heading. Skipped, not
        guessed."""
        grid = _grid()
        grid["merges"]["4,0"] = {"rows": 2, "cols": 2}   # across the band header
        ws = _write(grid, rows=[
            {"cells": {"Date": "01/03", "Amount": "$1.00"},
             "source": "Date Amount", "page": 1}])
        merged = [str(r) for r in ws.merged_cells.ranges]
        assert "A1:D1" in merged, "the safe merge must still be applied"
        assert not any(r.startswith("A5") for r in merged), merged

    def test_a_one_by_one_merge_is_not_a_merge(self):
        grid = _grid()
        grid["merges"]["2,0"] = {"rows": 1, "cols": 1}
        ws = _write(grid)
        assert not any(str(r).startswith("A3") for r in ws.merged_cells.ranges)

    def test_a_malformed_merge_is_ignored_rather_than_raising(self):
        grid = _grid()
        grid["merges"]["not,a,key"] = {"rows": 2, "cols": 2}
        grid["merges"]["9,9"] = None
        ws = _write(grid)
        assert "A1:D1" in [str(r) for r in ws.merged_cells.ranges]


class TestItReusesTheStyleHelperTheOtherWritersUse:
    def test_there_is_only_one_style_function(self):
        """`_apply_cell_style` already existed and four writers called it; the
        slot writer simply never did. A second copy would be a second thing to
        keep in step with the editor's CellStyle."""
        import inspect

        from app.api.routes import extract as E
        src = inspect.getsource(E)
        assert src.count("def _apply_cell_style(") == 1
        assert src.count("def _parse_hex_color(") == 1

    @pytest.mark.parametrize("given,expected", [
        ("#f9cb9c", "FFF9CB9C"), ("f9cb9c", "FFF9CB9C"), ("#ABC", "FFAABBCC"),
        ("", None), (None, None), ("#12345", None),
    ])
    def test_colour_parsing(self, given, expected):
        from app.api.routes.extract import _parse_hex_color
        assert _parse_hex_color(given) == expected

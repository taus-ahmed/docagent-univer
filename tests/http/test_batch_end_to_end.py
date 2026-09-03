"""
End-to-end batch extraction through the REAL HTTP routes and the REAL job
runner.

WHY THIS EXISTS. `tests/test_batch_isolation.py` proves that
`_extract_with_template` keeps documents apart. It calls that function
directly. Nothing called `_run_extraction_sync` — the background thread the
upload endpoint actually starts — so the path a user takes through the UI
(POST /api/extract/upload -> thread -> DocumentResult rows -> GET
/api/jobs/{id}/export) had no coverage at any point along it. A batch could be
extracted perfectly and still be lost, mis-stacked, or stranded between the
extraction function and the downloaded workbook, and every existing test would
still pass.

So these tests assert on what the user receives:

  1. five documents of one kind, one template, ONE job -> five results, each
     holding only its own values, and a DOWNLOADED workbook that says so;
  2. a mixed batch with NO template -> one sheet per inferred shape, with no
     value crossing between sheets;
  3. the accounting: five in, five out, none stranded at `processing`.

The thread is handed `settings.DATABASE_URL` (extract.py:311), not the request
session, so the fixture below points that at the same scratch SQLite file the
app's `get_db` override uses. That is not a workaround for the test — it is the
reason this layer needs its own fixture: the worker reaches the database by a
completely different route than the routes do.
"""
from __future__ import annotations

import io
import time

import pytest

from tests.harness import bootstrap as bs

bs.bootstrap()

# The contamination rule — "a value is present only if it appears as a number
# in its own right" — is defined once, in the isolation suite, and imported
# rather than restated. Two definitions of what counts as contamination is one
# too many.
from tests.test_batch_isolation import (  # noqa: E402
    SIBLINGS, _all_text, _numbers, _own_text,
)

STATEMENTS = list(SIBLINGS)                       # five bank statements

#: Five documents of five different kinds. Distinct kinds on purpose: the job
#: runner shares one inferred schema per document KIND within a job
#: (`batch_schemas`), so a mixed batch exercises grouping without exercising
#: reuse, and every document infers its own shape.
MIXED = ["BS-2024-Q1", "IS-2024-Q4", "INV-2024-0031", "CHQ-001847",
         "STMT-2024-01"]

POLL_TIMEOUT_SEC = 600


# ── fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module", autouse=True)
def _worker_env(_db_url):
    """Point the worker thread at the scratch database, drop the inter-document
    sleep, and replay recorded LLM responses.

    RATE_LIMIT_DELAY is production's 2 s pause between documents; five
    documents twice over is 16 s of the suite spent asleep, measuring nothing.
    Everything else is left exactly as production has it.
    """
    import os
    from pathlib import Path

    from app.config import settings
    from tests.harness.llm_cache import LLMCache

    cwd = Path.cwd()
    bs.chdir_backend()          # production runs from backend/; so does this

    saved = (settings.DATABASE_URL, settings.RATE_LIMIT_DELAY)
    settings.DATABASE_URL = _db_url
    settings.RATE_LIMIT_DELAY = 0.0

    cache = LLMCache(mode="replay")
    cache.install()
    try:
        yield cache
    finally:
        cache.uninstall()
        settings.DATABASE_URL, settings.RATE_LIMIT_DELAY = saved
        os.chdir(cwd)


@pytest.fixture(scope="module")
def bank_template(db, users):
    """The committed gold bank-statement template, owned by the acme tenant.

    The same grid the accuracy harness extracts with, so the batch here is the
    batch that scores 100% there — any difference in the result is this layer's
    doing, not the template's.
    """
    from app.models.models import ColumnTemplate

    grid = (bs.TEMPLATES_DIR / "bank_statement.json").read_text(encoding="utf-8")
    t = (db.query(ColumnTemplate)
           .filter(ColumnTemplate.name == "e2e_bank_statement").first())
    if not t:
        t = ColumnTemplate(name="e2e_bank_statement",
                           document_type="bank_statement",
                           description=grid, columns_json="[]",
                           user_id=users["acme"].id, client_id="acme_001")
        db.add(t)
        db.commit()
        db.refresh(t)
    db.commit()        # release the write lock; the worker thread needs it
    return t


def _upload(client, auth, stems, template_id=None):
    """POST a batch exactly as the browser does: multipart, one part per file."""
    files = [("files", (f"{s}.pdf", (bs.PDF_DIR / f"{s}.pdf").read_bytes(),
                        "application/pdf")) for s in stems]
    data = {"client_id": "acme_001"}
    if template_id is not None:
        data["template_id"] = str(template_id)
    r = client.post("/api/extract/upload", files=files, data=data,
                    headers=auth["acme"])
    assert r.status_code == 202, r.text
    return r.json()


def _await_job(client, auth, job_id):
    """Poll the status route until the job leaves pending/processing.

    A job that never leaves `processing` is finding #4 in the runbook's
    known-unfixed list, so the timeout is an assertion, not a convenience.
    """
    deadline = time.time() + POLL_TIMEOUT_SEC
    seen = []
    while time.time() < deadline:
        r = client.get(f"/api/jobs/{job_id}", headers=auth["acme"])
        assert r.status_code == 200, r.text
        job = r.json()
        seen.append(job["status"])
        if job["status"] not in ("pending", "processing"):
            return job
        time.sleep(0.25)
    pytest.fail(f"job {job_id} never left processing in {POLL_TIMEOUT_SEC}s "
                f"— stranded. statuses seen: {sorted(set(seen))}")


def _results(client, auth, job_id):
    r = client.get(f"/api/jobs/{job_id}/results", headers=auth["acme"])
    assert r.status_code == 200, r.text
    return r.json()


def _results_for(client, auth, job_id, stems):
    """The job's results, with the count checked FIRST.

    Every per-document assertion below is a loop, and a loop over an empty
    list passes. That is not hypothetical: the `storage` NameError produced
    zero result rows, and twelve of these tests went green over nothing while
    the pipeline returned literally nothing. Nothing in this file iterates
    results without coming through here.
    """
    docs = _results(client, auth, job_id)
    assert len(docs) == len(stems), (
        f"expected {len(stems)} results, got {len(docs)}: "
        f"{[d['filename'] for d in docs]}")
    return docs


def _download(client, auth, job_id):
    """The exported workbook, fetched through the download route and read back
    from the bytes the browser would receive — not from an in-memory result."""
    import openpyxl
    r = client.get(f"/api/jobs/{job_id}/export", headers=auth["acme"])
    assert r.status_code == 200, r.text[:400]
    assert "spreadsheetml" in r.headers.get("content-type", "")
    return openpyxl.load_workbook(io.BytesIO(r.content))


def _sheet_text(ws):
    return " | ".join(str(c.value) for row in ws.iter_rows()
                      for c in row if c.value is not None)


# ── 1. five documents, one template, one job ─────────────────────────────────

@pytest.fixture(scope="module")
def templated_job(client, auth, bank_template):
    up = _upload(client, auth, STATEMENTS, template_id=bank_template.id)
    job = _await_job(client, auth, up["job_id"])
    return {"upload": up, "job": job, "id": up["job_id"]}


class TestTemplatedBatchThroughTheRealRoutes:

    def test_the_job_completes(self, templated_job):
        assert templated_job["job"]["status"] == "completed", templated_job["job"]

    def test_every_document_uploaded_comes_back(self, templated_job, client, auth):
        docs = _results(client, auth, templated_job["id"])
        got = sorted(d["filename"] for d in docs)
        assert got == sorted(f"{s}.pdf" for s in STATEMENTS)

    def test_each_result_holds_its_own_statement_number(self, templated_job,
                                                        client, auth):
        for d in _results_for(client, auth, templated_job["id"], STATEMENTS):
            stem = d["filename"][:-4]
            text = _all_text(d["extracted_data"] or {})
            assert SIBLINGS[stem]["stmt"] in text, (
                f"{stem} came back without its own statement number")

    def test_no_result_holds_a_siblings_values(self, templated_job, client, auth):
        for d in _results_for(client, auth, templated_job["id"], STATEMENTS):
            stem = d["filename"][:-4]
            text = _all_text(d["extracted_data"] or {})
            for other, vals in SIBLINGS.items():
                if other == stem:
                    continue
                assert vals["stmt"] not in text, (
                    f"CONTAMINATION through the job runner: {stem} contains "
                    f"{other}'s statement number {vals['stmt']!r}")

    def test_every_number_in_a_result_is_in_that_documents_own_pdf(
            self, templated_job, client, auth):
        for d in _results_for(client, auth, templated_job["id"], STATEMENTS):
            stem = d["filename"][:-4]
            own = _numbers(_own_text(stem))
            strays = sorted(n for n in _numbers(_all_text(d["extracted_data"] or {}))
                            if n not in own and n > 100)
            assert not strays, (
                f"CONTAMINATION through the job runner: {stem} returned "
                f"{strays}, which are in no part of its own PDF")

    def test_the_downloaded_workbook_holds_every_document(
            self, templated_job, client, auth):
        wb = _download(client, auth, templated_job["id"])
        text = _sheet_text(wb.worksheets[0])
        missing = [s for s in STATEMENTS if SIBLINGS[s]["stmt"] not in text]
        assert not missing, f"missing from the downloaded workbook: {missing}"

    def test_each_documents_block_in_the_download_is_its_own(
            self, templated_job, client, auth):
        """The download is one sheet with five documents stacked by computed
        row offset. A wrong offset overwrites a neighbour, and the only place
        that is visible is the file the user opens."""
        ws = _download(client, auth, templated_job["id"]).worksheets[0]
        rows_of = {}
        for row in ws.iter_rows():
            for c in row:
                for stem, vals in SIBLINGS.items():
                    if c.value is not None and str(c.value).strip() == vals["stmt"]:
                        rows_of[stem] = c.row
        assert len(rows_of) == len(SIBLINGS), (
            f"only {sorted(rows_of)} are identifiable in the download")

        order = sorted(rows_of.items(), key=lambda kv: kv[1])
        for i, (stem, r0) in enumerate(order):
            r1 = order[i + 1][1] - 1 if i + 1 < len(order) else ws.max_row
            block = " ".join(str(c.value)
                             for row in ws.iter_rows(min_row=r0, max_row=r1)
                             for c in row if c.value is not None)
            own = _numbers(_own_text(stem))
            strays = sorted(n for n in _numbers(block) if n not in own and n > 100)
            assert not strays, (
                f"CONTAMINATION in the downloaded workbook: {stem}'s block "
                f"(rows {r0}-{r1}) contains {strays}, not in its own PDF")

    def test_the_download_agrees_with_the_stored_result(
            self, templated_job, client, auth):
        """The export is rebuilt from `extraction_json` on every download, so
        it can disagree with what the API reports. Each document's closing
        balance must be in both."""
        ws = _download(client, auth, templated_job["id"]).worksheets[0]
        in_sheet = _numbers(_sheet_text(ws))
        for d in _results_for(client, auth, templated_job["id"], STATEMENTS):
            stem = d["filename"][:-4]
            closing = round(float(SIBLINGS[stem]["closing"].replace(",", "")), 2)
            assert closing in _numbers(_all_text(d["extracted_data"] or {})), (
                f"{stem}: closing balance missing from the API result")
            assert closing in in_sheet, (
                f"{stem}: closing balance reached the API but not the download")


# ── 2. mixed batch, no template ──────────────────────────────────────────────

@pytest.fixture(scope="module")
def mixed_job(client, auth):
    up = _upload(client, auth, MIXED, template_id=None)
    job = _await_job(client, auth, up["job_id"])
    return {"upload": up, "job": job, "id": up["job_id"]}


@pytest.fixture(scope="module")
def exclusive_numbers():
    """Per document, the numbers that appear in ITS pdf and in no other.

    A shared number proves nothing either way — five financial documents from
    one fictional company repeat plenty of them. Only a number unique to one
    source can show that a value crossed from one sheet to another.
    """
    own = {s: _numbers(_own_text(s)) for s in MIXED}
    excl = {s: own[s] - set().union(*(own[o] for o in MIXED if o != s))
            for s in MIXED}
    # Without this the crossing test below is vacuous: a document with no
    # unique number can neither be located nor shown to have leaked.
    empty = [s for s, v in excl.items() if not v]
    assert not empty, (
        f"these documents share every number with a sibling, so nothing in "
        f"this fixture can distinguish them: {empty}")
    return excl


class TestMixedBatchWithNoTemplate:

    def test_the_job_completes(self, mixed_job):
        assert mixed_job["job"]["status"] == "completed", mixed_job["job"]

    def test_every_document_comes_back(self, mixed_job, client, auth):
        docs = _results(client, auth, mixed_job["id"])
        assert sorted(d["filename"] for d in docs) == sorted(
            f"{s}.pdf" for s in MIXED)

    def test_each_document_inferred_its_own_shape(self, mixed_job, client, auth):
        sigs = {d["filename"]: (d["extracted_data"] or {}).get("shape_signature")
                for d in _results_for(client, auth, mixed_job["id"], MIXED)}
        assert all(sigs.values()), f"a document inferred no shape: {sigs}"
        assert len(set(sigs.values())) == len(MIXED), (
            f"five different kinds of document produced "
            f"{len(set(sigs.values()))} distinct shapes: {sigs}")

    def test_the_download_has_one_sheet_per_inferred_shape(
            self, mixed_job, client, auth):
        docs = _results_for(client, auth, mixed_job["id"], MIXED)
        sigs = {(d["extracted_data"] or {}).get("shape_signature") for d in docs}
        wb = _download(client, auth, mixed_job["id"])
        assert len(wb.worksheets) > 1, (
            "a mixed batch collapsed into a single sheet — the documents "
            "share headings they do not have")
        assert len(wb.worksheets) == len(sigs), (
            f"{len(sigs)} inferred shapes produced {len(wb.worksheets)} "
            f"sheets: {[w.title for w in wb.worksheets]}")

    def test_no_value_crosses_between_sheets(self, mixed_job, client, auth,
                                             exclusive_numbers):
        """Every document claims exactly one sheet, and every sheet is claimed
        by exactly one document.

        Stated as a bijection on purpose. "No sheet holds a stray value" alone
        is satisfied by a workbook that dropped four of the five documents, so
        the mapping has to be checked in both directions at once.
        """
        wb = _download(client, auth, mixed_job["id"])
        per_sheet = {w.title: _numbers(_sheet_text(w)) for w in wb.worksheets}

        claims = {t: sorted(s for s in MIXED if exclusive_numbers[s] & nums)
                  for t, nums in per_sheet.items()}

        crossed = {t: c for t, c in claims.items() if len(c) > 1}
        assert not crossed, (
            f"CONTAMINATION between sheets — these sheets hold values unique "
            f"to more than one source document: {crossed}")

        claimed_by = {}
        for t, c in claims.items():
            for s in c:
                claimed_by.setdefault(s, []).append(t)

        unplaced = [s for s in MIXED if s not in claimed_by]
        assert not unplaced, (
            f"no value unique to these documents appears in any sheet — they "
            f"are missing from the workbook, not merely uncontaminated: "
            f"{unplaced}")

        spread = {s: t for s, t in claimed_by.items() if len(t) > 1}
        assert not spread, (
            f"CONTAMINATION between sheets — one document's unique values "
            f"appear in several sheets: {spread}")

    def test_every_document_is_identifiable_in_the_download(
            self, mixed_job, client, auth, exclusive_numbers):
        """The converse of the crossing test: a workbook that dropped a
        document entirely would pass every "no contamination" check."""
        wb = _download(client, auth, mixed_job["id"])
        all_nums = set().union(*(_numbers(_sheet_text(w)) for w in wb.worksheets))
        missing = [s for s in MIXED
                   if exclusive_numbers[s] and not (exclusive_numbers[s] & all_nums)]
        assert not missing, (
            f"no value unique to these documents reached the download: {missing}")


# ── 3. job and per-document accounting ───────────────────────────────────────

class TestJobAccounting:

    @pytest.mark.parametrize("fixture_name,stems",
                             [("templated_job", STATEMENTS), ("mixed_job", MIXED)])
    def test_five_in_five_out(self, request, fixture_name, stems, client, auth):
        job = request.getfixturevalue(fixture_name)
        j, docs = job["job"], _results(client, auth, job["id"])
        assert job["upload"]["total_files"] == 5
        assert j["total_docs"] == 5, j
        assert j["successful"] == 5, (
            f"{j['successful']} of 5 documents succeeded, {j['failed']} failed. "
            f"Per-document errors: "
            f"{[(d['filename'], d['validation_errors']) for d in docs]}")
        assert j["failed"] == 0, j
        assert len(docs) == 5, [d["filename"] for d in docs]

    @pytest.mark.parametrize("fixture_name,stems",
                             [("templated_job", STATEMENTS), ("mixed_job", MIXED)])
    def test_no_document_is_recorded_twice(self, request, fixture_name, stems,
                                           client, auth):
        job = request.getfixturevalue(fixture_name)
        names = [d["filename"] for d in _results_for(client, auth, job["id"], stems)]
        assert len(names) == len(set(names)), names

    @pytest.mark.parametrize("fixture_name", ["templated_job", "mixed_job"])
    def test_the_job_is_finished_and_timed(self, request, fixture_name):
        j = request.getfixturevalue(fixture_name)["job"]
        assert j["status"] == "completed"
        assert j["completed_at"], "a completed job with no completion time"
        assert j["error_message"] in (None, ""), j["error_message"]

    def test_nothing_is_left_at_processing(self, client, auth,
                                           templated_job, mixed_job):
        r = client.get("/api/jobs", headers=auth["acme"])
        assert r.status_code == 200, r.text
        stranded = [j["id"] for j in r.json()
                    if j["status"] in ("pending", "processing")]
        assert not stranded, f"jobs stranded mid-flight: {stranded}"

    @pytest.mark.parametrize("fixture_name", ["templated_job", "mixed_job"])
    def test_the_job_list_agrees_with_the_job_itself(self, request,
                                                     fixture_name, client, auth):
        job = request.getfixturevalue(fixture_name)
        r = client.get("/api/jobs", headers=auth["acme"])
        row = next(j for j in r.json() if j["id"] == job["id"])
        for k in ("status", "total_docs", "successful", "failed", "needs_review"):
            assert row[k] == job["job"][k], (k, row[k], job["job"][k])


# ── 4. what a failure looks like from outside ────────────────────────────────

GARBAGE = b"%PDF-1.4\nthis is not a pdf\n"


def _upload_raw(client, auth, parts, template_id=None):
    """Upload arbitrary (filename, bytes) parts — including files that cannot
    possibly extract."""
    files = [("files", (name, data, "application/pdf")) for name, data in parts]
    payload = {"client_id": "acme_001"}
    if template_id is not None:
        payload["template_id"] = str(template_id)
    r = client.post("/api/extract/upload", files=files, data=payload,
                    headers=auth["acme"])
    assert r.status_code == 202, r.text
    return r.json()


class TestAFailedDocumentIsVisible:
    """A document that fails must leave a row, and a job that produced nothing
    must not report success.

    Both are what made the `storage` NameError invisible: five documents
    failed, no row was written for any of them, and the job said `completed`.
    The status now names what the batch produced, and every failure carries its
    own message.
    """

    def test_a_batch_where_everything_fails_reports_failed(self, client, auth,
                                                           bank_template):
        up = _upload_raw(client, auth,
                         [(f"broken-{i}.pdf", GARBAGE) for i in range(3)],
                         template_id=bank_template.id)
        job = _await_job(client, auth, up["job_id"])
        assert job["status"] == "failed", (
            f"three unreadable documents produced status {job['status']!r} — "
            f"a batch that extracted nothing must not report success")
        assert job["successful"] == 0 and job["failed"] == 3, job
        assert job["error_message"], "a failed job with no explanation"

    def test_the_jobs_message_carries_the_documents_own_reason(
            self, client, auth, bank_template):
        """Not a count of failures — the reason for them.

        The message used to read "N of M document(s) failed — open the job to
        see each document's error", and there is no per-document error view to
        open. The reason was already on `DocumentResult.validation_errors` and
        already returned by the API; a grep across the frontend found exactly
        one reader, the TypeScript type declaration. So the extract page fell
        back to four hardcoded guesses about password-protected PDFs.
        """
        up = _upload_raw(client, auth,
                         [(f"broken-{i}.pdf", GARBAGE) for i in range(3)],
                         template_id=bank_template.id)
        job = _await_job(client, auth, up["job_id"])
        msg = job["error_message"] or ""

        docs = _results(client, auth, up["job_id"])
        reasons = {(d["validation_errors"] or "").strip()
                   for d in docs if (d["validation_errors"] or "").strip()}
        assert reasons, "the documents recorded no reason to surface"
        assert any(r in msg for r in reasons), (
            f"the job says {msg!r} but the documents say {reasons!r}")

    def test_identical_failures_are_stated_once(self, client, auth,
                                                bank_template):
        """Three documents failing the same way say it once, with a count."""
        up = _upload_raw(client, auth,
                         [(f"broken-{i}.pdf", GARBAGE) for i in range(3)],
                         template_id=bank_template.id)
        job = _await_job(client, auth, up["job_id"])
        msg = job["error_message"] or ""
        docs = _results(client, auth, up["job_id"])
        reasons = [(d["validation_errors"] or "").strip() for d in docs]
        if len(set(r for r in reasons if r)) == 1 and len(reasons) == 3:
            assert msg.count(reasons[0]) == 1, (
                f"the same reason is repeated in {msg!r}")

    def test_every_failed_document_carries_its_own_error(self, client, auth,
                                                         bank_template):
        up = _upload_raw(client, auth,
                         [(f"broken-{i}.pdf", GARBAGE) for i in range(3)],
                         template_id=bank_template.id)
        _await_job(client, auth, up["job_id"])
        docs = _results(client, auth, up["job_id"])
        assert len(docs) == 3, (
            f"3 documents failed but {len(docs)} rows were written — a failure "
            f"with no record is a failure nobody can diagnose")
        for d in docs:
            assert d["validation_errors"], (
                f"{d['filename']} failed with an empty error message")

    def test_a_batch_where_some_fail_reports_partial(self, client, auth,
                                                     bank_template):
        good = "STMT-2024-01"
        up = _upload_raw(
            client, auth,
            [(f"{good}.pdf", (bs.PDF_DIR / f"{good}.pdf").read_bytes()),
             ("broken.pdf", GARBAGE)],
            template_id=bank_template.id)
        job = _await_job(client, auth, up["job_id"])
        assert job["status"] == "partial", (
            f"one success and one failure produced status {job['status']!r}")
        assert job["successful"] == 1 and job["failed"] == 1, job

        docs = _results(client, auth, up["job_id"])
        assert len(docs) == 2, [d["filename"] for d in docs]
        by_name = {d["filename"]: d for d in docs}
        assert by_name[f"{good}.pdf"]["extracted_data"], (
            "the document that succeeded came back empty")
        assert by_name["broken.pdf"]["validation_errors"], (
            "the document that failed came back with no error")

    def test_a_partial_job_still_exports_what_worked(self, client, auth,
                                                     bank_template):
        """A partial job has real output for the documents that succeeded. The
        status says not everything worked; the download must still contain the
        part that did."""
        good = "STMT-2024-01"
        up = _upload_raw(
            client, auth,
            [(f"{good}.pdf", (bs.PDF_DIR / f"{good}.pdf").read_bytes()),
             ("broken.pdf", GARBAGE)],
            template_id=bank_template.id)
        _await_job(client, auth, up["job_id"])
        text = _sheet_text(_download(client, auth, up["job_id"]).worksheets[0])
        assert SIBLINGS[good]["stmt"] in text

    def test_an_exception_inside_the_worker_still_writes_a_row(
            self, client, auth, bank_template, monkeypatch):
        """The exact shape of the Phase 8 bug: something raises before any
        result object exists. The document must still be recorded, with the
        exception in its error message."""
        from app.api.routes import extract as ex

        def boom(storage, ref):
            raise RuntimeError("resolver exploded")

        monkeypatch.setattr(ex, "_resolve_source", boom)
        good = "STMT-2024-01"
        up = _upload_raw(
            client, auth,
            [(f"{good}.pdf", (bs.PDF_DIR / f"{good}.pdf").read_bytes())],
            template_id=bank_template.id)
        job = _await_job(client, auth, up["job_id"])

        assert job["status"] == "failed", job
        docs = _results(client, auth, up["job_id"])
        assert len(docs) == 1, (
            "the worker raised before producing a result and wrote no row — "
            "this is exactly how the storage NameError stayed invisible")
        assert "resolver exploded" in docs[0]["validation_errors"]
        assert docs[0]["filename"] == f"{good}.pdf"


# ── 5. the shape of a flag ───────────────────────────────────────────────────

class TestFlaggedFieldsHaveOneShape:
    """`validation.flagged_fields` had three shapes and two consumers that both
    assumed the third.

    The job runner summarised each entry with `f['ref']`. On the slot path the
    entries were plain strings, so that raised TypeError, the per-document
    handler counted the document as failed, and nothing was saved — every
    document carrying a single flagged field. It surfaced only once the
    `storage` NameError stopped hiding it, which is the argument for pinning
    the shape rather than the symptom.
    """

    def test_the_engine_emits_dicts_with_ref_value_reason(self, templated_job,
                                                          client, auth):
        seen = 0
        for d in _results_for(client, auth, templated_job["id"], STATEMENTS):
            for f in ((d["extracted_data"] or {})
                      .get("validation", {}).get("flagged_fields", [])):
                seen += 1
                assert isinstance(f, dict), (
                    f"{d['filename']}: flagged entry is {type(f).__name__}, "
                    f"not a dict: {f!r}")
                assert set(f) >= {"ref", "value", "reason"}, f
        assert seen, (
            "no document in this batch flagged anything, so this test proved "
            "nothing — pick a batch that does")

    def test_a_flagged_document_still_saves_with_its_warning(self, templated_job,
                                                             client, auth):
        """The regression itself: a document with flags must reach the database
        WITH its warning text, not be counted as a failure."""
        flagged_docs = [
            d for d in _results_for(client, auth, templated_job["id"], STATEMENTS)
            if (d["extracted_data"] or {}).get("validation", {}).get("flagged_count")
        ]
        assert flagged_docs, "no flagged document in this batch to check"
        for d in flagged_docs:
            assert d["validation_warnings"], (
                f"{d['filename']} has flagged fields but an empty "
                f"validation_warnings — the summary line was lost")

    def test_the_summary_renderer_survives_every_shape(self):
        """`_flag_summary` is on the save path. It must never be the reason a
        document fails, whatever an entry turns out to look like."""
        from app.api.routes.extract import _flag_summary

        assert _flag_summary([]) == ""
        assert "B7" in _flag_summary([{"ref": "B7", "value": "12", "reason": "x"}])
        assert "legacy string" in _flag_summary(["legacy string"])
        # the image path's older key, and an entry missing everything
        assert "no text layer" in _flag_summary(
            [{"ref": "image_upload", "value": "", "issue": "no text layer"}])
        assert isinstance(_flag_summary([{}, None, 3]), str)

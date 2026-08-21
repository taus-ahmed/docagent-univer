"""
Every parameter ExportRequest accepts must change the file it produces.

The pre-Phase-8 audit found five fields declared on ExportRequest and read by
nothing: selected_columns, column_order, doc_types, include_line_items,
include_needs_review_only. A caller asking for three columns of one document
type silently received every column of every document — and this is a PUBLIC
contract, visible in Swagger, so an integrator would have no way to tell.

Four are now implemented and pinned here. include_line_items was removed
rather than implemented: this export is a flat table of scalar fields and emits
no line items in any configuration, so the flag could only ever describe
something the writer cannot produce.
"""
import io

import pytest
from openpyxl import load_workbook

from tests.harness import bootstrap as bs

bs.bootstrap()


@pytest.fixture(scope="module")
def job(db, users):
    """One job, three documents of two types, one flagged for review."""
    import json

    from app.models.models import DocumentResult, ExtractionJob
    j = ExtractionJob(user_id=users["acme"].id, client_id="acme_001",
                      status="completed", total_docs=3, successful=3,
                      input_source="upload")
    db.add(j)
    db.commit()
    db.refresh(j)

    rows = [
        ("inv1.pdf", "sales_invoice", False,
         {"Invoice Number": "INV-1", "Total": "100", "Vendor": "Acme"}),
        ("inv2.pdf", "sales_invoice", True,
         {"Invoice Number": "INV-2", "Total": "200", "Vendor": "Beta"}),
        ("po1.pdf", "purchase_order", False,
         {"Invoice Number": "PO-1", "Total": "300", "Vendor": "Gamma"}),
    ]
    for fn, dt, review, fields in rows:
        d = DocumentResult(
            job_id=j.id, filename=fn, document_type=dt,
            overall_confidence="high", needs_review=review,
            extraction_json=json.dumps({
                "extracted_data": {k: {"value": v, "confidence": "high"}
                                   for k, v in fields.items()}}))
        db.add(d)
    db.commit()
    return j


def _sheet(client, auth, **params):
    r = client.post("/api/export/combined", headers=auth["acme"],
                    json={"job_id": params.pop("job_id"), **params})
    assert r.status_code == 200, r.text
    return load_workbook(io.BytesIO(r.content)).active


def _headers(ws):
    return [c.value for c in ws[1]]


def _files(ws):
    return [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]


class TestSelectedColumns:
    def test_omitting_it_returns_every_column(self, client, auth, job):
        h = _headers(_sheet(client, auth, job_id=job.id))
        assert {"Invoice Number", "Total", "Vendor"} <= set(h)

    def test_naming_two_columns_returns_two(self, client, auth, job):
        h = _headers(_sheet(client, auth, job_id=job.id,
                            selected_columns=["Invoice Number", "Total"]))
        assert "Vendor" not in h
        assert {"Invoice Number", "Total"} <= set(h)

    def test_the_fixed_columns_always_survive(self, client, auth, job):
        """File/Doc Type/Confidence/Status identify the row; filtering data
        columns must not strip the row's identity."""
        h = _headers(_sheet(client, auth, job_id=job.id,
                            selected_columns=["Total"]))
        assert h[:4] == ["File", "Doc Type", "Confidence", "Status"]

    def test_an_unknown_column_name_is_ignored_not_an_error(
            self, client, auth, job):
        """A saved column set should survive a template losing a field."""
        h = _headers(_sheet(client, auth, job_id=job.id,
                            selected_columns=["Total", "NoSuchField"]))
        assert "Total" in h

    def test_matching_ignores_case_and_spacing(self, client, auth, job):
        h = _headers(_sheet(client, auth, job_id=job.id,
                            selected_columns=["  invoice   number "]))
        assert "Invoice Number" in h
        assert "Total" not in h

    def test_an_empty_list_means_no_data_columns(self, client, auth, job):
        """Distinct from omitting it. [] is an explicit choice."""
        h = _headers(_sheet(client, auth, job_id=job.id, selected_columns=[]))
        assert h == ["File", "Doc Type", "Confidence", "Status"]


class TestColumnOrder:
    def test_named_columns_lead_in_the_order_given(self, client, auth, job):
        h = _headers(_sheet(client, auth, job_id=job.id,
                            column_order=["Vendor", "Invoice Number"]))
        data = h[4:]
        assert data[:2] == ["Vendor", "Invoice Number"]

    def test_unnamed_columns_follow_in_their_natural_order(
            self, client, auth, job):
        h = _headers(_sheet(client, auth, job_id=job.id,
                            column_order=["Vendor"]))
        data = h[4:]
        assert data[0] == "Vendor"
        assert data[1:] == ["Invoice Number", "Total"]

    def test_it_combines_with_selected_columns(self, client, auth, job):
        h = _headers(_sheet(client, auth, job_id=job.id,
                            selected_columns=["Total", "Vendor"],
                            column_order=["Vendor"]))
        assert h[4:] == ["Vendor", "Total"]


class TestDocTypes:
    def test_omitting_it_returns_every_document(self, client, auth, job):
        assert len(_files(_sheet(client, auth, job_id=job.id))) == 3

    def test_naming_one_type_returns_only_that_type(self, client, auth, job):
        files = _files(_sheet(client, auth, job_id=job.id,
                              doc_types=["purchase_order"]))
        assert files == ["po1.pdf"]

    def test_matching_ignores_case(self, client, auth, job):
        files = _files(_sheet(client, auth, job_id=job.id,
                              doc_types=["Purchase_Order"]))
        assert files == ["po1.pdf"]

    def test_a_type_that_matches_nothing_is_404_not_an_empty_sheet(
            self, client, auth, job):
        """An empty spreadsheet looks like a broken export. Say why instead."""
        r = client.post("/api/export/combined", headers=auth["acme"],
                        json={"job_id": job.id, "doc_types": ["cheque"]})
        assert r.status_code == 404
        assert "filter" in r.text.casefold()


class TestIncludeNeedsReviewOnly:
    def test_false_returns_everything(self, client, auth, job):
        assert len(_files(_sheet(client, auth, job_id=job.id,
                                 include_needs_review_only=False))) == 3

    def test_true_returns_only_flagged_documents(self, client, auth, job):
        files = _files(_sheet(client, auth, job_id=job.id,
                              include_needs_review_only=True))
        assert files == ["inv2.pdf"]

    def test_it_combines_with_doc_types(self, client, auth, job):
        files = _files(_sheet(client, auth, job_id=job.id,
                              doc_types=["sales_invoice"],
                              include_needs_review_only=True))
        assert files == ["inv2.pdf"]


class TestIncludeLineItemsIsGone:
    def test_the_field_is_no_longer_part_of_the_contract(self):
        from app.schemas.schemas import ExportRequest
        assert "include_line_items" not in ExportRequest.model_fields

    def test_sending_it_anyway_is_harmless(self, client, auth, job):
        """Old clients must not start failing — pydantic ignores extras."""
        r = client.post("/api/export/combined", headers=auth["acme"],
                        json={"job_id": job.id, "include_line_items": True})
        assert r.status_code == 200, r.text


class TestPerFileParameters:
    def test_doc_ids_selects_documents(self, client, auth, db, job):
        from app.models.models import DocumentResult
        ids = [d.id for d in db.query(DocumentResult).filter(
            DocumentResult.job_id == job.id).order_by(DocumentResult.id).all()]
        r = client.post("/api/export/perfile", headers=auth["acme"],
                        json={"job_id": job.id, "doc_ids": ids[:1]})
        assert r.status_code == 200, r.text
        assert len(load_workbook(io.BytesIO(r.content)).sheetnames) == 1

    def test_selected_columns_filters_each_sheet(self, client, auth, job):
        r = client.post("/api/export/perfile", headers=auth["acme"],
                        json={"job_id": job.id, "selected_columns": ["Total"]})
        assert r.status_code == 200, r.text
        ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
        labels = {ws.cell(row=i, column=1).value
                  for i in range(2, ws.max_row + 1)}
        assert "Total" in labels
        assert "Vendor" not in labels


class TestTheRouteTheFrontendCalls:
    def test_perfile_has_no_hyphen(self, client, auth, job):
        """api.ts posted to /api/export/per-file while the route is
        /api/export/perfile, so every per-file download 404'd."""
        assert client.post("/api/export/per-file", headers=auth["acme"],
                           json={"job_id": job.id}).status_code == 404
        assert client.post("/api/export/perfile", headers=auth["acme"],
                           json={"job_id": job.id}).status_code == 200
